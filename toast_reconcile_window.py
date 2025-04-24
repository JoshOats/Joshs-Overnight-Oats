import sys
import os
from PyQt5.QtWidgets import (QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QFileDialog, QTextEdit, QMessageBox, QListWidget, QApplication, QWidget, QDialog, QScrollArea)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from collections import defaultdict
from datetime import datetime
import csv
import logging
from pathlib import Path
from retro_style import RetroWindow, create_retro_central_widget
from PyQt5.QtGui import QIcon, QPixmap
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import traceback  # Added for better exception tracking

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Configure logging to be minimal so it doesn't interfere with our custom logging
logging.basicConfig(level=logging.WARNING)

# Location dictionary mapping
LOCATION_DICT = {
        "Aventura (Miami Gardens)": "Carrot Aventura Love LLC (Aventura)",
        "Aventura Mall": "Carrot Love Aventura Mall Operating LLC",
        "Boca Palmetto Park": "Carrot Love Palmetto Park Operating LLC",
        "Brickell": "Carrot Love Brickell Operating LLC",
        "Bryant Park": "Carrot Love Bryant Park Operating LLC",
        "Coconut Creek": "Carrot Love Coconut Creek Operating LLC",
        "Coconut Grove": "Carrot Love Coconut Grove Operating LLC",
        "Coral Gables": "Carrot Coral GablesLove LLC (Coral Gabes)",
        "Dadeland": "Carrot Love Dadeland Operating LLC",
        "Doral": "Carrot Love City Place Doral Operating LLC",
        "Downtown": "Carrot Downtown Love Two LLC",
        "Flatiron": "Carrot Flatiron Love Manhattan Operating LLC",
        "Hollywood": "Carrot Love Hollywood Operating LLC",
        "Las Olas": "Carrot Love Las Olas Operating LLC",
        "Lexington": "Carrot Love 600 Lexington LLC",
        "Miami Shores": "Carrot Express Miami Shores LLC",
        "Midtown": "Carrot Express Midtown LLC",
        "North Beach": "Carrot North Beach Love LL (North Beach)",
        "Pembroke Pines": "Carrot Love Pembroke Pines Operating LLC",
        "Plantation": "Carrot Love Plantation Operating LLC",
        "River Landing": "Carrot Love River Lading Operating LLC",
        "South Miami (Sunset)": "Carrot Love Sunset Operating LLC",
        "West Boca": "Carrot Love West Boca Operating LLC",
        "South Beach": "Carrot Sobe Love South Florida Operating C LLC"
    }

def read_csv_with_encoding(file_path, encodings=['utf-8', 'latin-1', 'cp1252']):
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                return list(reader)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Unable to decode the file {file_path} with any of the provided encodings.")

def date_range_overlaps(range1, range2):
    start1, end1 = range1
    start2, end2 = range2
    return start1 <= end2 and start2 <= end1

def replace_existing_file(output_dir, new_filename, file_prefix):
    new_date_range = new_filename.split('_')[-1].split('.')[0]
    new_start, new_end = new_date_range.split('-')
    new_range = (datetime.strptime(new_start, '%m%d%Y'), datetime.strptime(new_end, '%m%d%Y'))
    for filename in os.listdir(output_dir):
        if filename.startswith(file_prefix) and (filename.endswith('.csv') or filename.endswith('.xlsx')):
            existing_date_range = filename.split('_')[-1].split('.')[0]
            existing_start, existing_end = existing_date_range.split('-')
            existing_range = (datetime.strptime(existing_start, '%m%d%Y'), datetime.strptime(existing_end, '%m%d%Y'))
            if date_range_overlaps(new_range, existing_range):
                old_file_path = os.path.join(output_dir, filename)
                try:
                    os.remove(old_file_path)
                    return True
                except Exception as e:
                    return False
    return True

def save_excel_report(sales_data, gl_data, export_data, gift_card_outstanding, group_data, earliest_date, latest_date, output_dir, log_callback=None):
    """Save Excel report with detailed logging via callback function"""
    def log(message):
        """Log to callback if provided, otherwise just print"""
        if log_callback:
            log_callback(message)
        print(message)

    try:
        log("Starting Excel report creation...")
        output_filename = f"Net_Sales_Report_{earliest_date.strftime('%m%d%Y')}-{latest_date.strftime('%m%d%Y')}.xlsx"

        log(f"Checking for existing files to replace: {output_filename}")
        if not replace_existing_file(output_dir, output_filename, "Net_Sales"):
            log("Failed to replace existing Net Sales Report file. Aborting save operation.")
            return None

        output_path = os.path.join(output_dir, output_filename)
        log(f"Output path will be: {output_path}")

        # Create a new Excel workbook
        log("Creating new Excel workbook...")
        workbook = openpyxl.Workbook()

        # Set core workbook properties
        log("Setting workbook properties...")
        workbook.properties.creator = "Toast Net Sales Reconcile"
        workbook.properties.title = f"Net Sales Report {earliest_date.strftime('%m%d%Y')}-{latest_date.strftime('%m%d%Y')}"

        # Initialize properties for the active sheet
        log("Initializing active sheet properties...")
        active_sheet = workbook.active
        active_sheet.sheet_properties.filterMode = False
        if hasattr(active_sheet, 'page_setup'):
            active_sheet.page_setup.fitToHeight = 0
            active_sheet.page_setup.fitToWidth = 1

        # Summary Net Sales tab
        log("Setting up Summary Net Sales tab...")
        summary_sheet = workbook.active
        summary_sheet.title = "Summary Net Sales"

        # Set up headers for Summary Net Sales tab
        log("Adding headers to Summary sheet...")
        summary_sheet['A1'] = "Date Range"
        summary_sheet['B1'] = "Location"
        summary_sheet['C1'] = "Toast Net Sales"
        summary_sheet['D1'] = "R365 Net Sales"
        summary_sheet['E1'] = "Difference"

        # Apply header formatting
        log("Applying header formatting...")
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)

        for cell in summary_sheet['1:1']:
            cell.fill = header_fill
            cell.font = header_font

        # Extract plantation payable total if available
        plantation_payable_total = 0
        if "plantation_payable_total" in gl_data:
            plantation_payable_total = gl_data["plantation_payable_total"]
            log(f"Found plantation payable total: {plantation_payable_total}")
        else:
            log("No plantation payable total found")

        # Calculate total R365 sales by location for the entire period
        log("Calculating R365 totals...")
        r365_totals = defaultdict(float)
        for date in sorted(export_data.keys()):
            for location, amount in export_data[date].items():
                r365_totals[location] += amount
        log(f"Calculated R365 totals for {len(r365_totals)} locations")

        # Populate Summary Net Sales tab
        log("Populating Summary Net Sales tab...")
        date_range = f"{earliest_date.strftime('%m/%d/%Y')} - {latest_date.strftime('%m/%d/%Y')}"
        row = 2

        for location, toast_sales in group_data.items():
            # Skip excluded locations
            if location in ["Pinecrest", "West Kendall (London Square)", "Weston"]:
                continue

            # Get the corresponding R365 location name
            r365_location = LOCATION_DICT.get(location, location)
            r365_sales = r365_totals.get(r365_location, 0)

            # Apply special handling for Plantation location
            adjusted_toast_sales = toast_sales
            if location == "Plantation" and plantation_payable_total > 0:
                adjusted_toast_sales = toast_sales - plantation_payable_total
                log(f"Adjusted Plantation sales: {toast_sales} -> {adjusted_toast_sales}")

            difference = adjusted_toast_sales - r365_sales

            summary_sheet[f'A{row}'] = date_range
            summary_sheet[f'B{row}'] = location
            summary_sheet[f'C{row}'] = adjusted_toast_sales
            summary_sheet[f'D{row}'] = r365_sales
            summary_sheet[f'E{row}'] = difference

            # Format currency cells
            for col in ['C', 'D', 'E']:
                summary_sheet[f'{col}{row}'].number_format = '$#,##0.00'

            row += 1

        log(f"Summary sheet complete with {row-2} locations")

        # Create Discrepancies tab
        log("Creating Discrepancies tab...")
        discrepancies_sheet = workbook.create_sheet(title="Discrepancies")
        discrepancies_sheet.sheet_properties.filterMode = False
        if hasattr(discrepancies_sheet, 'page_setup'):
            discrepancies_sheet.page_setup.fitToHeight = 0
            discrepancies_sheet.page_setup.fitToWidth = 1

        # Set up headers for Discrepancies tab
        log("Adding headers to Discrepancies sheet...")
        discrepancies_sheet['A1'] = "Date"
        discrepancies_sheet['B1'] = "Location"
        discrepancies_sheet['C1'] = "Adjusted Net Sales"
        discrepancies_sheet['D1'] = "NetSales"
        discrepancies_sheet['E1'] = "GL Credit"
        discrepancies_sheet['F1'] = "Export NetSales"
        discrepancies_sheet['G1'] = "Difference"
        discrepancies_sheet['H1'] = "Gift Card Outstanding"  # Added column for gift card note

        # Apply header formatting
        for cell in discrepancies_sheet['1:1']:
            cell.fill = header_fill
            cell.font = header_font

        # Populate Discrepancies tab for locations with non-zero differences in summary
        log("Finding locations with differences...")
        locations_with_differences = []
        for row_num in range(2, summary_sheet.max_row + 1):
            # Only include locations with significant differences (greater than 0.01)
            if summary_sheet[f'E{row_num}'].value is not None and abs(summary_sheet[f'E{row_num}'].value) > 0.01:
                locations_with_differences.append(summary_sheet[f'B{row_num}'].value)

        log(f"Found {len(locations_with_differences)} locations with differences")

        row = 2
        discrepancy_count = 0
        for date in sorted(sales_data.keys()):
            for location, net_sales in sales_data[date].items():
                # Skip excluded locations
                if location in ["Pinecrest", "West Kendall (London Square)", "Weston"]:
                    continue

                if location in locations_with_differences:
                    gl_location = LOCATION_DICT.get(location, location)
                    gl_credit = gl_data[date][gl_location]
                    adjusted_net_sales = net_sales - gl_credit
                    export_net_sales = export_data[date][gl_location]
                    difference = adjusted_net_sales - export_net_sales

                    # Add gift card note if applicable
                    gift_card_note = ""
                    if gift_card_outstanding[date][gl_location] > 0:
                        gift_card_note = "Possible Deferred Gift Card Discount"

                    if abs(difference) > 0.01:  # Only include rows with non-zero differences
                        try:
                            date_obj = datetime.strptime(date, '%Y-%m-%d') if isinstance(date, str) else date
                            discrepancies_sheet[f'A{row}'] = date_obj.strftime('%Y-%m-%d')
                        except:
                            discrepancies_sheet[f'A{row}'] = str(date)
                        discrepancies_sheet[f'B{row}'] = location
                        discrepancies_sheet[f'C{row}'] = adjusted_net_sales
                        discrepancies_sheet[f'D{row}'] = net_sales
                        discrepancies_sheet[f'E{row}'] = gl_credit
                        discrepancies_sheet[f'F{row}'] = export_net_sales
                        discrepancies_sheet[f'G{row}'] = difference
                        discrepancies_sheet[f'H{row}'] = gift_card_note  # Add gift card note

                        # Format currency cells
                        for col in ['C', 'D', 'E', 'F', 'G']:
                            discrepancies_sheet[f'{col}{row}'].number_format = '$#,##0.00'

                        row += 1
                        discrepancy_count += 1

        log(f"Added {discrepancy_count} rows to Discrepancies sheet")

        # Create Gift Card tab
        log("Creating Gift Card tab...")
        gift_card_sheet = workbook.create_sheet(title="Deferred Gift Card Discount")
        gift_card_sheet.sheet_properties.filterMode = False
        if hasattr(gift_card_sheet, 'page_setup'):
            gift_card_sheet.page_setup.fitToHeight = 0
            gift_card_sheet.page_setup.fitToWidth = 1

        # Set up headers for Gift Card tab
        log("Adding headers to Gift Card sheet...")
        gift_card_sheet['A1'] = "Date"  # Just set header text directly
        gift_card_sheet['B1'] = "Location"
        gift_card_sheet['C1'] = "Adjusted Net Sales"
        gift_card_sheet['D1'] = "NetSales"
        gift_card_sheet['E1'] = "GL Credit"
        gift_card_sheet['F1'] = "Export NetSales"
        gift_card_sheet['G1'] = "Difference"
        gift_card_sheet['H1'] = "Gift Card Outstanding"

        # Apply header formatting
        for cell in gift_card_sheet['1:1']:
            cell.fill = header_fill
            cell.font = header_font

        # Populate Gift Card tab
        log("Populating Gift Card tab...")
        row = 2
        gift_card_count = 0
        for date in sorted(sales_data.keys()):
            for location, net_sales in sales_data[date].items():
                # Skip excluded locations
                if location in ["Pinecrest", "West Kendall (London Square)", "Weston"]:
                    continue

                gl_location = LOCATION_DICT.get(location, location)
                gift_card_amount = gift_card_outstanding[date][gl_location]

                if gift_card_amount > 0:  # Only include rows with gift card amounts
                    gl_credit = gl_data[date][gl_location]
                    adjusted_net_sales = net_sales - gl_credit
                    export_net_sales = export_data[date][gl_location]
                    difference = adjusted_net_sales - export_net_sales

                    try:
                        date_obj = datetime.strptime(date, '%Y-%m-%d') if isinstance(date, str) else date
                        gift_card_sheet[f'A{row}'] = date_obj.strftime('%Y-%m-%d')
                    except:
                        gift_card_sheet[f'A{row}'] = str(date)
                    gift_card_sheet[f'B{row}'] = location
                    gift_card_sheet[f'C{row}'] = adjusted_net_sales
                    gift_card_sheet[f'D{row}'] = net_sales
                    gift_card_sheet[f'E{row}'] = gl_credit
                    gift_card_sheet[f'F{row}'] = export_net_sales
                    gift_card_sheet[f'G{row}'] = difference
                    gift_card_sheet[f'H{row}'] = "Possible Deferred Gift Card Discount"

                    # Format currency cells
                    for col in ['C', 'D', 'E', 'F', 'G']:
                        gift_card_sheet[f'{col}{row}'].number_format = '$#,##0.00'

                    row += 1
                    gift_card_count += 1

        log(f"Added {gift_card_count} rows to Gift Card sheet")

        # Auto-adjust column width for all sheets
        log("Auto-adjusting column widths...")
        for sheet_name in workbook.sheetnames:
            log(f"Adjusting columns for sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            column_widths = {}

            # First pass: calculate max lengths
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        col_letter = cell.column_letter
                        try:
                            cell_value = str(cell.value)
                            current_width = column_widths.get(col_letter, 0)
                            column_widths[col_letter] = max(current_width, min(len(cell_value) + 2, 50))
                        except (TypeError, ValueError):
                            pass

            # Second pass: apply widths with reasonable limits
            for col_letter, width in column_widths.items():
                # Ensure width is between 8 and 50
                adjusted_width = max(min(width, 50), 8)
                sheet.column_dimensions[col_letter].width = adjusted_width

        # Set view properties for all sheets
        log("Setting view properties for all sheets...")
        for sheet in workbook.worksheets:
            if not hasattr(sheet, 'sheet_view'):
                sheet.sheet_view = openpyxl.worksheet.views.SheetView()
            sheet.sheet_view.showGridLines = True
            sheet.sheet_view.defaultGridColor = True

        # Save the workbook
        log("Attempting to save workbook...")
        try:
            # First attempt with default settings
            workbook.save(output_path)
            log(f"Excel report saved successfully to: {output_path}")
            return output_filename  # Return on successful save
        except Exception as e:
            log(f"First save attempt failed: {str(e)}")
            log("Trying alternative save method...")
            try:
                # Create a clean version by using a write-only workbook
                new_wb = openpyxl.Workbook(write_only=True)
                log("Created write-only workbook for alternative save method")

                # Copy each sheet
                for sheet_name in workbook.sheetnames:
                    log(f"Copying sheet {sheet_name} to new workbook")
                    new_sheet = new_wb.create_sheet(title=sheet_name)
                    old_sheet = workbook[sheet_name]
                    row_count = 0
                    for row in old_sheet.rows:
                        values = [cell.value for cell in row]
                        new_sheet.append(values)
                        row_count += 1
                    log(f"Copied {row_count} rows from {sheet_name}")

                # Save the clean version
                log("Saving with alternative method...")
                new_wb.save(output_path)
                log(f"Excel report saved using alternative method to: {output_path}")
                return output_filename
            except Exception as inner_e:
                log(f"Alternative save method failed: {str(inner_e)}")
                log(f"Full error: {traceback.format_exc()}")
                return None
    except Exception as e:
        log(f"Error in Excel report creation: {str(e)}")
        log(f"Full error details: {traceback.format_exc()}")
        return None


def debug_excel_file(file_path, log_callback=None):
    """Debug Excel file structure using openpyxl with logging callback"""
    def log(message):
        if log_callback:
            log_callback(message)
        print(message)

    log(f"Analyzing Excel file: {file_path}")
    try:
        # Try to open the file
        from openpyxl import load_workbook
        wb = load_workbook(file_path)

        # Log basic information
        log(f"Successfully opened file with openpyxl")
        log(f"Sheets: {wb.sheetnames}")
        log(f"Active sheet: {wb.active.title}")

        # Log workbook properties
        if hasattr(wb, 'properties'):
            props = wb.properties
            log(f"Workbook properties: Creator={props.creator}, Title={props.title}")

        # Check defined names
        if len(wb.defined_names) > 0:
            log(f"Defined names: {[name for name in wb.defined_names]}")

        # Check calculation properties
        if hasattr(wb, 'calculation'):
            log(f"Calculation properties exist")

        return True
    except Exception as e:
        log(f"Error analyzing Excel file: {str(e)}")
        log(f"Full error: {traceback.format_exc()}")
        return False


class ReconcileThread(QThread):
    update_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, input_files, output_dir):
        super().__init__()
        self.input_files = input_files
        self.output_dir = output_dir

    def log(self, message):
        """Log to UI console"""
        self.update_signal.emit(message)
        print(message)  # Also print to terminal for debugging

    def run(self):
        try:
            self.log("Starting reconciliation process")

            # Process sales data
            self.log("Processing sales data...")
            order_files = [f for f in self.input_files if os.path.basename(f).startswith("Order")]
            self.log(f"Found {len(order_files)} order files")
            sales_data, earliest_date, latest_date, valid_dates = self.process_sales_data(order_files)
            self.log(f"Processed sales data for date range: {earliest_date.strftime('%m/%d/%Y')} - {latest_date.strftime('%m/%d/%Y')}")

            # Process GL data
            self.log("Processing GL data...")
            gl_files = [f for f in self.input_files if os.path.basename(f).lower().startswith("gl")]
            self.log(f"Found {len(gl_files)} GL files")
            gl_data, gift_card_outstanding = self.process_gl_data(gl_files, valid_dates)
            self.log(f"Processed GL data with {len(gl_data)} entries")

            # Process export data
            self.log("Processing export data...")
            export_files = [f for f in self.input_files if os.path.basename(f).lower().startswith("export")]
            self.log(f"Found {len(export_files)} export files")
            export_data = self.process_export_data(export_files, valid_dates)
            self.log(f"Processed export data with {len(export_data)} entries")

            # Process group overview data
            self.log("Processing group overview data...")
            group_files = [f for f in self.input_files if os.path.basename(f).lower().startswith("group")]
            self.log(f"Found {len(group_files)} group files")
            group_data = self.process_group_data(group_files)
            self.log(f"Processed group data with {len(group_data)} locations")

            # Save Excel report
            self.log("Saving Excel report...")
            output_filename = save_excel_report(
                sales_data, gl_data, export_data, gift_card_outstanding,
                group_data, earliest_date, latest_date, self.output_dir,
                log_callback=self.log  # Pass the log function to the save function
            )

            if output_filename:
                self.log(f"Successfully saved report as: {output_filename}")
                self.finished_signal.emit(True, f"Reconciliation completed successfully for {earliest_date.strftime('%m/%d/%Y')}-{latest_date.strftime('%m/%d/%Y')}! \n \n \n Excel report saved as: {output_filename}")
            else:
                self.log("Failed to save the Excel report")
                self.finished_signal.emit(False, "Failed to save the Excel report.")
        except Exception as e:
            self.log(f"An error occurred: {str(e)}")
            self.log(f"Full error details: {traceback.format_exc()}")
            self.finished_signal.emit(False, f"An error occurred: {str(e)}")

    def process_sales_data(self, order_files):
        sales_by_date_location = defaultdict(lambda: defaultdict(float))
        earliest_date = None
        latest_date = None
        valid_dates = set()
        seen_orders = set()  # Track unique orders

        def format_order_number(order_num):
            # Convert scientific notation to full integer string
            try:
                float_num = float(order_num)
                int_num = int(float_num)
                return str(int_num)
            except (ValueError, TypeError):
                return str(order_num)

        def get_calendar_date(datetime_str):
            # Extract just the date portion from the datetime string
            return datetime_str.split()[0]

        total_processed = 0
        total_skipped = 0
        total_duplicates = 0

        self.log(f"Beginning to process {len(order_files)} order files")

        for file in order_files:
            self.log(f"Processing file: {os.path.basename(file)}")
            try:
                rows = read_csv_with_encoding(file)
                self.log(f"Read {len(rows)} rows from file")

                headers = rows[0]
                self.log(f"Headers: {headers}")

                try:
                    date_index = headers.index('Opened')
                    location_index = headers.index('Location')
                    amount_index = headers.index('Amount')
                    order_num_index = headers.index('Order #')
                except ValueError as e:
                    self.log(f"Error finding required columns: {str(e)}")
                    self.log(f"Available columns: {headers}")
                    continue

                processed_in_file = 0
                for i, row in enumerate(rows[1:], start=1):
                    if not row:
                        continue
                    try:
                        date_str = row[date_index].split()[0]
                        date = datetime.strptime(date_str, '%m/%d/%y')
                        location = row[location_index]
                        if location in ["Weston", "West Kendall", "Pinecrest", "West Kendall (London Square)"]:
                            total_skipped += 1
                            continue

                        # Format order number properly
                        order_num = format_order_number(row[order_num_index])
                        calendar_date = get_calendar_date(row[date_index])

                        # Create unique identifier that includes location and calendar date
                        order_key = f"{location}_{calendar_date}_{order_num}_{row[date_index]}"

                        # Check for duplicates
                        if order_key in seen_orders:
                            total_duplicates += 1
                            continue

                        seen_orders.add(order_key)
                        amount = float(row[amount_index])

                        date_key = date.strftime('%Y-%m-%d')
                        sales_by_date_location[date_key][location] += amount
                        valid_dates.add(date_key)

                        if earliest_date is None or date < earliest_date:
                            earliest_date = date
                        if latest_date is None or date > latest_date:
                            latest_date = date

                        processed_in_file += 1
                        total_processed += 1
                    except (ValueError, IndexError) as e:
                        self.log(f"Error processing row {i} in file {os.path.basename(file)}: {str(e)}")
                        continue

                self.log(f"Successfully processed {processed_in_file} rows from file: {os.path.basename(file)}")
            except Exception as e:
                self.log(f"Error processing file {os.path.basename(file)}: {str(e)}")
                self.log(traceback.format_exc())

        self.log(f"Sales data processing complete:")
        self.log(f"- Total orders processed: {total_processed}")
        self.log(f"- Total duplicates skipped: {total_duplicates}")
        self.log(f"- Total excluded locations skipped: {total_skipped}")
        self.log(f"- Date range: {earliest_date.strftime('%m/%d/%Y')} to {latest_date.strftime('%m/%d/%Y')}")
        self.log(f"- Number of unique dates: {len(valid_dates)}")
        self.log(f"- Number of unique locations: {sum(len(locations) for locations in sales_by_date_location.values())}")

        if not sales_by_date_location:
            raise ValueError("No valid data was processed from the Order files.")

        if earliest_date is None or latest_date is None:
            raise ValueError("No valid dates found in the processed data.")

        return sales_by_date_location, earliest_date, latest_date, valid_dates

    def process_gl_data(self, gl_files, valid_dates):
        gl_data = defaultdict(lambda: defaultdict(float))
        gift_card_outstanding = defaultdict(lambda: defaultdict(float))
        # Create a separate dictionary to track plantation payable
        plantation_payable_total = 0

        total_rows_processed = 0
        self.log(f"Beginning to process {len(gl_files)} GL files")

        for file in gl_files:
            self.log(f"Processing GL file: {os.path.basename(file)}")
            try:
                rows = read_csv_with_encoding(file)
                self.log(f"Read {len(rows)} rows from file")

                header_row_index = 3
                data_start_row = header_row_index + 1

                if len(rows) <= header_row_index:
                    self.log(f"File {os.path.basename(file)} has insufficient rows (expected header at row {header_row_index+1})")
                    continue

                headers = rows[header_row_index]
                self.log(f"Headers: {headers}")

                try:
                    date_index = headers.index('TrxDate')
                    location_index = headers.index('LocationName')
                    credit_index = headers.index('Credit')
                    debit_index = headers.index('Debit')
                    parent_account_index = headers.index('ParentAccountName')
                    trx_number_index = headers.index('TrxNumber')
                except ValueError as e:
                    self.log(f"Error finding required columns: {str(e)}")
                    self.log(f"Available columns: {headers}")
                    continue

                file_rows_processed = 0
                for i, row in enumerate(rows[data_start_row:], start=data_start_row):
                    if not row or len(row) <= max(date_index, location_index, credit_index, debit_index, parent_account_index, trx_number_index):
                        continue

                    try:
                        date_str = row[date_index].strip()
                        if not date_str:  # Skip rows without a date (summary rows)
                            continue

                        location = row[location_index]
                        credit = float(row[credit_index]) if row[credit_index] else 0
                        debit = float(row[debit_index]) if row[debit_index] else 0
                        parent_account = row[parent_account_index]
                        trx_number = row[trx_number_index]

                        # Skip excluded locations
                        if location in ["Pinecrest", "West Kendall (London Square)", "Weston"]:
                            continue

                        date = datetime.strptime(date_str, '%m/%d/%Y')
                        date_key = date.strftime('%Y-%m-%d')

                        if date_key in valid_dates:
                            if "21200 - Payable Donation" in parent_account:
                                gl_data[date_key][location] += credit
                            elif "25050 - Gift Cards Outstanding" in parent_account:
                                # Only include gift card transactions with "NJ" in the TrxNumber
                                if "NJ" in trx_number:
                                    gl_data[date_key][location] += credit
                                    if credit > 0:
                                        gift_card_outstanding[date_key][location] += credit
                            elif "73405 - Gift Cards" in parent_account:
                                # Only include gift card transactions with "NJ" in the TrxNumber
                                if "NJ" in trx_number:
                                    gl_data[date_key][location] -= debit
                            # Special case for Plantation location
                            elif "25101 - Plantation Walk Payable (1%)" in parent_account:
                                gl_data[date_key][location] += credit
                                # Track the total Plantation Walk Payable amount
                                plantation_payable_total += credit

                            file_rows_processed += 1
                            total_rows_processed += 1
                    except ValueError as e:
                        if "time data" in str(e):
                            # This is likely a summary row, so we'll skip it
                            continue
                        else:
                            self.log(f"Error processing row {i} in {os.path.basename(file)}: {str(e)}")
                    except (IndexError, KeyError) as e:
                        self.log(f"Error processing row {i} in {os.path.basename(file)}: {str(e)}")
                        continue

                self.log(f"Successfully processed {file_rows_processed} rows from GL file: {os.path.basename(file)}")
            except Exception as e:
                self.log(f"Error processing GL file {os.path.basename(file)}: {str(e)}")
                self.log(traceback.format_exc())

        # Store the plantation payable total in the gl_data dictionary
        gl_data["plantation_payable_total"] = plantation_payable_total

        self.log(f"GL data processing complete:")
        self.log(f"- Total rows processed: {total_rows_processed}")
        self.log(f"- Plantation payable total: {plantation_payable_total}")
        self.log(f"- Number of unique dates: {len(gl_data) - 1}")  # -1 for plantation_payable_total key
        self.log(f"- Gift card transactions: {sum(len(dates) for dates in gift_card_outstanding.values())}")

        return gl_data, gift_card_outstanding

    def process_export_data(self, export_files, valid_dates):
        export_data = defaultdict(lambda: defaultdict(float))

        total_rows_processed = 0
        self.log(f"Beginning to process {len(export_files)} export files")

        for file in export_files:
            self.log(f"Processing export file: {os.path.basename(file)}")
            try:
                rows = read_csv_with_encoding(file)
                self.log(f"Read {len(rows)} rows from file")

                if len(rows) <= 0:
                    self.log(f"File {os.path.basename(file)} is empty")
                    continue

                headers = rows[0]
                self.log(f"Headers: {headers}")

                try:
                    location_index = headers.index('Location')
                    date_index = headers.index('Date')
                    net_sales_index = headers.index('NetSales')
                except ValueError as e:
                    self.log(f"Error finding required columns: {str(e)}")
                    self.log(f"Available columns: {headers}")
                    continue

                file_rows_processed = 0
                for i, row in enumerate(rows[1:], start=1):
                    if not any(row):  # Skip completely empty rows
                        continue
                    try:
                        location = row[location_index].strip()

                        # Skip excluded locations
                        if location in ["Pinecrest", "West Kendall (London Square)", "Weston"]:
                            continue

                        date_str = row[date_index].split()[0].strip()
                        net_sales_str = row[net_sales_index].strip()

                        if not location or not date_str or not net_sales_str:
                            continue  # Skip rows with missing essential data

                        date = datetime.strptime(date_str, '%m/%d/%Y')
                        date_key = date.strftime('%Y-%m-%d')
                        net_sales = float(net_sales_str)

                        if date_key in valid_dates:
                            export_data[date_key][location] = net_sales
                            file_rows_processed += 1
                            total_rows_processed += 1
                    except (ValueError, IndexError) as e:
                        self.log(f"Error processing row {i} in export file {os.path.basename(file)}: {str(e)}")
                        continue

                self.log(f"Successfully processed {file_rows_processed} rows from export file: {os.path.basename(file)}")
            except Exception as e:
                self.log(f"Error processing export file {os.path.basename(file)}: {str(e)}")
                self.log(traceback.format_exc())

        self.log(f"Export data processing complete:")
        self.log(f"- Total rows processed: {total_rows_processed}")
        self.log(f"- Number of unique dates: {len(export_data)}")
        self.log(f"- Number of unique locations: {sum(len(locations) for locations in export_data.values())}")

        return export_data

    def process_group_data(self, group_files):
        group_data = {}

        total_rows_processed = 0
        self.log(f"Beginning to process {len(group_files)} group files")

        for file in group_files:
            self.log(f"Processing group file: {os.path.basename(file)}")
            try:
                rows = read_csv_with_encoding(file)
                self.log(f"Read {len(rows)} rows from file")

                if len(rows) <= 0:
                    self.log(f"File {os.path.basename(file)} is empty")
                    continue

                headers = rows[0]
                self.log(f"Headers: {headers}")

                try:
                    # Find the index of required columns
                    location_index = headers.index('Location')
                    net_sales_index = headers.index('Net Sales')
                except ValueError as e:
                    self.log(f"Error finding required columns: {str(e)}")
                    self.log(f"Available columns: {headers}")
                    continue

                file_rows_processed = 0
                for i, row in enumerate(rows[1:], start=1):
                    if not row or len(row) <= max(location_index, net_sales_index):
                        continue

                    try:
                        location = row[location_index].strip()

                        # Skip excluded locations
                        if location in ["Pinecrest", "West Kendall (London Square)", "Weston"]:
                            continue

                        net_sales_str = row[net_sales_index].strip()

                        if not location or not net_sales_str:
                            continue

                        # Convert net sales string to float
                        # Remove any currency symbols and commas
                        net_sales_str = net_sales_str.replace('$', '').replace(',', '')
                        net_sales = float(net_sales_str)

                        # Store in the dictionary
                        group_data[location] = net_sales
                        file_rows_processed += 1
                        total_rows_processed += 1

                    except (ValueError, IndexError) as e:
                        self.log(f"Error processing row {i} in group file {os.path.basename(file)}: {str(e)}")
                        continue

                self.log(f"Successfully processed {file_rows_processed} rows from group file: {os.path.basename(file)}")
            except Exception as e:
                self.log(f"Error processing group file {os.path.basename(file)}: {str(e)}")
                self.log(traceback.format_exc())

        self.log(f"Group data processing complete:")
        self.log(f"- Total rows processed: {total_rows_processed}")
        self.log(f"- Number of unique locations: {len(group_data)}")

        return group_data


class ToastReconcileWindow(RetroWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.selected_files = []
        self.output_dir = None

        icon_path = resource_path(os.path.join('assets', 'icon.png'))
        self.setWindowIcon(QIcon(icon_path))

    def initUI(self):
        central_widget = create_retro_central_widget(self)
        layout = QVBoxLayout(central_widget)

        # Add a retro-style title using the new title_label style
        title_label = QLabel("Toast Net Sales Reconcile", self)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setObjectName("title_label")  # This line applies the title_label style
        layout.addWidget(title_label)

        # Instructions button
        self.instructions_button = QPushButton('Instructions')
        self.instructions_button.clicked.connect(self.show_instructions)
        layout.addWidget(self.instructions_button)

        # Input Files button and list
        input_layout = QVBoxLayout()
        self.input_button = QPushButton('Input Files')
        self.input_button.clicked.connect(self.select_files)
        input_layout.addWidget(self.input_button)
        self.file_list = QListWidget()
        self.file_list.setFixedHeight(150)
        input_layout.addWidget(self.file_list)
        layout.addLayout(input_layout)

        # Output Directory button and label
        output_layout = QHBoxLayout()
        self.output_button = QPushButton('Output Directory')
        self.output_button.clicked.connect(self.select_output_directory)
        output_layout.addWidget(self.output_button)
        self.output_label = QLabel('No output directory selected')
        output_layout.addWidget(self.output_label)
        layout.addLayout(output_layout)

        # Run button
        self.run_button = QPushButton('RUN')
        self.run_button.clicked.connect(self.run_reconciliation)
        layout.addWidget(self.run_button)

        # Console output
        self.console_output = QTextEdit()
        self.console_output.setReadOnly(True)
        layout.addWidget(self.console_output)

        self.setWindowTitle('Toast Net Sales Reconcile')
        self.setFixedSize(1000, 738)
        # Center the window
        self.center()

    def center(self):
        frame_geometry = self.frameGeometry()
        center_point = QApplication.desktop().availableGeometry().center()
        frame_geometry.moveCenter(center_point)
        self.move(frame_geometry.topLeft())

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Input Files", "", "CSV Files (*.csv);;Excel Files (*.xlsx)"
        )
        if files:
            self.selected_files.extend(files)
            self.update_file_list()
            self.console_output.append(f"Selected {len(files)} file(s)")

    def update_file_list(self):
        self.file_list.clear()
        for file in self.selected_files:
            self.file_list.addItem(os.path.basename(file))

    def select_output_directory(self):
        self.output_dir = QFileDialog.getExistingDirectory(self, "Select Output Directory")
        if self.output_dir:
            self.output_label.setText(f"Output directory: {self.output_dir}")
            self.console_output.append(f"Selected output directory: {self.output_dir}")

    def run_reconciliation(self):
        if not self.selected_files or not self.output_dir:
            QMessageBox.warning(self, "Error", "Please select both input files and output directory.")
            return

        self.console_output.clear()
        self.console_output.append("Starting reconciliation process...")
        self.run_button.setEnabled(False)

        self.reconcile_thread = ReconcileThread(self.selected_files, self.output_dir)
        self.reconcile_thread.update_signal.connect(self.update_console)
        self.reconcile_thread.finished_signal.connect(self.reconciliation_finished)
        self.reconcile_thread.start()

    def update_console(self, message):
        self.console_output.append(message)

    def reconciliation_finished(self, success, message):
        self.run_button.setEnabled(True)
        if success:
            QMessageBox.information(self, "Success", message)
        else:
            QMessageBox.critical(self, "Error", message)
        self.console_output.append(message)

    def show_instructions(self):
        instructions = """
1. Get Toast net sales by downloading from "Orders" (not Order Details tab) tab in toast.
   You may upload several Order files from toast and the code will combine them for you.

2. From Financial Reports in R365 -> "GL Account Detail" -> "VIEW": Toast Net Sales Rec. GL -> Choose Date range

   "View Report" and save to CSV.

3. Download ALL DSS from Daily Sales Summary tab.

4. Download the Group Overview report from Toast for the same date range. "Sales" -> "Group sales overview"

5. Make sure not to modify any files once downloaded.

6. Click the "Input Files" button and select all the required CSV or Excel files.

7. Click the "Output Directory" button and select where you want the output files to be saved.

8. Click RUN to process the files.

Note: The program will ignore the locations "Pinecrest", "West Kendall (London Square)", and "Weston".
"""
        # Create a custom dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Instructions")

        # Create main layout
        main_layout = QVBoxLayout(dialog)

        # Create scroll area to handle potential overflow
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Add text instructions
        text_label = QLabel(instructions)
        text_label.setWordWrap(True)
        scroll_layout.addWidget(text_label)

        # Add some spacing at the bottom
        scroll_layout.addSpacing(10)
        # Set up scroll area
        scroll.setWidget(scroll_widget)
        main_layout.addWidget(scroll)
        # Create button layout
        button_layout = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(dialog.accept)
        button_layout.addStretch()
        button_layout.addWidget(ok_button)
        # Add button layout to main layout
        main_layout.addLayout(button_layout)
        # Set dialog size
        dialog.setMinimumWidth(1000)
        dialog.setMinimumHeight(800)
        dialog.exec_()
