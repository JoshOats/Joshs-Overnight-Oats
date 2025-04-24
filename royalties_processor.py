import csv
from collections import defaultdict
from datetime import datetime
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QThread, pyqtSignal

LOCATION_DICT = {
    "Aventura (Miami Gardens)": "Carrot Aventura Love LLC (Aventura)",
    "Aventura Mall": "Carrot Love Aventura Mall Operating LLC",
    "Boca Palmetto Park": "Carrot Love Palmetto Park Operating LLC",
    "Brickell": "Carrot Love Brickell Operating LLC",
    "Bryant Park": "Carrot Love Bryant Park Operating LLC",
    "Coconut Creek": "Carrot Love Coconut Creek Operating LLC",
    "Coconut Grove": "Carrot Love Coconut Grove Operating LLC",
    "Coral Gables": "Carrot Coral GablesLove LLC (Coral Gabes)",
    "Dadeland": "Carrot Love Dadeland Operating LLC",
    "Doral": "Carrot Love City Place Doral Operating LLC",
    "Downtown": "Carrot Downtown Love Two LLC",
    "Flatiron": "Carrot Flatiron Love Manhattan Operating LLC",
    "Hollywood": "Carrot Love Hollywood Operating LLC",
    "Las Olas": "Carrot Love Las Olas Operating LLC",
    "Lexington": "Carrot Love 600 Lexington LLC",
    "Miami Shores": "Carrot Express Miami Shores LLC",
    "Midtown": "Carrot Express Midtown LLC",
    "North Beach": "Carrot North Beach Love LL (North Beach)",
    "Pembroke Pines": "Carrot Love Pembroke Pines Operating LLC",
    "Plantation": "Carrot Love Plantation Operating LLC",
    "River Landing": "Carrot Love River Lading Operating LLC",
    "South Miami (Sunset)": "Carrot Love Sunset Operating LLC",
    "West Boca": "Carrot Love West Boca Operating LLC",
    "South Beach": "Carrot Sobe Love South Florida Operating C LLC"
}

# Define location groups
NEW_YORK_LOCATIONS = ["Bryant Park", "Lexington", "Flatiron"]
CARROT_LOVE_LOCATIONS = ["North Beach", "Aventura (Miami Gardens)", "Coral Gables"]
SOUTH_BEACH_LOCATIONS = ["South Beach"]
PLANTATION_LOCATIONS = ["Plantation"]

class RoyaltiesProcessThread(QThread):
    update_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, input_files, output_dir):
        super().__init__()
        self.input_files = input_files
        self.output_dir = output_dir

    def log_message(self, msg):
        print(msg)
        self.update_signal.emit(str(msg))

    def run(self):
        try:
            self.log_message("Starting Royalties processing...")

            # Categorize the input files
            group_files = [f for f in self.input_files if os.path.basename(f).startswith("GroupOverview")]
            order_files = [f for f in self.input_files if os.path.basename(f).startswith("Order")]
            gl_files = [f for f in self.input_files if os.path.basename(f).lower().startswith("gl")]
            profit_loss_files = [f for f in self.input_files if os.path.basename(f).lower().startswith("profit")]
            tax_files = [f for f in self.input_files if os.path.basename(f).startswith("Tax")]
            doordash_files = [f for f in self.input_files if os.path.basename(f).startswith("DoorDash")]
            grubhub_files = [f for f in self.input_files if os.path.basename(f).startswith("GrubHub")]
            ue_files = [f for f in self.input_files if os.path.basename(f).startswith("UE")]

            # Check if we have necessary files
            if not group_files:
                raise ValueError("No GroupOverview files found in the selected files.")

            # Process the data
            self.log_message("Processing sales data...")
            sales_data, earliest_date, latest_date = process_group_overview(group_files)

            self.log_message("Processing GL data...")
            r365_sales_tax_data, r365_resort_tax_data = process_gl_data(gl_files)

            self.log_message("Processing profit and loss data...")
            export_data = process_profit_loss_data(profit_loss_files)

            self.log_message("Processing tax-exempt data...")
            tax_exempt_data = process_tax_exempt_data(tax_files)

            self.log_message("Processing order data...")
            order_data = process_order_data_for_tax(order_files)

            self.log_message("Processing UberEats, GrubHub, and other delivery data...")
            uber_sales = process_uber_orders(order_files)
            grubhub_sales = process_grubhub_orders(order_files)
            profit_metrics = get_profit_metrics(profit_loss_files)

            # Process UE data
            ue_sales, ue_refunds, ue_discount, ue_sales_tax = process_ue_data(ue_files)

            # Add UE data to additional metrics
            additional_metrics = get_additional_metrics(profit_loss_files, doordash_files, grubhub_files, order_files)
            additional_metrics['ue_sales'] = ue_sales
            additional_metrics['ue_refunds'] = ue_refunds
            additional_metrics['ue_discount'] = ue_discount
            additional_metrics['ue_sales_tax'] = ue_sales_tax

            # Create output directory
            date_range = f"{earliest_date.strftime('%m%d%Y')}-{latest_date.strftime('%m%d%Y')}"
            output_dir = os.path.join(self.output_dir, f"Royalties {date_range}")
            Path(output_dir).mkdir(parents=True, exist_ok=True)

            # Save reports
            self.log_message("Generating royalties summary report...")
            save_royalties_report(sales_data, uber_sales, grubhub_sales, profit_metrics, additional_metrics, export_data,
                                 tax_exempt_data, order_data, r365_sales_tax_data, r365_resort_tax_data, earliest_date, latest_date, output_dir)

            self.log_message("Generating AR invoices...")
            ar_path = generate_ar_invoices(sales_data, uber_sales, profit_metrics, additional_metrics, export_data, r365_sales_tax_data, earliest_date, latest_date, output_dir)

            self.log_message("Generating AP invoices...")
            ap_path = generate_ap_invoices(sales_data, uber_sales, profit_metrics, additional_metrics, export_data, r365_sales_tax_data, earliest_date, latest_date, output_dir)

            self.log_message("Adding Tax tab and reordering tabs...")
            royalties_path = add_tax_tab_and_reorder(sales_data, order_data, tax_exempt_data, profit_metrics, earliest_date, latest_date, output_dir)

            self.log_message("Processing complete.")
            success_message = f"Processing complete. Files have been saved to:\n{output_dir}"
            self.finished_signal.emit(True, success_message)

        except Exception as e:
            error_message = f"Error processing royalties: {str(e)}"
            self.log_message(error_message)
            self.finished_signal.emit(False, error_message)

def read_csv_with_encoding(file_path, encodings=['utf-8', 'latin-1', 'cp1252']):
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                return list(reader)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Unable to decode file {file_path}")

def process_group_overview(group_files):
    sales_by_location = defaultdict(float)
    earliest_date = None
    latest_date = None

    for file in group_files:
        filename = os.path.basename(file)
        date_match = re.search(r'GroupOverview_(\d{4}_\d{2}_\d{2})-(\d{4}_\d{2}_\d{2})', filename)
        if date_match:
            start_date = datetime.strptime(date_match.group(1), '%Y_%m_%d')
            end_date = datetime.strptime(date_match.group(2), '%Y_%m_%d')

            if earliest_date is None or start_date < earliest_date:
                earliest_date = start_date
            if latest_date is None or end_date > latest_date:
                latest_date = end_date

            rows = read_csv_with_encoding(file)
            headers = rows[0]
            location_index = 0  # First column
            net_sales_index = 1  # Second column

            for row in rows[1:]:
                if not row:
                    continue
                try:
                    location = row[location_index]
                    if location in LOCATION_DICT.keys():  # Only process valid locations
                        net_sales_str = row[net_sales_index].replace('$', '').replace(',', '')
                        sales_by_location[location] = float(net_sales_str)
                except (IndexError, ValueError):
                    continue

    return sales_by_location, earliest_date, latest_date

def process_gl_data(gl_files):
    """
    Process GL data to get the R365 Sales Tax Payable for each location.
    For South Beach, specifically looks at Textbox49 to identify both Sales Tax and Resort Tax.
    """
    r365_sales_tax_by_location = defaultdict(float)
    r365_resort_tax_by_location = defaultdict(float)
    plantation_payable = 0

    for file in gl_files:
        rows = read_csv_with_encoding(file)
        header_row_index = 3
        headers = rows[header_row_index]

        location_index = headers.index('LocationName1')
        textbox33_index = headers.index('Textbox33')
        parent_account_index = headers.index('ParentAccountName')

        # Add Textbox49 index if it exists in the headers
        textbox49_index = headers.index('Textbox49') if 'Textbox49' in headers else -1

        south_beach_location = "Carrot Sobe Love South Florida Operating C LLC"

        for row in rows[header_row_index + 1:]:
            if not row or len(row) <= max(location_index, textbox33_index, parent_account_index):
                continue

            location = row[location_index]
            parent_account = row[parent_account_index] if len(row) > parent_account_index else ""

            # Regular sales tax for non-South Beach locations
            if "23000 - Sales Tax Payable" in parent_account and location != south_beach_location:
                sales_tax_payable = float(row[textbox33_index].replace(',', '')) if row[textbox33_index] else 0
                r365_sales_tax_by_location[location] = sales_tax_payable

            # Special handling for South Beach
            if location == south_beach_location and textbox49_index >= 0 and len(row) > textbox49_index:
                textbox49_value = row[textbox49_index] if row[textbox49_index] else ""
                textbox33_value = float(row[textbox33_index].replace(',', '')) if row[textbox33_index] else 0

                # Check both types in Textbox49
                if "Total Sales Tax Payable" in textbox49_value:
                    r365_sales_tax_by_location[location] = textbox33_value
                elif "Total Resort Tax Payable" in textbox49_value:
                    r365_resort_tax_by_location[location] = textbox33_value

            # Process Plantation payable for royalty calculations
            if "Carrot Love Plantation Operating LLC" in location:
                credit_index = headers.index('Credit')
                if "25101 - Plantation Walk Payable (1%)" in parent_account:
                    credit = float(row[credit_index].replace(',', '')) if row[credit_index] else 0
                    plantation_payable += credit

    # Store Plantation payable separately
    r365_sales_tax_by_location["Plantation_Payable"] = plantation_payable

    # Return both dictionaries
    return r365_sales_tax_by_location, r365_resort_tax_by_location

def process_profit_loss_data(profit_loss_files):
    profit_loss_by_location = defaultdict(float)
    processed_locations = set()

    for file in profit_loss_files:
        rows = read_csv_with_encoding(file)
        headers = rows[3]  # Headers start on row 4

        ganame_index = headers.index('gaName3')
        location_index = headers.index('ColumnGroupLabel')
        value_index = headers.index('ValueDisplay3')

        for row in rows[4:]:  # Start after headers
            if not row or len(row) <= max(ganame_index, location_index, value_index):
                continue

            ganame = row[ganame_index].strip()
            location = row[location_index].strip()

            # Skip if not "Total Sales" or location already processed
            if 'Total Sales' not in ganame or location in processed_locations:
                continue

            try:
                # Remove $ and , from value and convert to float
                value_str = row[value_index].strip().replace('$', '').replace(',', '')
                net_sales = float(value_str)
                profit_loss_by_location[location] = net_sales
                processed_locations.add(location)
            except (ValueError, IndexError):
                continue

    return profit_loss_by_location

def process_uber_orders(order_files):
    uber_sales_by_location = defaultdict(float)

    for file in order_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        dining_index = headers.index('Dining Options')
        location_index = headers.index('Location')
        amount_index = headers.index('Amount')

        for row in rows[1:]:
            if not row:
                continue

            dining_option = row[dining_index]
            if dining_option in ['UberEats (Pickup)', 'Uber Eats - Delivery!']:
                location = row[location_index]
                amount = float(row[amount_index])
                uber_sales_by_location[location] -= amount  # Make negative

    return uber_sales_by_location

def get_profit_metrics(profit_loss_files):
    metrics = {
        'uber_sales': defaultdict(float),
        'delivery_fee': defaultdict(float),
        'ez_catering': defaultdict(float),
        'uber_discount': defaultdict(float)
    }
    processed_locations = {metric: set() for metric in metrics.keys()}

    for file in profit_loss_files:
        rows = read_csv_with_encoding(file)
        headers = rows[3]

        ganame_index = headers.index('gaName3')
        location_index = headers.index('ColumnGroupLabel')
        value_index = headers.index('ValueDisplay3')

        for row in rows[4:]:
            if not row:
                continue

            ganame = row[ganame_index].strip()
            location = row[location_index].strip()

            try:
                value_str = row[value_index].strip().replace('$', '').replace(',', '')
                value = float(value_str)

                # Use "Total UberEats Sales" instead of "UberEats Sales"
                if 'Total UberEats Sales' in ganame and location not in processed_locations['uber_sales']:
                    metrics['uber_sales'][location] = value
                    processed_locations['uber_sales'].add(location)
                elif 'Delivery Fee Income' in ganame and location not in processed_locations['delivery_fee']:
                    metrics['delivery_fee'][location] = value
                    processed_locations['delivery_fee'].add(location)
                elif 'Ez catering' in ganame and location not in processed_locations['ez_catering']:
                    metrics['ez_catering'][location] = value
                    processed_locations['ez_catering'].add(location)
                elif 'UberEats Discount' in ganame and location not in processed_locations['uber_discount']:
                    metrics['uber_discount'][location] = value
                    processed_locations['uber_discount'].add(location)
            except ValueError:
                continue

    return metrics

def process_grubhub_orders(order_files):
    """Process order files to get Grubhub orders total"""
    grubhub_sales_by_location = defaultdict(float)

    for file in order_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        dining_index = headers.index('Dining Options')
        location_index = headers.index('Location')
        amount_index = headers.index('Amount')


        for row in rows[1:]:
            if not row:
                continue

            dining_option = row[dining_index]
            if dining_option in ['Grubhub (Delivery)', 'Grubhub (Takeout)']:
                location = row[location_index]
                amount = float(row[amount_index])
                grubhub_sales_by_location[location] -= amount  # Make negative like other delivery services


    return grubhub_sales_by_location

def process_doordash_data(doordash_files):
    """Process DoorDash files to get sales, refunds and discounts"""
    doordash_sales_by_location = defaultdict(float)
    doordash_refunds_by_location = defaultdict(float)
    doordash_discounts_by_location = defaultdict(float)  # Added for DoorDash discounts

    for file in doordash_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        # Get column indices
        location_index = headers.index('JELocation')
        account_index = headers.index('Account')
        credit_index = headers.index('Credit')
        debit_index = headers.index('Debit')

        for row in rows[1:]:
            if not row or len(row) <= max(location_index, account_index, credit_index, debit_index):
                continue

            location = row[location_index]
            account = row[account_index]

            # DoorDash Sales: Account is "DD Delivery" or "Pickup", sum the Credit column
            if account in ['DD Delivery', 'DD Pickup']:
                credit_value = float(row[credit_index]) if row[credit_index] else 0
                doordash_sales_by_location[location] += credit_value

            # DoorDash Refunds: Account is "Exchange", sum the Debit column and make it negative
            elif account == 'Refunds':
                debit_value = float(row[debit_index]) if row[debit_index] else 0
                doordash_refunds_by_location[location] -= debit_value  # Make it negative

            # DoorDash Discounts: Account contains "Doordash Discount", calculate Credit - Debit
            elif "Doordash Discount" in account:
                credit_value = float(row[credit_index]) if row[credit_index] else 0
                debit_value = float(row[debit_index]) if row[debit_index] else 0
                doordash_discounts_by_location[location] += credit_value - debit_value

    return doordash_sales_by_location, doordash_refunds_by_location, doordash_discounts_by_location

def process_grubhub_data(grubhub_files):
    """Process Grubhub files to get delivery fees and promotions"""
    grubhub_delivery_fees = defaultdict(float)
    grubhub_promotions = defaultdict(float)
    grubhub_refunds = defaultdict(float)

    for file in grubhub_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        # Get column indices
        location_index = headers.index('JELocation')
        account_index = headers.index('Account')
        credit_index = headers.index('Credit')
        debit_index = headers.index('Debit')
        detail_comment_index = headers.index('DetailComment') if 'DetailComment' in headers else -1

        for row in rows[1:]:
            if not row or len(row) <= max(location_index, account_index, credit_index, debit_index):
                continue

            location = row[location_index]
            account = row[account_index]

            # Check each condition independently
            if account == 'Delivery Fee Income':
                credit_value = float(row[credit_index]) if row[credit_index] else 0
                grubhub_delivery_fees[location] += credit_value

            # Special Bryant Park case checked independently
            if detail_comment_index >= 0 and location == "Carrot Love Bryant Park Operating LLC":
                detail_comment = row[detail_comment_index]
                if detail_comment and "Delivery Fee" in detail_comment:
                    credit_value = float(row[credit_index]) if row[credit_index] else 0
                    grubhub_delivery_fees[location] += credit_value

            # Promotions checked independently, not with elif
            if account in ['GrubHub Discount', 'Rewards']:
                debit_value = float(row[debit_index]) if row[debit_index] else 0
                grubhub_promotions[location] -= debit_value

            # Refunds checked independently
            if account == 'Refunds':
                debit_value = float(row[debit_index]) if row[debit_index] else 0
                grubhub_refunds[location] -= debit_value

    return grubhub_delivery_fees, grubhub_promotions, grubhub_refunds

def process_third_party_orders(order_files):
    """Process order files to get Third Parties total"""
    third_party_by_location = defaultdict(float)

    # For debugging - track what we're finding
    found_third_parties = defaultdict(list)

    for file in order_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        dining_index = headers.index('Dining Options')
        location_index = headers.index('Location')
        amount_index = headers.index('Amount')

        for row in rows[1:]:
            if not row:
                continue

            dining_option = row[dining_index]
            if dining_option in ['Sharebite', 'MealPal', 'Forkable', 'Foodie for All']:
                location = row[location_index]
                amount = float(row[amount_index])

                third_party_by_location[location] += amount

                # For debugging
                found_third_parties[location].append((dining_option, amount))

    return third_party_by_location

def process_ue_data(ue_files):
    """Process UE files to get sales, refunds, discounts and sales tax data"""
    ue_sales_by_location = defaultdict(float)
    ue_refunds_by_location = defaultdict(float)
    ue_discount_by_location = defaultdict(float)
    ue_sales_tax_by_location = defaultdict(float)

    for file in ue_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        # Get column indices
        location_index = headers.index('JELocation')
        account_index = headers.index('Account')
        credit_index = headers.index('Credit')
        debit_index = headers.index('Debit')

        for row in rows[1:]:
            if not row or len(row) <= max(location_index, account_index, credit_index, debit_index):
                continue

            location = row[location_index]
            account = row[account_index]

            # UberEats Sales: Account is "Pickup & Takeout" or "UE Delivery", sum Credit
            if account in ['UE Pickup & Takeout', 'UE Delivery']:
                credit_value = float(row[credit_index]) if row[credit_index] else 0
                ue_sales_by_location[location] += credit_value

            # UberEats Refunds: Account is "Refunds", calculate Credit - Debit
            elif account == 'Refunds':
                credit_value = float(row[credit_index]) if row[credit_index] else 0
                debit_value = float(row[debit_index]) if row[debit_index] else 0
                ue_refunds_by_location[location] += credit_value - debit_value

            # UberEats Discount: Account is "UberEats Discount", sum Debit
            elif account == 'UberEats Discount':
                debit_value = float(row[debit_index]) if row[debit_index] else 0
                ue_discount_by_location[location] -= debit_value  # Negate to match existing data format

            # UberEats Sales Tax: Account is "Sales Tax Payable", calculate Credit - Debit
            elif account == 'Sales Tax Payable':
                credit_value = float(row[credit_index]) if row[credit_index] else 0
                debit_value = float(row[debit_index]) if row[debit_index] else 0
                ue_sales_tax_by_location[location] += credit_value - debit_value

    return ue_sales_by_location, ue_refunds_by_location, ue_discount_by_location, ue_sales_tax_by_location

def get_additional_metrics(profit_loss_files, doordash_files, grubhub_files, order_files):
    """Get additional metrics for New York locations"""

    # Get metrics from profit loss files
    metrics = {
        'grubhub_sales': defaultdict(float),
        'doordash_sales': defaultdict(float),
        'third_parties': defaultdict(float)
    }
    processed_locations = {metric: set() for metric in metrics.keys()}

    for file in profit_loss_files:
        rows = read_csv_with_encoding(file)
        headers = rows[3]

        ganame_index = headers.index('gaName3')
        location_index = headers.index('ColumnGroupLabel')
        value_index = headers.index('ValueDisplay3')

        for row in rows[4:]:
            if not row:
                continue

            ganame = row[ganame_index].strip()
            location = row[location_index].strip()

            try:
                value_str = row[value_index].strip().replace('$', '').replace(',', '')
                value = float(value_str)

                if 'Total Grubhub Sales' in ganame and location not in processed_locations['grubhub_sales']:
                    metrics['grubhub_sales'][location] = value
                    processed_locations['grubhub_sales'].add(location)
                # We'll get additional metrics from the other files instead
            except ValueError:
                continue

    # Get DoorDash sales, refunds, and discounts from DoorDash files
    doordash_sales, doordash_refunds, doordash_discounts = process_doordash_data(doordash_files)

    # Get Grubhub delivery fees, promotions, and refunds from Grubhub files
    grubhub_delivery_fees, grubhub_promotions, grubhub_refunds = process_grubhub_data(grubhub_files)

    # Get third parties data from order files
    third_parties = process_third_party_orders(order_files)

    # Add these to the metrics dictionary
    metrics['doordash_sales'] = doordash_sales
    metrics['doordash_refunds'] = doordash_refunds
    metrics['doordash_discounts'] = doordash_discounts  # Add the new discounts data
    metrics['grubhub_delivery_fees'] = grubhub_delivery_fees
    metrics['grubhub_promotions'] = grubhub_promotions
    metrics['grubhub_refunds'] = grubhub_refunds
    metrics['third_parties'] = third_parties

    return metrics

def process_tax_exempt_data(tax_files):
    tax_exempt_by_location = defaultdict(float)

    for file in tax_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        location_index = headers.index('Location')
        amount_index = headers.index('Amount')

        for row in rows[1:]:
            if not row:
                continue
            try:
                location = row[location_index]
                amount = float(row[amount_index])
                tax_exempt_by_location[location] -= amount  # Make negative
            except (ValueError, IndexError):
                continue

    return tax_exempt_by_location

def process_order_data_for_tax(order_files):
    all_orders = []
    for file in order_files:
        rows = read_csv_with_encoding(file)
        headers = rows[0]

        # Convert to list of dictionaries for easier processing
        for row in rows[1:]:
            if row and len(row) >= len(headers):  # Skip empty or incomplete rows
                order_dict = {headers[i]: row[i] for i in range(len(headers))}
                all_orders.append(order_dict)

    return all_orders

def save_royalties_report(sales_data, uber_sales, grubhub_sales, profit_metrics, additional_metrics, export_data, tax_exempt_data, order_data, r365_sales_tax_data, r365_resort_tax_data, earliest_date, latest_date, output_dir):
    output_filename = f"Royalties_Summary_{earliest_date.strftime('%m%d%Y')}-{latest_date.strftime('%m%d%Y')}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    wb = Workbook()
    wb.remove(wb.active)

    headers = ['Metric', 'Amount']
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Styles for positive and negative numbers
    positive_format = '#,##0.00'
    negative_format = '[Red]#,##0.00'

    # Store FDOR tax values for Carrot Love locations
    carrot_love_tax_data = {}
    # Store Tax Reported values for Carrot Love locations
    carrot_love_tax_reported = {}

    # First pass to collect necessary data for Carrot Love locations
    for location, toast_sales in sales_data.items():
        if location in CARROT_LOVE_LOCATIONS:
            gl_location = LOCATION_DICT.get(location, location)

            # Calculate tax values needed for North Beach
            non_taxable = tax_exempt_data.get(location, 0)
            uber_sales_r365 = profit_metrics['uber_sales'].get(gl_location, 0)
            uber_refunds = -abs(toast_sales + uber_sales.get(location, 0) + uber_sales_r365 +
                            profit_metrics['uber_discount'].get(gl_location, 0) - export_data.get(gl_location, 0))
            uber_discount = profit_metrics['uber_discount'].get(gl_location, 0)
            uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount
            net_sales_pnl = export_data.get(gl_location, 0)

            # Calculate sales_reported as Net Sales PNL - UberEats total
            sales_reported = net_sales_pnl - uber_sales_total
            tax_seven_percent = round(sales_reported * 0.07, 2)

            # Get Total Tax for this location
            total_tax = sum(float(row['Tax']) for row in order_data if row['Location'] == location)

            # Calculate UberEats Tax
            uber_tax = -sum(float(row['Tax'])
                        for row in order_data
                        if row['Location'] == location
                        and row['Dining Options'] in ['UberEats (Pickup)', 'Uber Eats - Delivery!'])

            # Calculate Tax Reported
            tax_reported = total_tax + uber_tax

            carrot_love_tax_data[location] = {
                'non_taxable': non_taxable,
                'sales_reported': sales_reported,
                'tax_seven_percent': tax_seven_percent,
                'net_sales_pnl': net_sales_pnl,
                'uber_sales_total': uber_sales_total
            }

            carrot_love_tax_reported[location] = tax_reported

    # Calculate the sum of Tax Reported for all Carrot Love locations
    all_carrot_love_tax_reported = sum(reported for location, reported in carrot_love_tax_reported.items())

    # Main processing loop for all locations
    for location, toast_sales in sales_data.items():
        ws = wb.create_sheet(location)
        gl_location = LOCATION_DICT.get(location, location)

        # Set headers for main table
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border

        # Get metrics
        uber_sales_toast = uber_sales.get(location, 0)
        uber_sales_r365 = additional_metrics['ue_sales'].get(gl_location, 0)
        delivery_fee = profit_metrics['delivery_fee'].get(gl_location, 0)
        ez_catering = profit_metrics['ez_catering'].get(gl_location, 0)
        uber_discount = additional_metrics['ue_discount'].get(gl_location, 0)
        net_sales_pnl = export_data.get(gl_location, 0)

       # Get UberEats Refunds from UE file
        uber_refunds = additional_metrics['ue_refunds'].get(gl_location, 0)

        # Special handling for New York locations
        if location in NEW_YORK_LOCATIONS:
            # Get additional metrics for New York
            grubhub_toast = grubhub_sales.get(location, 0)
            grubhub_r365 = additional_metrics['grubhub_sales'].get(gl_location, 0)
            doordash_sales = additional_metrics['doordash_sales'].get(gl_location, 0)
            doordash_refunds = additional_metrics['doordash_refunds'].get(gl_location, 0)
            doordash_discounts = additional_metrics['doordash_discounts'].get(gl_location, 0)
            grubhub_refunds = additional_metrics['grubhub_refunds'].get(gl_location, 0)

            # Get grubhub delivery fees and promotions
            grubhub_delivery_fees = additional_metrics['grubhub_delivery_fees'].get(gl_location, 0)
            grubhub_promotions = additional_metrics['grubhub_promotions'].get(gl_location, 0)
            third_parties = additional_metrics['third_parties'].get(location, 0)

            # For non-Bryant Park NY locations, calculate adjusted delivery fee
            if location != "Bryant Park":
                delivery_fee_display = delivery_fee - grubhub_delivery_fees
            else:
                delivery_fee_display = delivery_fee

            # Build main_data array with proper values
            main_data = [
                ['Toast Net Sales', toast_sales],
                ['UberEats Sales - Toast', uber_sales_toast],
                ['UberEats Sales - R365', uber_sales_r365],
                ['Delivery Fee Income', delivery_fee_display],
                ['Ez Catering', ez_catering],
                ['UberEats Refunds', uber_refunds],
                ['UberEats Discount', uber_discount],
                ['Grubhub Toast', grubhub_toast],
                ['Grubhub R365', grubhub_r365],
                ['DoorDash Sales', doordash_sales],
                ['DoorDash Discount', doordash_discounts],  # Add this line
                ['DoorDash Refunds', doordash_refunds],
                ['Grubhub Delivery Fees', grubhub_delivery_fees],
                ['Grubhub Promotions', grubhub_promotions],
                ['Grubhub Refunds', grubhub_refunds],
                ['Third Parties', third_parties],
                ['Net Sales PNL', net_sales_pnl]
            ]
        else:
            # Special handling for Plantation - add Plantation Walk Payable to Delivery Fee
            if location in PLANTATION_LOCATIONS:
                plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
                delivery_fee += plantation_payable  # Add plantation payable to delivery fee
            # Standard data for non-New York locations
            main_data = [
                ['Toast Net Sales', toast_sales],
                ['UberEats Sales - Toast', uber_sales_toast],
                ['UberEats Sales - R365', uber_sales_r365],
                ['Delivery Fee Income', delivery_fee],
            ]

            # Add Non-Grat Svc Charges for Plantation location
            if location in PLANTATION_LOCATIONS:
                plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
                # Insert right after Delivery Fee Income with negative value
                main_data.append(['Non-Grat Svc Charges', -plantation_payable])

            # Continue with the rest of the items
            main_data.extend([
                ['Ez Catering', ez_catering],
                ['UberEats Refunds', uber_refunds],
                ['UberEats Discount', uber_discount],
            ])

            # Add Plantation Walk Payable for Plantation location (keeping this code as it was)
            if location in PLANTATION_LOCATIONS:
                plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
                # Insert before Net Sales PNL
                main_data.append(['Plantation Walk Shops 1%', plantation_payable])

            # Add Net Sales PNL as the final item
            main_data.append(['Net Sales PNL', net_sales_pnl])

        # Write main data with conditional formatting
        for row_idx, (metric, amount) in enumerate(main_data, 2):
            metric_cell = ws.cell(row=row_idx, column=1, value=metric)
            metric_cell.border = thin_border
            if metric == 'Net Sales PNL':
                metric_cell.font = Font(bold=True)  # Bold "Net Sales PNL"

            amount_cell = ws.cell(row=row_idx, column=2, value=amount)
            amount_cell.border = thin_border
            amount_cell.number_format = positive_format if amount >= 0 else negative_format

        # Add leadership table - fixed position to align with the end of the main data table
        if location in ["Midtown", "Miami Shores"]:
            leadership_headers = ['Cleadership - 2%', 'Leadership -1%']
        else:
            leadership_headers = ['Cleadership - 2%']

        if location in NEW_YORK_LOCATIONS:
            leadership_row = len(main_data) + 1   # Position at the end of the extended main data
        else:
            leadership_row = len(main_data) + 1  # Position after main data

        for col, header in enumerate(leadership_headers, 3):
            cell = ws.cell(row=leadership_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border

        # Calculate leadership values with updated formula
        if location in NEW_YORK_LOCATIONS and location != "Bryant Park":
            # For NY locations except Bryant Park, subtract Grubhub Delivery Fees
            grubhub_fees_index = -1
            for i, row in enumerate(main_data):
                if row[0] == 'Grubhub Delivery Fees':
                    grubhub_fees_index = i + 2  # +2 for Excel row number
                    break

            leadership_cell = ws.cell(row=leadership_row + 1, column=3, value="=(B" + str(main_data[-1][0] == 'Net Sales PNL' and row_idx or row_idx-1) + "-B" + str(main_data[3][0] == 'Delivery Fee Income' and 5 or 4) + "-B" + str(grubhub_fees_index) + ")*0.02")
        else:
            # Standard formula for other locations
            leadership_cell = ws.cell(row=leadership_row + 1, column=3, value="=(B" + str(main_data[-1][0] == 'Net Sales PNL' and row_idx or row_idx-1) + "-B" + str(main_data[3][0] == 'Delivery Fee Income' and 5 or 4) + ")*0.02")

        leadership_cell.number_format = positive_format
        leadership_cell.border = thin_border

        if location in ["Midtown", "Miami Shores"]:
            next_leadership_cell = ws.cell(row=leadership_row + 1, column=4, value="=(B" + str(main_data[2][0] == 'UberEats Sales - R365' and 4 or 3) + "+B" + str(main_data[5][0] == 'UberEats Refunds' and 7 or 6) + "+B" + str(main_data[6][0] == 'UberEats Discount' and 8 or 7) + ")*0.01")
            next_leadership_cell.number_format = positive_format
            next_leadership_cell.border = thin_border

        # Base values for tax tables
        non_taxable = tax_exempt_data.get(location, 0)
        uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount

        # Different sales_reported calculation based on location
        if location in PLANTATION_LOCATIONS:
            plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
            sales_reported = net_sales_pnl - uber_sales_total + non_taxable + plantation_payable
        elif location in NEW_YORK_LOCATIONS:
            # Modified calculation for New York locations
            sales_reported = net_sales_pnl + non_taxable
        else:
            sales_reported = net_sales_pnl - uber_sales_total + non_taxable

        # Tax rate depends on location
        if location in NEW_YORK_LOCATIONS:
            tax_rate = 0.08875  # 8.875% for New York
            tax_name = '8.875%'
        else:
            tax_rate = 0.07  # 7% for other locations
            tax_name = '7%'

        tax_amount = round(sales_reported * tax_rate, 2)

        # Calculate values for Toast tax table
        total_tax = sum(float(row['Tax']) for row in order_data if row['Location'] == location)

        # Special handling for New York locations - calculate Tax on Promotions
        if location in NEW_YORK_LOCATIONS:
            # Get grubhub promotions value (using placeholder)
            grubhub_promotions_value = additional_metrics['grubhub_promotions'].get(gl_location, 0)
            tax_on_promotions = round((uber_refunds + uber_discount + grubhub_promotions_value) * tax_rate, 2)
            uber_tax = tax_on_promotions  # For consistency in naming
        else:
            # Standard UberEats Sales Tax for non-NY locations
            uber_tax = -sum(float(row['Tax'])
                        for row in order_data
                        if row['Location'] == location
                        and row['Dining Options'] in ['UberEats (Pickup)', 'Uber Eats - Delivery!'])

        tax_reported = total_tax + uber_tax

        # Get R365 Sales Tax Payable for this location (or 0 if not found)
        r365_sales_tax_payable = r365_sales_tax_data.get(gl_location, 0)

        # For North Beach, calculate special Carrot Love value
        carrot_love_sales_tax_payable = 0
        if location == "North Beach":
            # Sum the R365 Sales Tax Payable for all Carrot Love locations
            carrot_love_r365_locations = [LOCATION_DICT.get(loc) for loc in CARROT_LOVE_LOCATIONS]
            carrot_love_sales_tax_payable = sum(r365_sales_tax_data.get(loc, 0) for loc in carrot_love_r365_locations)

        # Special handling for Carrot Love locations
        if location in CARROT_LOVE_LOCATIONS:
            if location == "North Beach":
                # SALES TAX - FDOR table for North Beach (using data from other Carrot Love locations)
                fdor_row = 1
                fdor_col = 6

                ws.merge_cells(start_row=fdor_row, start_column=fdor_col,
                              end_row=fdor_row, end_column=fdor_col + 1)
                header_cell = ws.cell(row=fdor_row, column=fdor_col, value='SALES TAX - FDOR')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=fdor_row, column=fdor_col + 1).border = thin_border

                # Correctly calculating combined_sales_reported
                combined_non_taxable = sum(data['non_taxable'] for loc, data in carrot_love_tax_data.items())
                combined_sales_reported = sum(data['net_sales_pnl'] - data['uber_sales_total']
                                             for loc, data in carrot_love_tax_data.items())
                combined_tax_seven_percent = round(combined_sales_reported * 0.07, 2)

                fdor_data = [
                    ['NON TAXABLE', combined_non_taxable],
                    ['SALES REPORTED', combined_sales_reported],
                    ['7%', combined_tax_seven_percent]
                ]

                for row_idx, (label, value) in enumerate(fdor_data, fdor_row + 1):
                    ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                    cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                    cell.border = thin_border
                    cell.number_format = positive_format if value >= 0 else negative_format

                # RESORT TAX table for North Beach
                resort_row = fdor_row + len(fdor_data) + 1

                ws.merge_cells(start_row=resort_row, start_column=fdor_col,
                              end_row=resort_row, end_column=fdor_col + 1)
                header_cell = ws.cell(row=resort_row, column=fdor_col, value='RESORT TAX')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=resort_row, column=fdor_col + 1).border = thin_border

                resort_non_taxable = non_taxable
                resort_sales_reported = net_sales_pnl
                resort_tax_two_percent = round((resort_sales_reported + resort_non_taxable) * 0.02, 2)

                resort_data = [
                    ['NON TAXABLE - RESORT', resort_non_taxable],
                    ['SALES REPORTED - RESORT', resort_sales_reported],
                    ['2%', resort_tax_two_percent]
                ]

                for row_idx, (label, value) in enumerate(resort_data, resort_row + 1):
                    ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                    cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                    cell.border = thin_border
                    cell.number_format = positive_format if value >= 0 else negative_format

                # Total Tax table
                total_tax_row = resort_row + len(resort_data) + 1

                total_tax_value = combined_tax_seven_percent + resort_tax_two_percent

                ws.cell(row=total_tax_row, column=fdor_col, value='TOTAL TAX').border = thin_border
                cell = ws.cell(row=total_tax_row, column=fdor_col + 1, value="=G4+G8")  # Formula
                cell.border = thin_border
                cell.number_format = positive_format

                # SALES TAX - Toast table for North Beach - with renamed 'Total Tax' to 'Total Tax Toast'
                tax_toast_row = total_tax_row + 2

                ws.merge_cells(start_row=tax_toast_row, start_column=fdor_col,
                              end_row=tax_toast_row, end_column=fdor_col + 1)
                header_cell = ws.cell(row=tax_toast_row, column=fdor_col, value='SALES TAX - Toast')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=tax_toast_row, column=fdor_col + 1).border = thin_border

                # For North Beach - calculate Total Tax Toast as just the tax from all orders for North Beach location
                total_tax_toast = total_tax  # Just North Beach tax from Orders file

                # Get UberEats Sales Tax directly from UE file
                ue_sales_tax_value = additional_metrics['ue_sales_tax'].get(gl_location, 0)

                tax_toast_data = [
                    ['Total Tax Toast', total_tax_toast],
                    ['UberEats Sales Tax', ue_sales_tax_value],  # Use UE sales tax data
                    ['Tax Reported', "=G12+G13"]  # Formula referencing Total Tax Toast minus UberEats Sales Tax
                ]

                for row_idx, (label, value) in enumerate(tax_toast_data, tax_toast_row + 1):
                    ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                    cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                    cell.border = thin_border
                    if isinstance(value, (int, float)):  # Check if value is a number
                        cell.number_format = positive_format if value >= 0 else negative_format
                    else:  # It's a formula or other string
                        cell.number_format = positive_format  # Default to positive format for formulas

                # Add R365 table
                r365_row = tax_toast_row + len(tax_toast_data) + 1

                ws.merge_cells(start_row=r365_row, start_column=fdor_col,
                              end_row=r365_row, end_column=fdor_col + 1)
                header_cell = ws.cell(row=r365_row, column=fdor_col, value='R365')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=r365_row, column=fdor_col + 1).border = thin_border

                r365_data = [
                    ['Sales Tax Payable', ""],
                    ['Resort Tax Payable', ""],# Add the new Resort Tax Payable row
                    ['Carrot Love TOTAL Sales Tax Payable', "=G17+G16+'Coral Gables'!G6+'Aventura (Miami Gardens)'!G6"]  # Update formula reference to account for new row
                ]

                for row_idx, (label, value) in enumerate(r365_data, r365_row + 1):
                    ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                    cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                    cell.border = thin_border
                    if isinstance(value, (int, float)):  # Check if value is a number
                        cell.number_format = positive_format if value >= 0 else negative_format
                    else:  # It's a formula or other string
                        cell.number_format = positive_format  # Default to positive format for formulas

                # Highlight Sales Tax Payable and Carrot Love Sales Tax Payable cells
                ws.cell(row=r365_row + 1, column=fdor_col + 1).fill = yellow_fill
                ws.cell(row=r365_row + 2, column=fdor_col + 1).fill = yellow_fill
                ws.cell(row=r365_row + 3, column=fdor_col + 1).fill = yellow_fill

                # Add note next to Sales Tax Payable
                ws.cell(row=r365_row + 1, column=fdor_col + 2, value="<-- Enter Sales Tax Payable For North Beach").alignment = Alignment(horizontal='left')
                ws.cell(row=r365_row + 2, column=fdor_col + 2, value="<-- Enter Resort Tax Payable For North Beach").alignment = Alignment(horizontal='left')

                # Add 'Differences' table with updated calculations
                diff_row = r365_row + len(r365_data) + 1

                ws.merge_cells(start_row=diff_row, start_column=fdor_col,
                              end_row=diff_row, end_column=fdor_col + 1)
                header_cell = ws.cell(row=diff_row, column=fdor_col, value='Differences')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=diff_row, column=fdor_col + 1).border = thin_border

                diff_data = [
                    ['DIFFERENCE Toast - R365', "=G14-G16"],
                    ['DIFFERENCE Excel - R365', "=G9-G16"],
                    ['DIFFERENCE Toast - Excel', "=G14-G9"]
                ]

                for row_idx, (label, value) in enumerate(diff_data, diff_row + 1):
                    ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                    cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                    cell.border = thin_border
                    if isinstance(value, (int, float)):  # Check if value is a number
                        cell.number_format = positive_format if value >= 0 else negative_format
                    else:  # It's a formula or other string
                        cell.number_format = positive_format  # Default to positive format for formulas

                # Highlight all three DIFFERENCE cells
                ws.cell(row=diff_row + 1, column=fdor_col + 1).fill = yellow_fill
                ws.cell(row=diff_row + 2, column=fdor_col + 1).fill = yellow_fill
                ws.cell(row=diff_row + 3, column=fdor_col + 1).fill = yellow_fill  # Highlight the new DIFFERENCE Toast - Excel cell

                # Add 'Carrot Love LLC' table
                cl_row = diff_row + len(diff_data) + 1

                ws.merge_cells(start_row=cl_row, start_column=fdor_col,
                               end_row=cl_row, end_column=fdor_col + 1)
                header_cell = ws.cell(row=cl_row, column=fdor_col, value='Carrot Love LLC')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=cl_row, column=fdor_col + 1).border = thin_border

                # Just one row with Tax Reported (combined from all 3 Carrot Love locations)
                ws.cell(row=cl_row + 1, column=fdor_col, value='Tax Reported').border = thin_border
                cell = ws.cell(row=cl_row + 1, column=fdor_col + 1, value="=G14+'Coral Gables'!G4+'Aventura (Miami Gardens)'!G4")
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = positive_format if cell.value >= 0 else negative_format
                else:
                    cell.number_format = positive_format  # Default to positive format for formulas

                # Add 15(d) value (1% of combined_sales_reported)
                fifteen_d_row = cl_row + 2
                fifteen_d_value = round(combined_sales_reported * 0.01, 2)

                fifteen_d_cell = ws.cell(row=fifteen_d_row, column=fdor_col, value='15(d)')
                fifteen_d_cell.border = thin_border

                value_cell = ws.cell(row=fifteen_d_row, column=fdor_col + 1, value="=G3*0.01")  # Formula: 1% of SALES REPORTED
                value_cell.border = thin_border
                value_cell.number_format = positive_format

            else:
                # For Aventura and Coral Gables, just add SALES TAX - Toast table and R365 table
                tax_toast_row = 1
                tax_toast_col = 6

                ws.merge_cells(start_row=tax_toast_row, start_column=tax_toast_col,
                              end_row=tax_toast_row, end_column=tax_toast_col + 1)
                header_cell = ws.cell(row=tax_toast_row, column=tax_toast_col, value='SALES TAX - Toast')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=tax_toast_row, column=tax_toast_col + 1).border = thin_border

                tax_toast_data = [
                    ['Total Tax Toast', total_tax],
                    ['UberEats Sales Tax', uber_tax],
                    ['Tax Reported', tax_reported]
                ]

                for row_idx, (label, value) in enumerate(tax_toast_data, tax_toast_row + 1):
                    ws.cell(row=row_idx, column=tax_toast_col, value=label).border = thin_border
                    cell = ws.cell(row=row_idx, column=tax_toast_col + 1, value=value)
                    cell.border = thin_border
                    cell.number_format = positive_format if value >= 0 else negative_format

                # Add R365 table
                r365_row = tax_toast_row + len(tax_toast_data) + 1

                ws.merge_cells(start_row=r365_row, start_column=tax_toast_col,
                              end_row=r365_row, end_column=tax_toast_col + 1)
                header_cell = ws.cell(row=r365_row, column=tax_toast_col, value='R365')
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.font = Font(bold=True)
                header_cell.fill = header_fill
                header_cell.border = thin_border
                ws.cell(row=r365_row, column=tax_toast_col + 1).border = thin_border

                # Just one row with Sales Tax Payable
                ws.cell(row=r365_row + 1, column=tax_toast_col, value='Sales Tax Payable').border = thin_border
                cell = ws.cell(row=r365_row + 1, column=tax_toast_col + 1, value="")  # Empty string
                cell.border = thin_border
                cell.fill = yellow_fill  # Highlight with yellow

                # Add note next to Sales Tax Payable
                ws.cell(row=r365_row + 1, column=tax_toast_col + 2, value=" <-- Enter Sales Tax Payable calculated from R365").alignment = Alignment(horizontal='left')

                # Add DIFFERENCE calculation
                difference_row = r365_row + 2

                # Only add Toast - R365 difference
                diff_toast_r365_cell = ws.cell(row=difference_row, column=tax_toast_col, value='DIFFERENCE Toast - R365')
                diff_toast_r365_cell.border = thin_border

                value_cell = ws.cell(row=difference_row, column=tax_toast_col + 1, value="=G4-G6")  # Formula
                value_cell.border = thin_border
                value_cell.fill = yellow_fill  # Highlight with yellow
                value_cell.number_format = positive_format  # Default to positive format for formulas

        elif location in SOUTH_BEACH_LOCATIONS:
            # For South Beach, similar to North Beach but using its own values
            fdor_row = 1
            fdor_col = 6

            ws.merge_cells(start_row=fdor_row, start_column=fdor_col,
                          end_row=fdor_row, end_column=fdor_col + 1)
            header_cell = ws.cell(row=fdor_row, column=fdor_col, value='SALES TAX - FDOR')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=fdor_row, column=fdor_col + 1).border = thin_border

            # Calculate the 7% value correctly for South Beach
            tax_seven_percent = round(sales_reported * 0.07, 2)

            fdor_data = [
                ['NON TAXABLE', non_taxable],
                ['SALES REPORTED', sales_reported],
                ['7%', tax_seven_percent]
            ]
            for row_idx, (label, value) in enumerate(fdor_data, fdor_row + 1):
                ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # RESORT TAX table for South Beach
            resort_row = fdor_row + len(fdor_data) + 1

            ws.merge_cells(start_row=resort_row, start_column=fdor_col,
                          end_row=resort_row, end_column=fdor_col + 1)
            header_cell = ws.cell(row=resort_row, column=fdor_col, value='RESORT TAX')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=resort_row, column=fdor_col + 1).border = thin_border

            resort_non_taxable = non_taxable
            resort_sales_reported = net_sales_pnl
            resort_tax_two_percent = round((resort_sales_reported + resort_non_taxable) * 0.02, 2)

            resort_data = [
                ['NON TAXABLE - RESORT', resort_non_taxable],
                ['SALES REPORTED - RESORT', resort_sales_reported],
                ['2%', resort_tax_two_percent]
            ]

            for row_idx, (label, value) in enumerate(resort_data, resort_row + 1):
                ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # Total Tax table
            total_tax_row = resort_row + len(resort_data) + 1

            total_tax_value = tax_seven_percent + resort_tax_two_percent

            ws.cell(row=total_tax_row, column=fdor_col, value='TOTAL TAX').border = thin_border
            cell = ws.cell(row=total_tax_row, column=fdor_col + 1, value="=G4+G8")  # Formula
            cell.border = thin_border
            cell.number_format = positive_format

            # SALES TAX - Toast table for South Beach - with renamed Total Tax to Total Tax Toast
            tax_toast_row = total_tax_row + 2

            ws.merge_cells(start_row=tax_toast_row, start_column=fdor_col,
                          end_row=tax_toast_row, end_column=fdor_col + 1)
            header_cell = ws.cell(row=tax_toast_row, column=fdor_col, value='SALES TAX - Toast')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=tax_toast_row, column=fdor_col + 1).border = thin_border

            tax_toast_data = [
                ['Total Tax Toast', total_tax],
                ['UberEats Sales Tax', uber_tax],
                ['Tax Reported', tax_reported]
            ]

            for row_idx, (label, value) in enumerate(tax_toast_data, tax_toast_row + 1):
                ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format


            # Get R365 Sales Tax Payable and Resort Tax Payable for South Beach
            r365_sales_tax_payable = r365_sales_tax_data.get(gl_location, 0)
            resort_tax_payable = r365_resort_tax_data.get(gl_location, 0)
            total_tax_payable = r365_sales_tax_payable + resort_tax_payable

            # Add R365 table with both tax values
            r365_row = tax_toast_row + len(tax_toast_data) + 1

            ws.merge_cells(start_row=r365_row, start_column=fdor_col,
                        end_row=r365_row, end_column=fdor_col + 1)
            header_cell = ws.cell(row=r365_row, column=fdor_col, value='R365')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=r365_row, column=fdor_col + 1).border = thin_border

            # Set Sales Tax Payable value from GL
            ws.cell(row=r365_row + 1, column=fdor_col, value='Sales Tax Payable').border = thin_border
            sales_tax_cell = ws.cell(row=r365_row + 1, column=fdor_col + 1, value=r365_sales_tax_payable)
            sales_tax_cell.border = thin_border
            sales_tax_cell.number_format = positive_format if r365_sales_tax_payable >= 0 else negative_format

            # Add Resort Tax Payable row with value from GL
            ws.cell(row=r365_row + 2, column=fdor_col, value='Resort Tax Payable').border = thin_border
            resort_tax_cell = ws.cell(row=r365_row + 2, column=fdor_col + 1, value=resort_tax_payable)
            resort_tax_cell.border = thin_border
            resort_tax_cell.number_format = positive_format if resort_tax_payable >= 0 else negative_format

            # Calculate Total Tax Payable as sum of both values
            ws.cell(row=r365_row + 3, column=fdor_col, value='TOTAL Tax Payable').border = thin_border
            total_tax_cell = ws.cell(row=r365_row + 3, column=fdor_col + 1, value=total_tax_payable)
            total_tax_cell.border = thin_border
            total_tax_cell.number_format = positive_format

            # For the Differences table, use total_tax_payable instead of just sales tax
            diff_row = r365_row + 4

            ws.merge_cells(start_row=diff_row, start_column=fdor_col,
                        end_row=diff_row, end_column=fdor_col + 1)
            header_cell = ws.cell(row=diff_row, column=fdor_col, value='Differences')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=diff_row, column=fdor_col + 1).border = thin_border

            # Update difference calculations to use total_tax_payable
            diff_toast_r365 = tax_reported - total_tax_payable
            diff_excel_r365 = total_tax_value - total_tax_payable
            diff_toast_excel = tax_reported - total_tax_value

            # Round to 0 if absolute value is less than 0.01
            diff_toast_r365 = 0 if abs(diff_toast_r365) < 0.01 else diff_toast_r365
            diff_excel_r365 = 0 if abs(diff_excel_r365) < 0.01 else diff_excel_r365
            diff_toast_excel = 0 if abs(diff_toast_excel) < 0.01 else diff_toast_excel

            diff_data = [
                ['DIFFERENCE Toast - R365', diff_toast_r365],
                ['DIFFERENCE Excel - R365', diff_excel_r365],
                ['DIFFERENCE Toast - Excel', diff_toast_excel]
            ]

            for row_idx, (label, value) in enumerate(diff_data, diff_row + 1):
                ws.cell(row=row_idx, column=fdor_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=fdor_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # Add 15(d) value (1% of SALES REPORTED)
            fifteen_d_row = diff_row + len(diff_data) + 1
            fifteen_d_value = round(sales_reported * 0.01, 2)

            fifteen_d_cell = ws.cell(row=fifteen_d_row, column=fdor_col, value='15(d)')
            fifteen_d_cell.border = thin_border

            value_cell = ws.cell(row=fifteen_d_row, column=fdor_col + 1, value="=G3*0.01")  # Formula: 1% of SALES REPORTED
            value_cell.border = thin_border
            value_cell.number_format = positive_format

        elif location in NEW_YORK_LOCATIONS:
            # Special handling for New York locations
            tax_excel_row = 1
            tax_excel_col = 6  # Column F
            ws.merge_cells(start_row=tax_excel_row, start_column=tax_excel_col,
                            end_row=tax_excel_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=tax_excel_row, column=tax_excel_col, value='SALES TAX - Excel')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=tax_excel_row, column=tax_excel_col + 1).border = thin_border

            tax_excel_data = [
                ['NON TAXABLE', non_taxable],
                ['SALES REPORTED', sales_reported],
                ['8.875%', tax_amount]  # 8.875% for New York
            ]

            for row_idx, (label, value) in enumerate(tax_excel_data, tax_excel_row + 1):
                ws.cell(row=row_idx, column=tax_excel_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=tax_excel_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # Add Sales Tax - Toast table
            tax_toast_row = 6

            ws.merge_cells(start_row=tax_toast_row, start_column=tax_excel_col,
                            end_row=tax_toast_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=tax_toast_row, column=tax_excel_col, value='SALES TAX - Toast')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=tax_toast_row, column=tax_excel_col + 1).border = thin_border

            # For New York: renamed 'Total Tax' to 'Total Tax Toast'
            tax_toast_data = [
                ['Total Tax Toast', total_tax],
                ['Tax on Promotions', tax_on_promotions],
                ['Tax Reported', tax_reported]
            ]

            for row_idx, (label, value) in enumerate(tax_toast_data, tax_toast_row + 1):
                ws.cell(row=row_idx, column=tax_excel_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=tax_excel_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # Add R365 table
            r365_row = tax_toast_row + len(tax_toast_data) + 1

            ws.merge_cells(start_row=r365_row, start_column=tax_excel_col,
                            end_row=r365_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=r365_row, column=tax_excel_col, value='R365')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=r365_row, column=tax_excel_col + 1).border = thin_border

            # Just one row with Sales Tax Payable
            ws.cell(row=r365_row + 1, column=tax_excel_col, value='Sales Tax Payable').border = thin_border
            cell = ws.cell(row=r365_row + 1, column=tax_excel_col + 1, value=r365_sales_tax_payable)
            cell.border = thin_border
            cell.number_format = positive_format if r365_sales_tax_payable >= 0 else negative_format

            # Add the differences table
            diff_row = r365_row + 2

            ws.merge_cells(start_row=diff_row, start_column=tax_excel_col,
                            end_row=diff_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=diff_row, column=tax_excel_col, value='Differences')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=diff_row, column=tax_excel_col + 1).border = thin_border

            # Calculate differences for the two rows
            diff_toast_r365 = tax_reported - r365_sales_tax_payable
            diff_r365_excel = tax_amount - r365_sales_tax_payable  # Updated formula

            # Calculate additional difference: Tax Reported - tax_amount
            diff_toast_excel = tax_reported - tax_amount

            # Round to 0 if absolute value is less than 0.01
            diff_toast_r365 = 0 if abs(diff_toast_r365) < 0.01 else diff_toast_r365
            diff_r365_excel = 0 if abs(diff_r365_excel) < 0.01 else diff_r365_excel
            diff_toast_excel = 0 if abs(diff_toast_excel) < 0.01 else diff_toast_excel

            diff_data = [
                ['DIFFERENCE Toast - R365', diff_toast_r365],
                ['DIFFERENCE Excel - R365', diff_r365_excel],
                ['DIFFERENCE Toast - Excel', diff_toast_excel]
            ]

            for row_idx, (label, value) in enumerate(diff_data, diff_row + 1):
                ws.cell(row=row_idx, column=tax_excel_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=tax_excel_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # Add 15(d) value (1% of SALES REPORTED)
            fifteen_d_row = diff_row + len(diff_data) + 1
            fifteen_d_value = round(sales_reported * 0.01, 2)

            fifteen_d_cell = ws.cell(row=fifteen_d_row, column=tax_excel_col, value='15(d)')
            fifteen_d_cell.border = thin_border

            value_cell = ws.cell(row=fifteen_d_row, column=tax_excel_col + 1, value="=G3*0.01")  # Formula: 1% of SALES REPORTED
            value_cell.border = thin_border
            value_cell.number_format = positive_format

        else:  # Standard location handling for locations not in special groups
            # Add Sales Tax - Excel table
            tax_excel_row = 1
            tax_excel_col = 6  # Column F

            ws.merge_cells(start_row=tax_excel_row, start_column=tax_excel_col,
                            end_row=tax_excel_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=tax_excel_row, column=tax_excel_col, value='SALES TAX - Excel')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=tax_excel_row, column=tax_excel_col + 1).border = thin_border

            # For standard locations - fix tax calculation
            tax_excel_data = [
                ['NON TAXABLE', non_taxable],
                ['SALES REPORTED', sales_reported],
                ['7%', 0]  # Placeholder, will be calculated after
            ]

            for row_idx, (label, value) in enumerate(tax_excel_data, tax_excel_row + 1):
                ws.cell(row=row_idx, column=tax_excel_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=tax_excel_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # Calculate 7% based on NON TAXABLE + SALES REPORTED from this table
            tax_value = round((non_taxable + sales_reported) * 0.07, 2)
            tax_cell = ws.cell(row=tax_excel_row + 3, column=tax_excel_col + 1, value="=(G2+G3)*0.07")  # Formula
            tax_cell.number_format = positive_format

            # Add Sales Tax - Toast table
            tax_toast_row = 6

            ws.merge_cells(start_row=tax_toast_row, start_column=tax_excel_col,
                            end_row=tax_toast_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=tax_toast_row, column=tax_excel_col, value='SALES TAX - Toast')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=tax_toast_row, column=tax_excel_col + 1).border = thin_border

            # Renamed 'Total Tax' to 'Total Tax Toast'
            tax_toast_data = [
                ['Total Tax Toast', total_tax],
                ['UberEats Sales Tax', uber_tax],
                ['Tax Reported', tax_reported]
            ]

            for row_idx, (label, value) in enumerate(tax_toast_data, tax_toast_row + 1):
                ws.cell(row=row_idx, column=tax_excel_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=tax_excel_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

# Add R365 table
            r365_row = tax_toast_row + len(tax_toast_data) + 1

            ws.merge_cells(start_row=r365_row, start_column=tax_excel_col,
                            end_row=r365_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=r365_row, column=tax_excel_col, value='R365')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=r365_row, column=tax_excel_col + 1).border = thin_border

            # Just one row with Sales Tax Payable
            ws.cell(row=r365_row + 1, column=tax_excel_col, value='Sales Tax Payable').border = thin_border
            cell = ws.cell(row=r365_row + 1, column=tax_excel_col + 1, value=r365_sales_tax_payable)
            cell.border = thin_border
            cell.number_format = positive_format if r365_sales_tax_payable >= 0 else negative_format

            # Add the differences table
            diff_row = r365_row + 2

            ws.merge_cells(start_row=diff_row, start_column=tax_excel_col,
                            end_row=diff_row, end_column=tax_excel_col + 1)
            header_cell = ws.cell(row=diff_row, column=tax_excel_col, value='Differences')
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = header_fill
            header_cell.border = thin_border
            ws.cell(row=diff_row, column=tax_excel_col + 1).border = thin_border

            # Calculate differences
            diff_toast_r365 = tax_reported - r365_sales_tax_payable
            diff_r365_excel = tax_value - r365_sales_tax_payable  # Updated formula

            # Calculate additional difference: Tax Reported - tax_value
            diff_toast_excel = tax_reported - tax_value

            # Round to 0 if absolute value is less than 0.01
            diff_toast_r365 = 0 if abs(diff_toast_r365) < 0.01 else diff_toast_r365
            diff_r365_excel = 0 if abs(diff_r365_excel) < 0.01 else diff_r365_excel
            diff_toast_excel = 0 if abs(diff_toast_excel) < 0.01 else diff_toast_excel

            diff_data = [
                ['DIFFERENCE Toast - R365', diff_toast_r365],
                ['DIFFERENCE Excel - R365', diff_r365_excel],
                ['DIFFERENCE Toast - Excel', diff_toast_excel]
            ]

            for row_idx, (label, value) in enumerate(diff_data, diff_row + 1):
                ws.cell(row=row_idx, column=tax_excel_col, value=label).border = thin_border
                cell = ws.cell(row=row_idx, column=tax_excel_col + 1, value=value)
                cell.border = thin_border
                cell.number_format = positive_format if value >= 0 else negative_format

            # Add 15(d) value (1% of SALES REPORTED)
            fifteen_d_row = diff_row + len(diff_data) + 1
            fifteen_d_value = round(sales_reported * 0.01, 2)

            fifteen_d_cell = ws.cell(row=fifteen_d_row, column=tax_excel_col, value='15(d)')
            fifteen_d_cell.border = thin_border

            value_cell = ws.cell(row=fifteen_d_row, column=tax_excel_col + 1, value="=G3*0.01")  # Formula: 1% of SALES REPORTED
            value_cell.border = thin_border
            value_cell.number_format = positive_format

        # Add Royalty Fees section
        # Determine royalty_start_row more dynamically for New York locations
        if location in CARROT_LOVE_LOCATIONS:
            if location == "North Beach":
                royalty_start_row = fifteen_d_row + 2
            else:
                # Ensure we have enough rows for main data + leadership table in Aventura and Coral Gables
                main_data_end = len(main_data) + 2  # +2 for headers
                leadership_table_end = leadership_row + 2  # +2 for headers and values
                tax_toast_end = fifteen_d_row + 1
                royalty_start_row = max(main_data_end, leadership_table_end, tax_toast_end) + 2
        elif location in SOUTH_BEACH_LOCATIONS:
            royalty_start_row = fifteen_d_row + 2
        elif location in NEW_YORK_LOCATIONS:
            # For New York locations, ensure enough space for extended main data table
            main_data_end = len(main_data) + 2  # +2 for headers
            leadership_table_end = leadership_row + 2  # +2 for headers and values
            tax_tables_end = fifteen_d_row + 1
            royalty_start_row = max(main_data_end, leadership_table_end, tax_tables_end) + 4  # Add more spacing
        else:
            royalty_start_row = fifteen_d_row + 2

        ws.merge_cells(start_row=royalty_start_row, start_column=1, end_row=royalty_start_row, end_column=4)
        header_cell = ws.cell(row=royalty_start_row, column=1, value='Royalty Fees')
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.font = Font(bold=True)
        header_cell.fill = header_fill
        for col in range(1, 5):
            ws.cell(row=royalty_start_row, column=col).border = thin_border

        royalty_headers = ['Metric', 'Sales', 'Royalty Fee 5%', 'Royalty Fee 3%']
        for col, header in enumerate(royalty_headers, 1):
            cell = ws.cell(row=royalty_start_row + 1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border

        uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount

        # Special handling for New York locations
        if location in NEW_YORK_LOCATIONS:
            # Get additional metrics for New York
            grubhub_toast = grubhub_sales.get(location, 0)
            grubhub_r365 = additional_metrics['grubhub_sales'].get(gl_location, 0)
            doordash_sales = additional_metrics['doordash_sales'].get(gl_location, 0)
            doordash_refunds = additional_metrics['doordash_refunds'].get(gl_location, 0)
            doordash_discounts = additional_metrics['doordash_discounts'].get(gl_location, 0)
            grubhub_refunds = additional_metrics['grubhub_refunds'].get(gl_location, 0)
            grubhub_delivery_fees = additional_metrics['grubhub_delivery_fees'].get(gl_location, 0)
            grubhub_promotions = additional_metrics['grubhub_promotions'].get(gl_location, 0)
            third_parties = additional_metrics['third_parties'].get(location, 0)

            main_data = [
                ['Toast Net Sales', toast_sales],
                ['UberEats Sales - Toast', uber_sales_toast],
                ['UberEats Sales - R365', uber_sales_r365],
                ['Delivery Fee Income', delivery_fee],
                ['Ez Catering', ez_catering],
                ['UberEats Refunds', uber_refunds],
                ['UberEats Discount', uber_discount],
                ['Grubhub Toast', grubhub_toast],
                ['Grubhub R365', grubhub_r365],
                ['DoorDash Sales', doordash_sales],  # Changed from 'DoorDash' to 'DoorDash Sales'
                ['DoorDash Refunds', doordash_refunds],
                ['Grubhub Delivery Fees', grubhub_delivery_fees],
                ['Grubhub Promotions', grubhub_promotions],
                ['Grubhub Refunds', grubhub_refunds],
                ['Third Parties', third_parties],
                ['Net Sales PNL', net_sales_pnl]
            ]

        # For NY locations, calculate the royalty section
        if location in NEW_YORK_LOCATIONS:
            # Calculate Grubhub total for royalty
            grubhub_r365 = additional_metrics['grubhub_sales'].get(gl_location, 0)
            grubhub_promotions = additional_metrics['grubhub_promotions'].get(gl_location, 0)
            grubhub_refunds = additional_metrics['grubhub_refunds'].get(gl_location, 0)
            grubhub_sales_total = grubhub_r365 + grubhub_promotions + grubhub_refunds

            # Calculate DoorDash total
            doordash_sales = additional_metrics['doordash_sales'].get(gl_location, 0)
            doordash_refunds = additional_metrics['doordash_refunds'].get(gl_location, 0)
            doordash_total = doordash_sales + doordash_refunds
            doordash_discounts = additional_metrics['doordash_discounts'].get(gl_location, 0)

            # Get third parties
            third_parties_value = additional_metrics['third_parties'].get(location, 0)

            # NEW APPROACH: Find indices by name instead of tuple matching
            # Find the index of the delivery fee row by searching for the correct first element
            delivery_fee_index = -1
            grubhub_fees_index = -1

            for i, row in enumerate(main_data):
                if row[0] == 'Delivery Fee Income':
                    delivery_fee_index = i + 2  # +2 for Excel row number
                elif row[0] == 'Grubhub Delivery Fees':
                    grubhub_fees_index = i + 2


            # Safety check to make sure we found the indices
            if delivery_fee_index == -1:
                print(f"Warning: 'Delivery Fee Income' not found in main_data for {location}")
                delivery_fee_index = 5  # Fallback to a reasonable guess

            if grubhub_fees_index == -1 and location != "Bryant Park":
                print(f"Warning: 'Grubhub Delivery Fees' not found in main_data for {location}")
                grubhub_fees_index = 12  # Fallback to a reasonable guess

            # Calculate formula references for Sales w/o 3rd Party deliveries
            if location == "Bryant Park":
                # For Bryant Park: Net Sales PNL - Delivery Fee - UberEats - Grubhub - Third Parties - DoorDash (including discounts)
                sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - grubhub_sales_total - (doordash_sales + doordash_refunds + doordash_discounts) - third_parties
            else:
                # For other NY locations: delivery_fee already includes grubhub_delivery_fees, so DON'T subtract it again
                sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - grubhub_sales_total - (doordash_sales + doordash_refunds + doordash_discounts) - third_parties

            # Now find the indices for the other values using the same approach
            uber_sales_r365_index = -1
            uber_refunds_index = -1
            uber_discount_index = -1
            grubhub_r365_index = -1
            grubhub_promotions_index = -1
            grubhub_refunds_index = -1
            third_parties_index = -1
            doordash_sales_index = -1
            doordash_refunds_index = -1
            doordash_discounts_index = -1

            for i, row in enumerate(main_data):
                if row[0] == 'UberEats Sales - R365':
                    uber_sales_r365_index = i + 2
                elif row[0] == 'UberEats Refunds':
                    uber_refunds_index = i + 2
                elif row[0] == 'UberEats Discount':
                    uber_discount_index = i + 2
                elif row[0] == 'Grubhub R365':
                    grubhub_r365_index = i + 2
                elif row[0] == 'Grubhub Promotions':
                    grubhub_promotions_index = i + 2
                elif row[0] == 'Grubhub Refunds':
                    grubhub_refunds_index = i + 2
                elif row[0] == 'Third Parties':
                    third_parties_index = i + 2
                elif row[0] == 'DoorDash Sales':
                    doordash_sales_index = i + 2
                elif row[0] == 'DoorDash Refunds':
                    doordash_refunds_index = i + 2
                elif row[0] == 'DoorDash Discount':
                    doordash_discounts_index = i + 2

            # Update royalty data - Third Parties now includes DoorDash
            # Set up hardcoded formulas for NY locations
            if location == "Bryant Park":
                sales_wo_3rd_formula = "=B18-B5-B27-B28-B29"  # Hard-coded for Bryant Park
            else:
                sales_wo_3rd_formula = "=B18-B14-B5-B27-B28-B29"  # Hard-coded for other NY locations

            royalty_data = [
                ['Sales w/o 3rd Party deliveries', sales_wo_3rd_formula, f"=B{royalty_start_row+2}*0.05", 0],
                ['UberEats', "=B4+B7+B8", 0, f"=B{royalty_start_row+3}*0.03"],
                ['Grubhub', "=B10+B16+B15", 0, f"=B{royalty_start_row+4}*0.03"],
                ['Third Parties', "=B17+B11+B12+B13", 0, f"=B{royalty_start_row+5}*0.03"],
                ['Totals', f"=SUM(B{royalty_start_row+2}:B{royalty_start_row+5})", f"=SUM(C{royalty_start_row+2}:C{royalty_start_row+5})", f"=SUM(D{royalty_start_row+2}:D{royalty_start_row+5})"]
            ]
        else:
            # Standard royalty calculation for other locations
            # Find indices for needed rows
            delivery_fee_index = -1
            for i, row in enumerate(main_data):
                if row[0] == 'Delivery Fee Income':
                    delivery_fee_index = i + 2  # +2 for Excel row number
                    break

            if delivery_fee_index == -1:
                delivery_fee_index = 5  # Fallback to a reasonable guess

            # Updated formula for non-NY locations: Net Sales PNL - Delivery Fee - UberEats - Ezcater
            sales_wo_3rd_formula = f"=B{len(main_data)+1}-B{delivery_fee_index}-B{royalty_start_row+3}-B{royalty_start_row+4}"

            royalty_data = [
                ['Sales w/o 3rd Party deliveries', sales_wo_3rd_formula, f"=B{royalty_start_row+2}*0.05", 0],
                ['UberEats', f"=B{main_data.index(['UberEats Sales - R365', uber_sales_r365])+2}+B{main_data.index(['UberEats Refunds', uber_refunds])+2}+B{main_data.index(['UberEats Discount', uber_discount])+2}", 0, f"=B{royalty_start_row+3}*0.03"],
                ['Ezcater', f"=B{main_data.index(['Ez Catering', ez_catering])+2}", 0, f"=B{royalty_start_row+4}*0.03"],
                ['Totals', f"=SUM(B{royalty_start_row+2}:B{royalty_start_row+4})", f"=SUM(C{royalty_start_row+2}:C{royalty_start_row+4})", f"=SUM(D{royalty_start_row+2}:D{royalty_start_row+4})"]
            ]

        for row_idx, row_data in enumerate(royalty_data, royalty_start_row + 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col in [2, 3, 4]:  # Amount columns
                    if isinstance(value, (int, float)):
                        cell.number_format = positive_format if value >= 0 else negative_format
                    else:  # It's a formula or other string
                        cell.number_format = positive_format  # Default to positive format for formulas

        # Add total royalty fees row with yellow highlight
        total_row = royalty_start_row + len(royalty_data) + 2
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)

        total_cell = ws.cell(row=total_row, column=1, value='ROYALTY FEES')
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
        total_cell.fill = yellow_fill
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=2).fill = yellow_fill

        # The final royalty fee is the sum of 5% and 3% totals
        last_royalty_row = royalty_start_row + len(royalty_data) + 1
        amount_cell = ws.cell(row=total_row, column=3, value=f"=C{last_royalty_row}+D{last_royalty_row}")
        amount_cell.number_format = positive_format
        amount_cell.border = thin_border
        amount_cell.fill = yellow_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 180 / 7  # Convert pixels to Excel width units
        ws.column_dimensions['F'].width = 220 / 7  # Same width as column A
        ws.column_dimensions['G'].width = 90 / 7   # Half width of column F
        for col in range(2, 6):  # B through E columns
            ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(output_path)


def generate_ar_invoices(sales_data, uber_sales, profit_metrics, additional_metrics, export_data, r365_sales_tax_data, earliest_date, latest_date, output_dir):
    """
    Generate AR invoices for each location.
    Two invoices per location: one from Carrot Leadership LLC and one from Carrot Express Franchise System LLC
    (with Midtown having an additional invoice from Carrot Leadership LLC)
    """
    import csv
    import os
    from datetime import datetime

    # Add this dictionary to translate location names to vendor names
    VENDOR_NAME_DICT = {
        'Aventura (Miami Gardens)':'Carrot Love Aventura LLC',
        'Aventura Mall':'Carrot Love Aventura Mall Operating LLC',
        'Boca Palmetto Park':'Carrot Love Palmetto Park Op LLC',
        'Brickell':'Carrot Love Brickell Operating LLC',
        'Bryant Park':'Carrot Love Bryant Park Operating LLC',
        'Coconut Creek':'Carrot Love Coconut Creek Operating  LLC',
        'Coconut Grove':'Carrot Love Coconut Grove Operating LLC',
        'Coral Gables':'Carrot Love Coral Gables',
        'Dadeland':'Carrot Love Dadeland Operating LLC',
        'Doral':'Carrot Love City Place Doral Operating LLC',
        'Downtown':'Carrot Love Two LLC',
        'Flatiron':'Carrot Love Manhattan Operating LLC',
        'Hollywood':'Carrot Love Hollywood Operating LLC',
        'Las Olas':'Carrot Love Las Olas Operating LLC',
        'Lexington':'Carrot Love Lexington 52 LLC',
        'Miami Shores':'Carrot Express Miami Shores LLC',
        'Midtown':'Carrot Express Midtown LLC',
        'North Beach':'Carrot Love North Beach',
        'Pembroke Pines':'Carrot Love Pembroke Pines Operating LLC',
        'Plantation':'Carrot Love Plantation Operating LLC',
        'River Landing':'Carrot Love River Landing Op LLC',
        'South Beach':'Carrot Love South Florida Operating C LLC',
        'South Miami (Sunset)':'Carrot Love Sunset Operating LLC',
        'West Boca':'Carrot Love West Boca Operating LLC'
    }

    # Format dates
    start_date_str = earliest_date.strftime('%m/%d/%Y')
    end_date_str = latest_date.strftime('%m/%d/%Y')
    date_range_str = f"{start_date_str} - {end_date_str}"

    # For invoice numbers (mmddyyyy format)
    date_suffix = latest_date.strftime('%m%d%Y')

    # Initialize counters for invoice numbers
    cefs_counter = 1  # Counter for Carrot Express Franchise System LLC
    clead_counter = 1  # Counter for Carrot Leadership LLC

    # Prepare output file
    output_filename = f"AR_Invoices_Fees_{date_suffix}.csv"
    output_path = os.path.join(output_dir, output_filename)

    # Define fieldnames based on the example
    fieldnames = [
        'Type', 'Location', 'Vendor', 'Number', 'Date', 'Gl Date', 'Amount',
        'Payment Terms', 'Due Date', 'Comment', 'Detail Location', 'Detail Comment',
        'Detail Account', 'Detail Amount'
    ]

    invoices = []

    # For each location, generate two invoices (three for Midtown)
    for location, toast_sales in sales_data.items():
        gl_location = LOCATION_DICT.get(location, location)

        # Skip locations that don't have royalty info calculated or West Boca
        if location not in LOCATION_DICT.keys() or location == "West Boca":
            continue

        # Get vendor name - this should be the AP Location from the left side of the image
        vendor_name = VENDOR_NAME_DICT.get(location, location)  # Use the dictionary to get the vendor name

        # Get necessary data for calculations
        net_sales_pnl = export_data.get(gl_location, 0)
        delivery_fee = profit_metrics['delivery_fee'].get(gl_location, 0)
        uber_sales_r365 = additional_metrics['ue_sales'].get(gl_location, 0)
        uber_refunds = additional_metrics['ue_refunds'].get(gl_location, 0)
        uber_discount = additional_metrics['ue_discount'].get(gl_location, 0)
        ez_catering = profit_metrics['ez_catering'].get(gl_location, 0)

        # Calculate cleadership fee
        if location in NEW_YORK_LOCATIONS and location != "Bryant Park":
            # For NY locations, only subtract grubhub_delivery_fees once
            # Don't subtract grubhub_delivery_fees if it's already included in delivery_fee
            grubhub_delivery_fees = additional_metrics['grubhub_delivery_fees'].get(gl_location, 0)

            # Check if we're using Bryant Park special case
            delivery_fee_includes_grubhub = True  # Set to True as the delivery_fee likely already includes grubhub fees

            if delivery_fee_includes_grubhub:
                cleadership_fee = round((net_sales_pnl - delivery_fee) * 0.02, 2)
            else:
                cleadership_fee = round((net_sales_pnl - delivery_fee - grubhub_delivery_fees) * 0.02, 2)
        elif location in PLANTATION_LOCATIONS:
            # Special handling for Plantation - include plantation_payable in delivery fee
            plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
            cleadership_fee = round((net_sales_pnl - (delivery_fee + plantation_payable)) * 0.02, 2)
        else:
            # Standard calculation for all other locations
            cleadership_fee = round((net_sales_pnl - delivery_fee) * 0.02, 2)

        # Calculate leadership fee for Midtown only (not Miami Shores)
        if location == "Midtown":
            leadership_fee = round((uber_sales_r365 + uber_refunds + uber_discount) * 0.01, 2)
        else:
            leadership_fee = 0

        # Create franchise invoice for all locations EXCEPT Midtown
        if location != "Midtown":
            # Calculate royalty values based on location type
            uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount

            if location in NEW_YORK_LOCATIONS:
                # New York location calculations
                grubhub_r365 = additional_metrics['grubhub_sales'].get(gl_location, 0)
                grubhub_promotions = additional_metrics['grubhub_promotions'].get(gl_location, 0)
                grubhub_refunds = additional_metrics['grubhub_refunds'].get(gl_location, 0)
                doordash_sales = additional_metrics['doordash_sales'].get(gl_location, 0)
                doordash_refunds = additional_metrics['doordash_refunds'].get(gl_location, 0)
                doordash_discounts = additional_metrics['doordash_discounts'].get(gl_location, 0)
                third_parties = additional_metrics['third_parties'].get(location, 0)

                # Calculate values matching the report exactly
                uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount
                grubhub_sales_total = grubhub_r365 + grubhub_promotions + grubhub_refunds

                # For Bryant Park, we don't need to worry about grubhub_delivery_fees
                if location == "Bryant Park":
                    # For Bryant Park: Net Sales PNL - Delivery Fee - UberEats - Grubhub - Third Parties - DoorDash (including discounts)
                    sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - grubhub_sales_total - (doordash_sales + doordash_refunds + doordash_discounts) - third_parties
                else:
                    # For other NY locations: delivery_fee already includes grubhub_delivery_fees, so DON'T subtract it again
                    sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - grubhub_sales_total - (doordash_sales + doordash_refunds + doordash_discounts) - third_parties

                # Calculate royalty components
                royalty_5_percent = sales_wo_3rd * 0.05
                royalty_3_percent_uber = uber_sales_total * 0.03
                royalty_3_percent_grubhub = grubhub_sales_total * 0.03
                royalty_3_percent_third = (third_parties + doordash_sales + doordash_refunds + doordash_discounts) * 0.03
                total_royalty = round(royalty_5_percent + royalty_3_percent_uber + royalty_3_percent_grubhub + royalty_3_percent_third, 2)
            else:
                # Standard location calculations
                uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount

                # For Plantation, add plantation payable to delivery fee
                if location in PLANTATION_LOCATIONS:
                    plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
                    delivery_fee += plantation_payable

                # Calculate Sales w/o 3rd Party deliveries exactly as in royalties report
                sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - ez_catering

                # Calculate royalty components
                royalty_5_percent = sales_wo_3rd * 0.05
                royalty_3_percent_uber = uber_sales_total * 0.03
                royalty_3_percent_ez = ez_catering * 0.03

                total_royalty = round(royalty_5_percent + royalty_3_percent_uber + royalty_3_percent_ez, 2)

            # Create Carrot Express Franchise System LLC invoice
            cefs_invoice_num = f"AR-CEFS{date_suffix}-{cefs_counter:02d}"
            cefs_counter += 1

            cefs_invoice = {
                'Type': 'AR Invoice',
                'Location': 'Carrot Express Franchise System LLC',
                'Vendor': vendor_name,
                'Number': cefs_invoice_num,
                'Date': end_date_str,
                'Gl Date': end_date_str,
                'Amount': total_royalty,
                'Payment Terms': 'Due Upon Receipt',
                'Due Date': end_date_str,
                'Comment': f"Royalty Fees // {date_range_str}",
                'Detail Location': 'Carrot Express Franchise System LLC',
                'Detail Comment': f"Royalty Fees // {date_range_str}",
                'Detail Account': 'Royalty Fees',
                'Detail Amount': total_royalty
            }
            invoices.append(cefs_invoice)

        # INVOICE 2: Carrot Leadership LLC (for Cleadership 2% Fee) - for ALL locations
        clead_invoice_num = f"AR-CLEAD{date_suffix}-{clead_counter:02d}"
        clead_counter += 1

        clead_invoice = {
            'Type': 'AR Invoice',
            'Location': 'Carrot Leadership LLC',
            'Vendor': vendor_name,
            'Number': clead_invoice_num,
            'Date': end_date_str,
            'Gl Date': end_date_str,
            'Amount': cleadership_fee,
            'Payment Terms': 'Due Upon Receipt',
            'Due Date': end_date_str,
            'Comment': f"Cleadership - 2% Fee // {date_range_str}",
            'Detail Location': 'Carrot Leadership LLC',
            'Detail Comment': f"Cleadership - 2% Fee // {date_range_str}",
            'Detail Account': 'Admnistrative Fee Revenue',
            'Detail Amount': cleadership_fee
        }
        invoices.append(clead_invoice)

        # INVOICE 3: Additional Carrot Leadership LLC invoice for Midtown only (for Leadership 1% Fee)
        if location == "Midtown":  # Changed from ["Midtown", "Miami Shores"] to just "Midtown"
            midtown_clead_invoice_num = f"AR-CLEAD{date_suffix}-{clead_counter:02d}"
            clead_counter += 1

            midtown_clead_invoice = {
                'Type': 'AR Invoice',
                'Location': 'Carrot Leadership LLC',
                'Vendor': vendor_name,
                'Number': midtown_clead_invoice_num,
                'Date': end_date_str,
                'Gl Date': end_date_str,
                'Amount': leadership_fee,
                'Payment Terms': 'Due Upon Receipt',
                'Due Date': end_date_str,
                'Comment': f"Leadership - 1% Fee // {date_range_str}",
                'Detail Location': 'Carrot Leadership LLC',
                'Detail Comment': f"Leadership - 1% Fee // {date_range_str}",
                'Detail Account': 'UberEats Rebate/Bonus',
                'Detail Amount': leadership_fee
            }
            invoices.append(midtown_clead_invoice)

    # Write all invoices to CSV file - ensure all amounts are rounded to 2 decimal places
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for invoice in invoices:
            # Ensure amounts are properly rounded
            if 'Amount' in invoice:
                invoice['Amount'] = round(invoice['Amount'], 2)
            if 'Detail Amount' in invoice:
                invoice['Detail Amount'] = round(invoice['Detail Amount'], 2)
            writer.writerow(invoice)

    return output_path

def generate_ap_invoices(sales_data, uber_sales, profit_metrics, additional_metrics, export_data, r365_sales_tax_data, earliest_date, latest_date, output_dir):
    """
    Generate AP invoices based on AR invoices, with switched vendor/location and mapped account names.
    """
    import csv
    import os
    from datetime import datetime

    # Format dates
    start_date_str = earliest_date.strftime('%m/%d/%Y')
    end_date_str = latest_date.strftime('%m/%d/%Y')
    date_range_str = f"{start_date_str} - {end_date_str}"

    # For invoice numbers (mmddyyyy format)
    date_suffix = latest_date.strftime('%m%d%Y')

    # Initialize counters for invoice numbers - must match AR invoice numbers
    cefs_counter = 1  # Counter for Carrot Express Franchise System LLC
    clead_counter = 1  # Counter for Carrot Leadership LLC

    # Prepare output file
    output_filename = f"AP_Invoices_Fees_{date_suffix}.csv"
    output_path = os.path.join(output_dir, output_filename)

    # Define fieldnames based on the example
    fieldnames = [
        'Type', 'Location', 'Vendor', 'Number', 'Date', 'Gl Date', 'Amount',
        'Payment Terms', 'Due Date', 'Comment', 'Detail Location', 'Detail Comment',
        'Detail Account', 'Detail Amount'
    ]

    invoices = []

    # For each location, generate two invoices (three for Midtown)
    for location, toast_sales in sales_data.items():
        gl_location = LOCATION_DICT.get(location, location)

        # Skip locations that don't have royalty info calculated or West Boca
        if location not in LOCATION_DICT.keys() or location == "West Boca":
            continue

        # Get necessary data for calculations
        net_sales_pnl = export_data.get(gl_location, 0)
        delivery_fee = profit_metrics['delivery_fee'].get(gl_location, 0)
        uber_sales_r365 = additional_metrics['ue_sales'].get(gl_location, 0)
        uber_refunds = additional_metrics['ue_refunds'].get(gl_location, 0)
        uber_discount = additional_metrics['ue_discount'].get(gl_location, 0)
        ez_catering = profit_metrics['ez_catering'].get(gl_location, 0)

        # Calculate cleadership fee
        if location in NEW_YORK_LOCATIONS and location != "Bryant Park":
            # For NY locations, only subtract grubhub_delivery_fees once
            # Don't subtract grubhub_delivery_fees if it's already included in delivery_fee
            grubhub_delivery_fees = additional_metrics['grubhub_delivery_fees'].get(gl_location, 0)

            # Check if we're using Bryant Park special case
            delivery_fee_includes_grubhub = True  # Set to True as the delivery_fee likely already includes grubhub fees

            if delivery_fee_includes_grubhub:
                cleadership_fee = round((net_sales_pnl - delivery_fee) * 0.02, 2)
            else:
                cleadership_fee = round((net_sales_pnl - delivery_fee - grubhub_delivery_fees) * 0.02, 2)
        elif location in PLANTATION_LOCATIONS:
            # Special handling for Plantation - include plantation_payable in delivery fee
            plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
            cleadership_fee = round((net_sales_pnl - (delivery_fee + plantation_payable)) * 0.02, 2)
        else:
            # Standard calculation for all other locations
            cleadership_fee = round((net_sales_pnl - delivery_fee) * 0.02, 2)

        # Calculate leadership fee for Midtown only (not Miami Shores)
        if location == "Midtown":
            leadership_fee = round((uber_sales_r365 + uber_refunds + uber_discount) * 0.01, 2)
        else:
            leadership_fee = 0

        # Create franchise invoice for all locations EXCEPT Midtown
        if location != "Midtown":
            # Calculate royalty values based on location type
            uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount

            if location in NEW_YORK_LOCATIONS:
                # New York location calculations
                grubhub_r365 = additional_metrics['grubhub_sales'].get(gl_location, 0)
                grubhub_promotions = additional_metrics['grubhub_promotions'].get(gl_location, 0)
                grubhub_refunds = additional_metrics['grubhub_refunds'].get(gl_location, 0)
                doordash_sales = additional_metrics['doordash_sales'].get(gl_location, 0)
                doordash_refunds = additional_metrics['doordash_refunds'].get(gl_location, 0)
                doordash_discounts = additional_metrics['doordash_discounts'].get(gl_location, 0)
                third_parties = additional_metrics['third_parties'].get(location, 0)

                # Calculate values matching the report exactly
                uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount
                grubhub_sales_total = grubhub_r365 + grubhub_promotions + grubhub_refunds

                # For Bryant Park, we don't need to worry about grubhub_delivery_fees
                if location == "Bryant Park":
                    # For Bryant Park: Net Sales PNL - Delivery Fee - UberEats - Grubhub - Third Parties - DoorDash (including discounts)
                    sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - grubhub_sales_total - (doordash_sales + doordash_refunds + doordash_discounts) - third_parties
                else:
                    # For other NY locations: delivery_fee already includes grubhub_delivery_fees, so DON'T subtract it again
                    sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - grubhub_sales_total - (doordash_sales + doordash_refunds + doordash_discounts) - third_parties

                # Calculate royalty components
                royalty_5_percent = sales_wo_3rd * 0.05
                royalty_3_percent_uber = uber_sales_total * 0.03
                royalty_3_percent_grubhub = grubhub_sales_total * 0.03
                royalty_3_percent_third = (third_parties + doordash_sales + doordash_refunds + doordash_discounts) * 0.03
                total_royalty = round(royalty_5_percent + royalty_3_percent_uber + royalty_3_percent_grubhub + royalty_3_percent_third, 2)
            else:
                # Standard location calculations
                uber_sales_total = uber_sales_r365 + uber_refunds + uber_discount

                # For Plantation, add plantation payable to delivery fee
                if location in PLANTATION_LOCATIONS:
                    plantation_payable = r365_sales_tax_data.get("Plantation_Payable", 0)
                    delivery_fee += plantation_payable

                # Calculate Sales w/o 3rd Party deliveries exactly as in royalties report
                sales_wo_3rd = net_sales_pnl - delivery_fee - uber_sales_total - ez_catering

                # Calculate royalty components
                royalty_5_percent = sales_wo_3rd * 0.05
                royalty_3_percent_uber = uber_sales_total * 0.03
                royalty_3_percent_ez = ez_catering * 0.03

                total_royalty = round(royalty_5_percent + royalty_3_percent_uber + royalty_3_percent_ez, 2)

            # Create AP invoice - using the same number as corresponding AR invoice
            cefs_invoice_num = f"AR-CEFS{date_suffix}-{cefs_counter:02d}"
            cefs_counter += 1

            ap_cefs_invoice = {
                'Type': 'AP Invoice',
                'Location': gl_location,  # Use R365 location name (right side of dictionary)
                'Vendor': 'Carrot Express Franchise System LLC',  # Switch from Location to Vendor
                'Number': cefs_invoice_num,  # Same number as AR invoice
                'Date': end_date_str,
                'Gl Date': end_date_str,
                'Amount': total_royalty,
                'Payment Terms': 'Due Upon Receipt',
                'Due Date': end_date_str,
                'Comment': f"Royalty Fees // {date_range_str}",
                'Detail Location': gl_location,  # Match the Location field
                'Detail Comment': f"Royalty Fees // {date_range_str}",
                'Detail Account': 'Frachisee Fees',  # Changed from "Royalty Fees"
                'Detail Amount': total_royalty
            }
            invoices.append(ap_cefs_invoice)

        # INVOICE 2: AP Invoice corresponding to Carrot Leadership LLC (for Cleadership 2% Fee)
        clead_invoice_num = f"AR-CLEAD{date_suffix}-{clead_counter:02d}"
        clead_counter += 1

        ap_clead_invoice = {
            'Type': 'AP Invoice',
            'Location': gl_location,  # Use R365 location name
            'Vendor': 'Carrot Leadership LLC',  # Switch from Location to Vendor
            'Number': clead_invoice_num,  # Same number as AR invoice
            'Date': end_date_str,
            'Gl Date': end_date_str,
            'Amount': cleadership_fee,
            'Payment Terms': 'Due Upon Receipt',
            'Due Date': end_date_str,
            'Comment': f"Cleadership - 2% Fee // {date_range_str}",
            'Detail Location': gl_location,  # Match the Location field
            'Detail Comment': f"Cleadership - 2% Fee // {date_range_str}",
            'Detail Account': 'Management Fees',  # Changed from "Admnistrative Fee Revenue"
            'Detail Amount': cleadership_fee
        }
        invoices.append(ap_clead_invoice)

        # INVOICE 3: Additional AP invoice for Midtown only (for Leadership 1% Fee)
        if location == "Midtown":  # Changed from ["Midtown", "Miami Shores"] to just "Midtown"
            midtown_clead_invoice_num = f"AR-CLEAD{date_suffix}-{clead_counter:02d}"
            clead_counter += 1

            ap_midtown_clead_invoice = {
                'Type': 'AP Invoice',
                'Location': gl_location,  # Use R365 location name
                'Vendor': 'Carrot Leadership LLC',  # Switch from Location to Vendor
                'Number': midtown_clead_invoice_num,  # Same number as AR invoice
                'Date': end_date_str,
                'Gl Date': end_date_str,
                'Amount': leadership_fee,
                'Payment Terms': 'Due Upon Receipt',
                'Due Date': end_date_str,
                'Comment': f"Leadership - 1% Fee // {date_range_str}",
                'Detail Location': gl_location,  # Match the Location field
                'Detail Comment': f"Leadership - 1% Fee // {date_range_str}",
                'Detail Account': 'Advertising & Promotion',  # Changed from "UberEats Rebate/Bonus"
                'Detail Amount': leadership_fee
            }
            invoices.append(ap_midtown_clead_invoice)

    # Write all invoices to CSV file - ensure all amounts are rounded to 2 decimal places
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for invoice in invoices:
            # Ensure amounts are properly rounded
            if 'Amount' in invoice:
                invoice['Amount'] = round(invoice['Amount'], 2)
            if 'Detail Amount' in invoice:
                invoice['Detail Amount'] = round(invoice['Detail Amount'], 2)
            writer.writerow(invoice)

    return output_path

def add_tax_tab_and_reorder(sales_data, order_data, tax_exempt_data, profit_metrics, earliest_date, latest_date, output_dir):
    """
    Adds a Tax tab to the royalties summary workbook and reorders tabs according to the specified order.
    """
    import os
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    # Determine the filename of the existing royalties summary file
    royalties_filename = f"Royalties_Summary_{earliest_date.strftime('%m%d%Y')}-{latest_date.strftime('%m%d%Y')}.xlsx"
    royalties_path = os.path.join(output_dir, royalties_filename)

    # Load the workbook
    wb = load_workbook(royalties_path)

    # Create the Tax tab
    tax_ws = wb.create_sheet("Tax")

    # Set up styles
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Set column widths
    tax_ws.column_dimensions['A'].width = 30
    tax_ws.column_dimensions['B'].width = 15

    # Set header row
    headers = ['Store', 'Vendor', 'Invoice Number', 'Invoice Date', 'GL Date',
               'Due Date', 'Total', 'Comments_1', 'Comments_2', 'Approved Payment Date', 'Payment Type']

    for col, header in enumerate(headers, 1):
        cell = tax_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border

    # Get all location names from the workbook sheets
    locations = []
    north_beach_tax_value = 0
    south_beach_sales_tax_value = 0
    south_beach_resort_tax_value = 0

    # First, collect all tax values and get the North Beach value
    tax_values = {}
    for sheet_name in wb.sheetnames:
        if sheet_name != 'Tax':  # Skip the new Tax sheet
            locations.append(sheet_name)
            if sheet_name == 'North Beach':
                # Find the Sales Tax Payable value
                ws = wb[sheet_name]
                found = False
                for row in range(1, 50):  # Search through rows
                    tax_label_cell = ws.cell(row=row, column=6)
                    if tax_label_cell.value == 'Sales Tax Payable':
                        north_beach_tax_value = ws.cell(row=row, column=7).value
                        tax_values[sheet_name] = north_beach_tax_value  # Use actual value for North Beach
                        found = True
                        break
                    # Also check for Carrot Love Sales Tax Payable
                    elif tax_label_cell.value == 'Carrot Love TOTAL Sales Tax Payable':
                        north_beach_tax_value = ws.cell(row=row, column=7).value
                        tax_values[sheet_name] = north_beach_tax_value
                        found = True
                        break

                if not found:
                    tax_values[sheet_name] = 0  # Default to 0 if not found

            elif sheet_name == 'South Beach':
                # Find both Sales Tax Payable and Resort Tax Payable values
                ws = wb[sheet_name]
                for row in range(1, 50):  # Search through rows
                    tax_label_cell = ws.cell(row=row, column=6)
                    if tax_label_cell.value == 'Sales Tax Payable':
                        south_beach_sales_tax_value = ws.cell(row=row, column=7).value
                    elif tax_label_cell.value == 'Resort Tax Payable':
                        south_beach_resort_tax_value = ws.cell(row=row, column=7).value

                # Use sales tax as the default tax value
                tax_values[sheet_name] = south_beach_sales_tax_value
            else:
                # Find the Sales Tax Payable value for other locations
                ws = wb[sheet_name]
                found = False
                for row in range(1, 50):  # Search through rows
                    tax_label_cell = ws.cell(row=row, column=6)
                    if tax_label_cell.value == 'Sales Tax Payable':
                        tax_values[sheet_name] = ws.cell(row=row, column=7).value
                        found = True
                        break

                if not found:
                    tax_values[sheet_name] = 0  # Default to 0 if not found

    # Override values for special cases
    tax_values['Coral Gables'] = 0
    if 'Aventura (Miami Gardens)' in tax_values:
        tax_values['Aventura (Miami Gardens)'] = north_beach_tax_value

    # Populate the Tax tab data according to the specified store order
    store_order = [
        'Carrot Aventura Love LLC (Aventura)',
        'Carrot Coral GablesLove LLC (Coral Gabes)',
        'Carrot Downtown Love Two LLC',
        'Carrot Express Miami Shores LLC',
        'Carrot Express Midtown LLC',
        'Carrot Flatiron Love Manhattan Operating LLC',
        'Carrot Love 600 Lexington LLC',
        'Carrot Love Aventura Mall Operating LLC',
        'Carrot Love Brickell Operating LLC',
        'Carrot Love Bryant Park Operating LLC',
        'Carrot Love City Place Doral Operating LLC',
        'Carrot Love Coconut Creek Operating LLC',
        'Carrot Love Coconut Grove Operating LLC',
        'Carrot Love Dadeland Operating LLC',
        'Carrot Love Hollywood Operating LLC',
        'Carrot Love Las Olas Operating LLC',
        'Carrot Love Palmetto Park Operating LLC',
        'Carrot Love Pembroke Pines Operating LLC',
        'Carrot Love Plantation Operating LLC',
        'Carrot Love River Lading Operating LLC',
        'Carrot Love Sunset Operating LLC',
        'Carrot Love West Boca Operating LLC',
        'Carrot North Beach Love LL (North Beach)',
        'Carrot Sobe Love South Florida Operating C LLC'
    ]

    # Create a mapping from R365 location to AP location
    reverse_location_dict = {v: k for k, v in LOCATION_DICT.items()}

    row = 2
    for r365_store in store_order:
        # Find the corresponding location (sheet name) for this R365 store
        location = reverse_location_dict.get(r365_store)

        # If not found in the dict, try to match by name
        if not location:
            for loc in locations:
                if LOCATION_DICT.get(loc) == r365_store:
                    location = loc
                    break

        # Get the tax value for this location
        tax_value = tax_values.get(location, 0) if location else 0

        # Special handling for specific stores
        if r365_store == "Carrot Aventura Love LLC (Aventura)":
            tax_value = "='North Beach'!G18-'North Beach'!G17"  # Formula reference to "Carrot Love Sales Tax Payable - Resort Tax Payable"
            # Write the data
            tax_ws.cell(row=row, column=1, value=r365_store).border = thin_border
            tax_ws.cell(row=row, column=2, value="Sales Tax").border = thin_border
            cell = tax_ws.cell(row=row, column=7, value=tax_value)
            cell.border = thin_border
            cell.fill = yellow_fill  # Highlight with yellow
            tax_ws.cell(row=row, column=8, value=" <-- Make Sure All \"Sales Tax Payable\" For Carrot Love Is Updated").alignment = Alignment(horizontal='left')
            tax_ws.cell(row=row, column=11, value="No postear").border = thin_border

        elif r365_store == "Carrot North Beach Love LL (North Beach)":
            # Set North Beach tax to use the formula for Resort Tax
            tax_ws.cell(row=row, column=1, value=r365_store).border = thin_border
            tax_ws.cell(row=row, column=2, value="Resort Tax").border = thin_border

            # Add the formula cell with yellow highlight
            cell = tax_ws.cell(row=row, column=7, value="='North Beach'!G17")  # Formula to North Beach Resort Tax
            cell.border = thin_border
            cell.fill = yellow_fill  # Highlight with yellow

            # Add the explanatory note in column 8 (Comments_1)
            tax_ws.cell(row=row, column=8, value="<-- Make Sure \"Resort Tax Payable\" For North Beach Is Updated").alignment = Alignment(horizontal='left')

            tax_ws.cell(row=row, column=11, value="No postear").border = thin_border

        # South Beach - update for Sales Tax value
        elif r365_store == "Carrot Sobe Love South Florida Operating C LLC":
            # Set South Beach Sales Tax row
            tax_ws.cell(row=row, column=1, value=r365_store).border = thin_border
            tax_ws.cell(row=row, column=2, value="Sales Tax").border = thin_border
            tax_ws.cell(row=row, column=7, value=south_beach_sales_tax_value).border = thin_border
            tax_ws.cell(row=row, column=11, value="No postear").border = thin_border

            # Add an additional row for South Beach Resort Tax
            row += 1
            tax_ws.cell(row=row, column=1, value=r365_store).border = thin_border
            tax_ws.cell(row=row, column=2, value="Resort Tax").border = thin_border
            tax_ws.cell(row=row, column=7, value=south_beach_resort_tax_value).border = thin_border
            tax_ws.cell(row=row, column=11, value="No postear").border = thin_border

        else:
            # Write the data for other stores normally
            tax_ws.cell(row=row, column=1, value=r365_store).border = thin_border
            tax_ws.cell(row=row, column=2, value="Sales Tax").border = thin_border
            tax_ws.cell(row=row, column=7, value=tax_value).border = thin_border
            tax_ws.cell(row=row, column=11, value="No postear").border = thin_border

        row += 1

    # Now reorder sheets according to the specified order
    desired_order = [
        'Aventura (Miami Gardens)','Coral Gables', 'North Beach', 'Downtown', 'Brickell',
        'Dadeland','Boca Palmetto Park', 'Doral', 'Coconut Creek', 'Aventura Mall',
        'West Boca', 'Pembroke Pines', 'South Miami (Sunset)', 'Plantation',
        'Las Olas', 'Coconut Grove', 'River Landing', 'Hollywood', 'Flatiron',
        'Bryant Park', 'Lexington', 'Midtown', 'Miami Shores', 'South Beach',
         'Tax'
    ]

    # Create a mapping of sheet indices
    sheet_indices = {sheet_name: idx for idx, sheet_name in enumerate(wb.sheetnames)}

    # Reorder according to desired_order, accounting for missing sheets
    for idx, sheet_name in enumerate(desired_order):
        if sheet_name in wb.sheetnames:
            current_index = wb.sheetnames.index(sheet_name)
            target_index = idx

            # Only move if not already in correct position
            if current_index != target_index:
                # Get the sheet
                sheet = wb[sheet_name]

                # Remove sheet and reinsert at target position
                wb.remove(sheet)
                wb._sheets.insert(target_index, sheet)
    # Set column widths
    tax_ws.column_dimensions['A'].width = 290 / 7  # Convert pixels to Excel width units (approx)
    tax_ws.column_dimensions['B'].width = 15
    # Add width for column H (Comments_1)
    tax_ws.column_dimensions['H'].width = 390 / 7  # Convert pixels to Excel width units (approx)

    # Save the modified workbook
    wb.save(royalties_path)
    return royalties_path
