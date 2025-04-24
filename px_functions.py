import os
from datetime import datetime
import calendar


def get_next_business_day(date_obj, days_forward=1):
    """
    Get a business day N days forward from a given date (skipping weekends).

    Parameters:
    date_obj (datetime): The input date object
    days_forward (int): Number of business days to move forward

    Returns:
    datetime: The resulting business day
    """
    from datetime import datetime, timedelta

    result_date = date_obj.replace(hour=0, minute=0, second=0, microsecond=0)
    business_days_added = 0

    while business_days_added < days_forward:
        # Add one calendar day
        result_date += timedelta(days=1)

        # Skip weekend days
        if result_date.weekday() < 5:  # Weekday (0-4 is Monday through Friday)
            business_days_added += 1

    return result_date

def process_special_stores(redemptions_df, fee_percentages):
    """Process special stores and create separate report with Excel formatting."""
    import pandas as pd
    special_data = []
    special_stores = {"West Kendall (London Square)", "Weston", "Pinecrest"}

    special_stores_df = redemptions_df[
        (redemptions_df['Store Name'].isin(special_stores)) &
        (redemptions_df['Card Template'] == 'eGift')
    ].copy()

    # Check if we have any data before proceeding
    if special_stores_df.empty:
        print("No eGift transactions found for special stores")
        # Return an empty dataframe with the expected columns
        return pd.DataFrame(columns=[
            'Store', 'Month', 'Online Gift Card', 'Online Gift Card Fee', 'Net Payout', 'Fee %'
        ])

    special_stores_df['Date'] = pd.to_datetime(special_stores_df['Date'])
    special_stores_df['Month'] = special_stores_df['Date'].dt.strftime('%m/%Y')
    grouped = special_stores_df.groupby(['Store Name', 'Month'])['Dollars Redeemed'].sum().abs().round(2)

    for (store, month), gift_card_amount in grouped.items():
        fee_pct = fee_percentages.get(month, 0)
        fee_amount = round(gift_card_amount * fee_pct, 2)
        net_payout = round(gift_card_amount - fee_amount, 2)

        special_data.append({
            'Store': store,
            'Month': month,
            'Online Gift Card': round(gift_card_amount, 2),
            'Online Gift Card Fee': round(-fee_amount, 2),  # Make fee negative
            'Net Payout': net_payout,
            'Fee %': fee_pct  # Keep as decimal for Excel percentage formatting
        })

    df = pd.DataFrame(special_data)

    # Calculate totals
    totals = {
        'Store': 'Grand Total',
        'Month': '',
        'Online Gift Card': df['Online Gift Card'].sum(),
        'Online Gift Card Fee': df['Online Gift Card Fee'].sum(),
        'Net Payout': df['Net Payout'].sum(),
        'Fee %': None
    }

    df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)
    return df

def save_special_stores_excel(df, raw_redemptions_df, output_path):
    """Save special stores data to Excel with formatting."""
    import pandas as pd
    from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    # Convert DataFrame to Excel
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # Write summary sheet
    df.to_excel(writer, index=False, sheet_name='Gift Card Summary')

    # Filter and write details sheet
    special_stores = {"West Kendall (London Square)", "Weston", "Pinecrest"}
    details_df = raw_redemptions_df[
        (raw_redemptions_df['Store Name'].isin(special_stores)) &
        (raw_redemptions_df['Card Template'] == 'eGift')
    ].copy()
    details_df.to_excel(writer, index=False, sheet_name='Gift Card Details')

    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = workbook['Gift Card Summary']

    # Define styles
    header_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')  # Light orange
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)

    # Get the dimensions
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    # Format headers
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.border = thick_border
        cell.font = bold_font

    # Format cells and apply borders
    money_format = '#,##0.00'  # Comma style
    accounting_format = '"$"#,##0.00_);("$"#,##0.00)'  # Accounting format

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border

            # Apply number formatting to numeric columns
            if col in [3, 4, 5]:  # Online Gift Card, Fee, and Net Payout columns
                cell.number_format = accounting_format  # All rows use accounting format
                if row == max_row:
                    cell.font = bold_font

            # Format Fee % column
            elif col == 6:  # Fee % column
                if row < max_row:  # All rows except total
                    cell.number_format = '0.00%'

    # Set specific width for Month column (135 pixels)
    # Excel column width is in characters where 1 character is approximately 7 pixels
    month_col = get_column_letter(2)  # Column B is the Month column
    worksheet.column_dimensions[month_col].width = 135/7  # Convert pixels to Excel width units

    # Autofit other columns
    for col in range(1, max_col + 1):
        if col != 2:  # Skip Month column since we set it manually
            column_letter = get_column_letter(col)
            length = max(len(str(cell.value)) for cell in worksheet[column_letter])
            worksheet.column_dimensions[column_letter].width = length + 2

    # Create table without filter buttons by manually applying formatting
    for row in range(2, max_row + 1):
        if row % 2 == 0:  # Apply alternating row colors
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    # Add thick border around the entire table and first/last rows
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)

            # Regular cell borders (non-edge cells)
            if row not in [1, max_row] and col not in [1, max_col]:
                cell.border = thin_border

            # First row
            if row == 1:
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
            # Last row
            elif row == max_row:
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )

            # Left edge
            if col == 1:
                cell.border = Border(
                    top=Side(style='thick') if row in [1, max_row] else Side(style='thin'),
                    bottom=Side(style='thick') if row in [1, max_row] else Side(style='thin'),
                    left=Side(style='thick'),
                    right=Side(style='thin')
                )
            # Right edge
            elif col == max_col:
                cell.border = Border(
                    top=Side(style='thick') if row in [1, max_row] else Side(style='thin'),
                    bottom=Side(style='thick') if row in [1, max_row] else Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thick')
                )

            # Corner cases
            if row == 1 and col == 1:  # Top-left corner
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick'),
                    left=Side(style='thick'),
                    right=Side(style='thin')
                )
            elif row == 1 and col == max_col:  # Top-right corner
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thick')
                )
            elif row == max_row and col == 1:  # Bottom-left corner
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick'),
                    left=Side(style='thick'),
                    right=Side(style='thin')
                )
            elif row == max_row and col == max_col:  # Bottom-right corner
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thick')
                )

    # Format Details Sheet
    details_sheet = workbook['Gift Card Details']

    # Autofit columns for details sheet
    for column_cells in details_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        details_sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Save the workbook
    writer.close()


def get_last_day_of_month(date_str):
    """Get the last day of the month for a given date string."""
    import pandas as pd
    try:
        date_obj = pd.to_datetime(date_str)
        last_day = calendar.monthrange(date_obj.year, date_obj.month)[1]
        return datetime(date_obj.year, date_obj.month, last_day).strftime('%m/%d/%Y')
    except:
        print(f"Error processing date: {date_str}")
        raise

def get_month_name(date_str):
    import pandas as pd
    """Get the month name from a date string."""
    return pd.to_datetime(date_str).strftime('%B')

def get_month_year(date_str):
    import pandas as pd
    """Get month and year in mmyyyy format."""
    return pd.to_datetime(date_str).strftime('%m%Y')

def load_chase_data(directory):
    """Load and filter Chase bank data."""
    import pandas as pd
    import os

    # Update file detection to match actual filename pattern
    chase_files = [f for f in os.listdir(directory) if 'Chase' in f and f.endswith('.CSV')]
    if not chase_files:
        raise FileNotFoundError("No Chase CSV file found in the directory")

    file_path = os.path.join(directory, chase_files[0])
    print(f"\nReading Chase file: {file_path}")

    try:
        # Read CSV with explicit fixed column names and handling
        chase_df = pd.read_csv(
            file_path,
            names=["Details", "Posting Date", "Description", "Amount", "Type", "Balance", "Check or Slip #"],
            dtype={
                "Details": str,
                "Posting Date": str,
                "Description": str,
                "Amount": str,
                "Type": str,
                "Balance": str,
                "Check or Slip #": str
            },
            skipinitialspace=True,
            skiprows=1  # Skip the original header row since we're providing our own
        )

        print("\nInitial data shape:", chase_df.shape)
        print("Initial columns:", chase_df.columns.tolist())
        print("\nFirst few rows of raw data:")
        print(chase_df.head())

        # Filter for Paytronix transactions
        filtered_df = chase_df[
            chase_df['Description'].str.contains('ORIG CO NAME:Paytronix', case=False, na=False, regex=True)
        ].copy()

        # Convert Amount to float
        filtered_df['Amount'] = pd.to_numeric(filtered_df['Amount'].str.strip(), errors='coerce')

        print(f"\nFound {len(filtered_df)} Paytronix transactions")
        print("\nFiltered data sample:")
        print(filtered_df.head())

        return filtered_df

    except Exception as e:
        print(f"\nError reading Chase file: {str(e)}")
        print(f"File path: {file_path}")
        print(f"Directory contents: {os.listdir(directory)}")
        raise

def load_paytronix_data(directory):
    import pandas as pd
    payout_files = [f for f in os.listdir(directory) if f.startswith('Payouts') and f.endswith('.csv')]
    if not payout_files:
        raise FileNotFoundError("No Paytronix Payout CSV file found in the directory")

    # Read CSV specifying all columns as strings
    px_df = pd.read_csv(
        os.path.join(directory, payout_files[0]),
        dtype={
            'Unnamed: 0': str,
            'Payout ID': str,
            'Payout Status': str,
            'Description': str,
            'Payout Created Date': str,
            'Payout Arrival Date': str,
            'Gross': str,
            'Fees': str,
            'Total': str
        }
    )

    # Convert date
    px_df['Payout Created Date'] = pd.to_datetime(px_df['Payout Created Date']).dt.strftime('%m/%d/%Y')

    # Clean currency values
    for col in ['Gross', 'Fees', 'Total']:
        # Ensure column is string type before using str accessor
        px_df[col] = px_df[col].astype(str).str.replace('$', '').str.replace(',', '').astype(float)

    return px_df

def load_redemptions_data(directory):
    """Load PX redemptions data from multiple files, skipping the first row of each."""
    import pandas as pd
    redemption_files = [f for f in os.listdir(directory) if 'Stored' in f and f.endswith('.csv')]
    if not redemption_files:
        raise FileNotFoundError("No Redemption CSV file found in the directory")

    all_data = []
    for file in redemption_files:
        df = pd.read_csv(
            os.path.join(directory, file),
            skiprows=1,
            dtype={'Store Name': str, 'Card Template': str, 'Dollars Redeemed': float, 'Date': str}
        )
        all_data.append(df)

    # Combine all dataframes
    combined_df = pd.concat(all_data, ignore_index=True)

    # Ensure Date is properly formatted
    combined_df['Date'] = pd.to_datetime(combined_df['Date']).dt.strftime('%m/%d/%Y')

    return combined_df

def calculate_monthly_fee_percentages(px_df):
   """Calculate fee percentage for each month from payout data."""
   import pandas as pd
   px_df['Month'] = pd.to_datetime(px_df['Payout Created Date']).dt.strftime('%m/%Y')
   monthly_fees = {}

   for month in px_df['Month'].unique():
       month_data = px_df[px_df['Month'] == month]
       total_fees = month_data['Fees'].sum()
       total_gross = month_data['Gross'].sum()
       monthly_fees[month] = total_fees / total_gross if total_gross != 0 else 0

   return monthly_fees


SPECIAL_STORES = {
    "West Kendall (London Square)",
    "Weston",
    "Pinecrest"
}

STORE_TO_DETAIL_LOCATION = {
   "Aventura (Miami Gardens)": "Carrot Aventura Love LLC (Aventura)",
   "Aventura Mall": "Carrot Love Aventura Mall Operating LLC",
   "Boca Raton East": "Carrot Love Palmetto Park Operating LLC",
   "Brickell": "Carrot Love Brickell Operating LLC",
   "Bryant Park (NY)": "Carrot Love Bryant Park Operating LLC",
   "Coconut Creek": "Carrot Love Coconut Creek Operating LLC",
   "Coconut Grove": "Carrot Love Coconut Grove Operating LLC",
   "Coral Gables (Miracle Mile)": "Carrot Coral GablesLove LLC (Coral Gabes)",
   "Dadeland": "Carrot Love Dadeland Operating LLC",
   "Doral": "Carrot Love City Place Doral Operating LLC",
   "Downtown Miami": "Carrot Downtown Love Two LLC",
   "Flatiron (NY)": "Carrot Flatiron Love Manhattan Operating LLC",
   "Hollywood": "Carrot Love Hollywood Operating LLC",
   "Las Olas": "Carrot Love Las Olas Operating LLC",
   "Lexington (NY)": "Carrot Love 600 Lexington LLC",
   "Miami Shores": "Carrot Express Miami Shores LLC",
   "Midtown Miami": "Carrot Express Midtown LLC",
   "North Beach": "Carrot North Beach Love LL (North Beach)",
   "Pembroke Pines": "Carrot Love Pembroke Pines Operating LLC",
   "Plantation": "Carrot Love Plantation Operating LLC",
   "River Landing": "Carrot Love River Lading Operating LLC",
   "South Miami (Sunset)": "Carrot Love Sunset Operating LLC",
   "West Boca": "Carrot Love West Boca Operating LLC",
   "Sunset Harbour (Sobe)": "Carrot Sobe Love South Florida Operating C LLC"
}

STORE_TO_CHECKING = {
   "Aventura (Miami Gardens)": "Checking Carrot Love LLC",
   "Aventura Mall": "Checking Carrot Love Aventura Mall Operating LLC",
   "Boca Raton East": "Checking Carrot Love Palmetto Park Operating LLC",
   "Brickell": "Checking Carrot Love Brickell Operating LLC",
   "Bryant Park (NY)": "Checking Carrot Love Bryant Park Operating LLC",
   "Coconut Creek": "Checking Carrot Love Coconut Creek Operating LLC",
   "Coconut Grove": "Checking Carrot Love Coconut Grove Operating LLC",
   "Coral Gables (Miracle Mile)": "Checking Carrot Love LLC",
   "Dadeland": "Checking Carrot Love Dadeland Operating LLC",
   "Doral": "Checking Carrot Love City Place Doral Operating LLC",
   "Downtown Miami": "Checking Carrot Love Two LLC",
   "Flatiron (NY)": "Checking Carrot Love Manhattan Operating LLC",
   "Hollywood": "Checking Carrot Love Hollywood Operating LLC",
   "Las Olas": "Checking Carrot Love Las Olas Operating LLC",
   "Lexington (NY)": "Checking Carrot Love Lexington 52 LLC",
   "Miami Shores": "Checking Carrot Express Miami Shores LLC",
   "Midtown Miami": "Checking Carrot Express Midtown LLC",
   "North Beach": "Checking Carrot Love LLC",
   "Pembroke Pines": "Checking Carrot Love Pembroke Pines Operating LLC",
   "Plantation": "Checking Carrot Love Plantation Operating LLC",
   "River Landing": "Checking Carrot Love River Landing LLC",
   "South Miami (Sunset)": "Checking Carrot Love Sunset Operating LLC",
   "West Boca": "Checking Carrot Love West Boca Operating LLC",
   "Sunset Harbour (Sobe)": "Checking Carrot Love South Florida Operating C LLC"
}

CHECKING_TO_ACCOUNT = {
   "Checking Carrot Leadership LLC": "30000480952",
   "Checking Carrot Love LLC": "30000481123",
   "Checking Carrot Love Two LLC": "30000481258",
   "Checking Carrot Love City Place Doral Operating LLC": "30000481978",
   "Checking Carrot Love Palmetto Park Operating LLC": "30000482122",
   "Checking Carrot Love Brickell Operating LLC": "30000482104",
   "Checking Carrot Love West Boca Operating LLC": "30000482140",
   "Checking Carrot Love Aventura Mall Operating LLC": "30000482023",
   "Checking Carrot Love Coconut Creek Operating LLC": "30000482167",
   "Checking Carrot Love Coconut Grove Operating LLC": "30000482176",
   "Checking Carrot Love Sunset Operating LLC": "30000482212",
   "Checking Carrot Love Pembroke Pines Operating LLC": "30000594757",
   "Checking Carrot Love Plantation Operating LLC": "30000482149",
   "Checking Carrot Love River Landing LLC": "30000482230",
   "Checking Carrot Love Las Olas Operating LLC": "30000482158",
   "Checking Carrot Love Hollywood Operating LLC": "30000482203",
   "Checking Carrot Love South Florida Operating C LLC": "30000633502",
   "Checking Carrot Love South Florida Operating A LLC": "30000633448",
   "Checking Carrot Love Manhattan Operating LLC": "30000482131",
   "Checking Carrot Love Bryant Park Operating LLC": "30000482410",
   "Checking Carrot Love Lexington 52 LLC": "30000510616",
   "Checking Carrot Love Dadeland Operating LLC": "30000481834",
   "Checking Carrot Express Miami Shores LLC": "##N/A##",
   "Checking Carrot Express Midtown LLC": "##N/A##"
}

def create_redemption_entries(redemptions_df, fee_percentages):
   """Create journal entries for redemptions data."""
   import pandas as pd
   from datetime import datetime, timedelta

   # Get today's date
   today = datetime.now()

   # Calculate deposited dates based on today
   next_business_day = get_next_business_day(today)
   next_business_day_str = next_business_day.strftime('%m/%d/%Y')

   # For Miami Shores and Midtown, use 2 business days later
   two_business_days_later = get_next_business_day(today, days_forward=2)
   two_business_days_later_str = two_business_days_later.strftime('%m/%d/%Y')

   special_stores = {"West Kendall (London Square)", "Weston", "Pinecrest"}
   redemptions_df = redemptions_df[~redemptions_df['Store Name'].isin(special_stores)].copy()

   redemptions_df['Date'] = pd.to_datetime(redemptions_df['Date'])
   redemptions_df['Month'] = redemptions_df['Date'].dt.strftime('%m/%Y')

   journal_entries = []
   counter = 1

   egift_data = redemptions_df[redemptions_df['Card Template'] == 'eGift'].copy()
   grouped = egift_data.groupby(['Store Name', 'Month'])['Dollars Redeemed'].sum().abs().round(2)

   for (store, month), gift_card_amount in grouped.items():
       fee_pct = fee_percentages.get(month, 0)
       fee_amount = round(gift_card_amount * fee_pct, 2)
       net_amount = round(gift_card_amount - fee_amount, 2)

       detail_location = STORE_TO_DETAIL_LOCATION.get(store, store)
       checking_account = STORE_TO_CHECKING.get(store, "Unknown")

       month_date = datetime.strptime(month, '%m/%Y')
       last_day = calendar.monthrange(month_date.year, month_date.month)[1]
       date_str = month_date.replace(day=last_day).strftime('%m/%d/%Y')
       je_number = f"TRANSFER-PX{month_date.strftime('%m%Y')}-{str(counter).zfill(2)}"
       month_name = month_date.strftime('%B')

       # Determine which date to use in comments based on the store
       comment_date = two_business_days_later_str if store in ["Miami Shores", "Midtown Miami"] else next_business_day_str

       base_entry = {
           'JENumber': je_number,
           'Date': date_str,
           'JEComment': f"Deposited {comment_date} // Gift Cards Redeemed - {month_name}",
           'JELocation': "Carrot Leadership LLC",
           'DetailComment': f"Deposited {comment_date} // Gift Cards Redeemed - {month_name}"
       }

       entries = [
           {**base_entry,
            'Account': "Checking Carrot Leadership LLC",
            'Debit': 0,
            'Credit': net_amount,
            'DetailLocation': "Carrot Leadership LLC"},
           {**base_entry,
            'Account': "Online Gift Card Fee",
            'Debit': 0,
            'Credit': fee_amount,
            'DetailLocation': "Carrot Leadership LLC"},
           {**base_entry,
            'Account': "Online Gift Card",
            'Debit': gift_card_amount,
            'Credit': 0,
            'DetailLocation': "Carrot Leadership LLC"},
           {**base_entry,
            'Account': checking_account,
            'Debit': net_amount,
            'Credit': 0,
            'DetailLocation': detail_location},
           {**base_entry,
            'Account': "Merchant Account Fees",
            'Debit': fee_amount,
            'Credit': 0,
            'DetailLocation': detail_location},
           {**base_entry,
            'Account': "Gift Cards Outstanding",
            'Debit': 0,
            'Credit': gift_card_amount,
            'DetailLocation': detail_location}
       ]

       journal_entries.extend(entries)
       counter += 1

   return pd.DataFrame(journal_entries)

def create_payout_entries(chase_df, px_df, redemptions_df):
    """Create journal entries based on Chase, Paytronix data, and redemptions data."""
    import pandas as pd
    from datetime import datetime
    journal_entries = []
    monthly_totals = {}  # Dictionary to store monthly totals
    grand_total = 0

    # Calculate fee percentages
    fee_percentages = calculate_monthly_fee_percentages(px_df)

    # Process Paytronix deposits first
    px_daily = px_df.groupby('Payout Created Date').agg({
        'Gross': 'sum',
        'Fees': 'sum',
        'Total': 'sum'
    }).reset_index()

    for posting_date in chase_df['Posting Date'].unique():
        # Keep original sign for determining debit/credit
        raw_chase_amount = float(chase_df[chase_df['Posting Date'] == posting_date]['Amount'].iloc[0])
        chase_amount = round(abs(raw_chase_amount), 2)
        month_name = get_month_name(posting_date)
        je_number = f"PxDep-{get_month_year(posting_date)}"
        last_day_of_month = get_last_day_of_month(posting_date)
        px_entry = px_daily[px_daily['Payout Created Date'] == posting_date]

        if px_entry.empty:
            entries = [
                {
                    'JENumber': je_number,
                    'Date': last_day_of_month,
                    'JEComment': f"PX Gift Card Deposits {month_name}",
                    'JELocation': "Carrot Leadership LLC",
                    'Account': "Savings Chase Carrot Leadership LLC",
                    'Debit': chase_amount if raw_chase_amount > 0 else 0,
                    'Credit': chase_amount if raw_chase_amount < 0 else 0,
                    'DetailLocation': "Carrot Leadership LLC",
                    'DetailComment': f"Deposited {posting_date} // Pending - Missing payout information from PX Payout Report"
                },
                {
                    'JENumber': je_number,
                    'Date': last_day_of_month,
                    'JEComment': f"PX Gift Card Deposits {month_name}",
                    'JELocation': "Carrot Leadership LLC",
                    'Account': "Exchange",
                    'Debit': 0,
                    'Credit': chase_amount,
                    'DetailLocation': "Carrot Leadership LLC",
                    'DetailComment': f"Deposited {posting_date} // Pending - Missing payout information from PX Payout Report"
                }
            ]
        else:
            # Keep original signs for Paytronix values
            raw_px_gross = float(px_entry['Gross'].iloc[0])
            px_gross = round(abs(raw_px_gross), 2)
            px_fees = round(abs(float(px_entry['Fees'].iloc[0])), 2)
            px_total = round(abs(float(px_entry['Total'].iloc[0])), 2)

            entries = [
                {
                    'JENumber': je_number,
                    'Date': last_day_of_month,
                    'JEComment': f"PX Gift Card Deposits {month_name}",
                    'JELocation': "Carrot Leadership LLC",
                    'Account': "Savings Chase Carrot Leadership LLC",
                    'Debit': chase_amount if raw_chase_amount > 0 else 0,
                    'Credit': chase_amount if raw_chase_amount < 0 else 0,
                    'DetailLocation': "Carrot Leadership LLC",
                    'DetailComment': f"Deposited {posting_date}"
                },
                {
                    'JENumber': je_number,
                    'Date': last_day_of_month,
                    'JEComment': f"PX Gift Card Deposits {month_name}",
                    'JELocation': "Carrot Leadership LLC",
                    'Account': "Online Gift Card",
                    'Debit': px_gross if raw_px_gross < 0 else 0,
                    'Credit': px_gross if raw_px_gross > 0 else 0,
                    'DetailLocation': "Carrot Leadership LLC",
                    'DetailComment': f"Deposited {posting_date}"
                },
                {
                    'JENumber': je_number,
                    'Date': last_day_of_month,
                    'JEComment': f"PX Gift Card Deposits {month_name}",
                    'JELocation': "Carrot Leadership LLC",
                    'Account': "Online Gift Card Fee",
                    'Debit': px_fees,
                    'Credit': 0,
                    'DetailLocation': "Carrot Leadership LLC",
                    'DetailComment': f"Deposited {posting_date}"
                }
            ]

            if abs(chase_amount - px_total) > 0.01:
                discrepancy = round(chase_amount - px_total, 2)
                entries.append({
                    'JENumber': je_number,
                    'Date': last_day_of_month,
                    'JEComment': f"PX Gift Card Deposits {month_name}",
                    'JELocation': "Carrot Leadership LLC",
                    'Account': "Exchange",
                    'Debit': abs(discrepancy) if discrepancy < 0 else 0,
                    'Credit': abs(discrepancy) if discrepancy > 0 else 0,
                    'DetailLocation': "Carrot Leadership LLC",
                    'DetailComment': f"Deposited {posting_date} // Pending - Discrepancy in PX Payout Report"
                })

        journal_entries.extend(entries)

    # Calculate monthly transfer amounts based on eGift redemptions
    # Include ALL stores (including special stores) for the transfer calculation
    redemptions_df['Date'] = pd.to_datetime(redemptions_df['Date'])
    redemptions_df['month'] = redemptions_df['Date'].dt.strftime('%m/%Y')

    # Group by store and month first, then calculate fees and net amounts
    store_month_amounts = redemptions_df[
        redemptions_df['Card Template'] == 'eGift'
    ].groupby(['Store Name', 'month'])['Dollars Redeemed'].sum().abs().round(2)

    # Calculate monthly totals by processing each store's amounts separately
    for month in store_month_amounts.index.get_level_values('month').unique():
        month_stores = store_month_amounts[store_month_amounts.index.get_level_values('month') == month]
        fee_pct = fee_percentages.get(month, 0)

        # Calculate fees and net amounts for each store separately
        month_total = 0
        for store_amount in month_stores:
            fee_amount = round(store_amount * fee_pct, 2)
            net_amount = round(store_amount - fee_amount, 2)
            month_total += net_amount

        month_date = pd.to_datetime(month, format='%m/%Y')
        last_day = get_last_day_of_month(month_date)
        month_name = get_month_name(month)

        monthly_totals[month] = round(month_total, 2)
        grand_total += round(month_total, 2)

        # Use today's date as the reference point for next business day
        today = datetime.now()

        # Get the next business day for TRANSFER-PxDep entries based on today
        next_business_day = get_next_business_day(today)
        next_business_day_str = next_business_day.strftime('%m/%d/%Y')

        # Add transfer entries - use next business day for Date
        transfer_entries = [
            {
                'JENumber': f"TRANSFER-PxDep-{get_month_year(last_day)}",
                'Date': next_business_day_str,  # Use next business day
                'JEComment': f"Deposited {next_business_day_str} // Manually transferred from Chase to CNB",
                'JELocation': "Carrot Leadership LLC",
                'Account': "Savings Chase Carrot Leadership LLC",
                'Debit': 0,
                'Credit': month_total,
                'DetailLocation': "Carrot Leadership LLC",
                'DetailComment': f"Deposited {next_business_day_str} // Manually transferred from Chase to CNB"
            },
            {
                'JENumber': f"TRANSFER-PxDep-{get_month_year(last_day)}",
                'Date': next_business_day_str,  # Use next business day
                'JEComment': f"Deposited {next_business_day_str} // Manually transferred from Chase to CNB",
                'JELocation': "Carrot Leadership LLC",
                'Account': "Checking Carrot Leadership LLC",
                'Debit': month_total,
                'Credit': 0,
                'DetailLocation': "Carrot Leadership LLC",
                'DetailComment': f"Deposited {next_business_day_str} // Manually transferred from Chase to CNB"
            }
        ]
        journal_entries.extend(transfer_entries)

    # Create DataFrame and ensure all monetary values are rounded
    df = pd.DataFrame(journal_entries)
    if not df.empty:
        df['Debit'] = round(df['Debit'], 2)
        df['Credit'] = round(df['Credit'], 2)

    return df, monthly_totals, round(grand_total, 2)

def create_transfer_files(redemption_entries):
    """Create CNB transfer files and ACHB payment file."""
    import pandas as pd
    import math
    cnb_transfers = []

    for je_number in redemption_entries['JENumber'].unique():
        je_data = redemption_entries[redemption_entries['JENumber'] == je_number]

        # Get leadership entries
        leadership_entries = je_data[
            (je_data['DetailLocation'] == "Carrot Leadership LLC") &
            (je_data['Credit'] > 0)
        ]

        if leadership_entries.empty:
            continue

        leadership_credit = leadership_entries['Credit'].iloc[0]
        transaction_date = pd.to_datetime(je_data['Date'].iloc[0])
        month = transaction_date.strftime('%B %Y')

        # Get non-Leadership location data
        non_leadership_data = je_data[je_data['DetailLocation'] != "Carrot Leadership LLC"]
        if non_leadership_data.empty:
            continue

        # Get checking account entries
        checking_entries = non_leadership_data[
            non_leadership_data['Account'].str.startswith('Checking')
        ]
        if checking_entries.empty:
            continue

        detail_location = non_leadership_data['DetailLocation'].iloc[0]
        checking_account = checking_entries['Account'].iloc[0]
        account_number = CHECKING_TO_ACCOUNT.get(checking_account)

        # Only add to CNB transfers if it's not Miami Shores or Midtown
        if account_number != "##N/A##":
            cnb_transfers.append({
                'Month': month,
                'From': 30000480952,
                'To': int(account_number),
                'Amount': round(leadership_credit, 2),
                'From account ---> To account': f"Checking Carrot Leadership LLC ---> {checking_account}"
            })

    today = datetime.now().strftime('%m%d%Y')

    # Process CNB transfers in batches of 35
    cnb_df = pd.DataFrame(cnb_transfers)

    # Reorder columns to put Month first
    if not cnb_df.empty:
        cnb_df = cnb_df[['Month', 'From', 'To', 'Amount', 'From account ---> To account']]

        # Calculate number of files needed
        num_files = math.ceil(len(cnb_df) / 35)

        # Split into multiple dataframes if needed
        cnb_dfs = []
        for i in range(num_files):
            start_idx = i * 35
            end_idx = min((i + 1) * 35, len(cnb_df))
            file_suffix = f"-{str(i + 1).zfill(2)}" if num_files > 1 else ""
            cnb_dfs.append({
                'df': cnb_df.iloc[start_idx:end_idx],
                'filename': f"PX_CNB_Transfer_{today}{file_suffix}.csv"
            })

    return cnb_dfs if not cnb_df.empty else [], today

def save_all_output_files(chase_df, px_df, redemptions_df, modified_dir):
    """Save all output files."""
    import pandas as pd
    from datetime import datetime
    import os

    # Get date range
    date_range = pd.to_datetime(redemptions_df['Date'])
    start_date = date_range.min()
    end_date = date_range.max()
    start_month = start_date.strftime('%B %Y')
    end_month = end_date.strftime('%B %Y')
    month_range = start_month if start_month == end_month else f"{start_month}-{end_month}"

    # Calculate fee percentages
    fee_percentages = calculate_monthly_fee_percentages(px_df)

    # 1. Process special stores (Maduro file)
    special_stores_df = process_special_stores(redemptions_df, fee_percentages)
    special_filename = f"MaduroPX_{month_range}.xlsx"
    special_filepath = os.path.join(modified_dir, special_filename)
    save_special_stores_excel(special_stores_df, redemptions_df, special_filepath)
    print(f"Created Maduro file: {special_filename}")
    # Create AP Invoices file
    ap_invoices_filename = create_ap_invoices(special_stores_df, modified_dir)
    print(f"Created AP Invoices file: {ap_invoices_filename}")

    # 2. Create redemption entries
    redemption_entries = create_redemption_entries(redemptions_df, fee_percentages)
    redemption_filename = f"PX_Redemptions_{datetime.now().strftime('%m%d%Y')}.csv"
    redemption_filepath = os.path.join(modified_dir, redemption_filename)
    redemption_entries.to_csv(redemption_filepath, index=False)
    print(f"Created redemptions file: {redemption_filename}")

    # 3. Create payout entries - updated to include redemptions_df
    payout_entries, monthly_totals, grand_total = create_payout_entries(chase_df, px_df, redemptions_df)
    payout_filename = f"PX_LeadershipPayouts_{month_range}.csv"
    payout_filepath = os.path.join(modified_dir, payout_filename)
    payout_entries.to_csv(payout_filepath, index=False)
    print(f"Created leadership payouts file: {payout_filename}")

    # Print transfer amounts summary
    print("\nTransfer Amounts Summary:")
    for month, amount in monthly_totals.items():
        month_date = pd.to_datetime(month, format='%m/%Y')
        month_name = month_date.strftime('%B %Y')
        print(f"{month_name}: ${amount:,.2f}")
    print(f"\nTotal Transfer Amount from Chase to CNB: ${grand_total:,.2f}\n")

    # 4. Create transfer files
    cnb_dfs, today = create_transfer_files(redemption_entries)

    # Save CNB transfer files
    for file_info in cnb_dfs:
        file_info['df'].to_csv(os.path.join(modified_dir, file_info['filename']), index=False)
        print(f"Created CNB transfer file: {file_info['filename']}")

    # Create ACHB payment file if needed
    achb_filename = create_achb_payment(redemption_entries, modified_dir)
    if achb_filename:
        print(f"Created ACHB Payment file: {achb_filename}")

    return month_range, start_month, end_month

def create_ap_invoices(special_stores_df, output_dir):
    """
    Create AP Invoices file for special stores based on the MaduroPX data.

    Parameters:
    special_stores_df (pandas.DataFrame): DataFrame containing the special stores data
    output_dir (str): Directory path where the output file should be saved
    """
    import pandas as pd
    import os
    from datetime import datetime

    # Check if there's any data to process (other than the Grand Total row)
    store_data = special_stores_df[special_stores_df['Store'] != 'Grand Total']
    if store_data.empty:
        print("No special store data to process for AP invoices")
        return "No_AP_Invoices_Required.txt"

    # Store name translation dictionary
    STORE_TO_VENDOR = {
        "West Kendall (London Square)": "5A Healthy Restaurants LLC WKendall",
        "Weston": "5M Healthy Restaurants LLC Weston",
        "Pinecrest": "5AM Healthy Restaurants LLC Pinecre"
    }

    # Get today's date in different formats
    today = datetime.now()
    today_mmddyyyy = today.strftime('%m%d%Y')
    today_slashes = today.strftime('%m/%d/%Y')

    # Initialize list to store invoice rows
    invoice_rows = []

    # Filter out the Grand Total row and get month range for comment
    store_data = special_stores_df[special_stores_df['Store'] != 'Grand Total']
    months = pd.Series(store_data['Month'].unique())
    months = pd.to_datetime(months, format='%m/%Y')
    start_month = months.min().strftime('%B %Y')
    end_month = months.max().strftime('%B %Y')
    month_range = f"{start_month} - {end_month}" if start_month != end_month else start_month
    comment = f"PX Gift Cards Redeemed // {month_range}"

    # Group by store and calculate totals (excluding Grand Total row)
    store_totals = store_data.groupby('Store').agg({
        'Online Gift Card': lambda x: round(sum(x), 2),
        'Online Gift Card Fee': lambda x: round(sum(x), 2),
        'Net Payout': lambda x: round(sum(x), 2)
    }).reset_index()

    # Create invoice rows for each store
    for idx, row in store_totals.iterrows():
        invoice_number = f"PX5A-{today_mmddyyyy}-{idx + 1}"
        vendor = STORE_TO_VENDOR[row['Store']]

        # Common fields for both rows
        common_fields = {
            'Type': 'AP Invoice',
            'Location': 'Carrot Leadership LLC',
            'Vendor': vendor,
            'Number': invoice_number,
            'Date': today_slashes,
            'Gl Date': today_slashes,
            'Amount': round(abs(row['Net Payout']), 2),  # Ensure positive amount and round
            'Payment Terms': 'Due Upon Receipt',
            'Due Date': today_slashes,
            'Comment': comment,
            'Detail Location': 'Carrot Leadership LLC',
            'Detail Comment': comment
        }

        # First row - Online Gift Card
        invoice_rows.append({
            **common_fields,
            'Detail Account': 'Online Gift Card',
            'Detail Amount': round(abs(row['Online Gift Card']), 2)  # Positive amount and round
        })

        # Second row - Online Gift Card Fee
        invoice_rows.append({
            **common_fields,
            'Detail Account': 'Online Gift Card Fee',
            'Detail Amount': round(-abs(row['Online Gift Card Fee']), 2)  # Negative amount and round
        })

    # Create DataFrame and ensure all columns match template
    invoices_df = pd.DataFrame(invoice_rows)

    # Save to CSV
    output_filename = f"AP_Invoices_Maduro_{today_mmddyyyy}.csv"
    output_path = os.path.join(output_dir, output_filename)
    invoices_df.to_csv(output_path, index=False)

    return output_filename

def create_achb_payment(redemption_entries, output_dir):
    """
    Create ACHB Payment file for external transfers (Miami Shores and Midtown).

    Parameters:
    redemption_entries (pandas.DataFrame): DataFrame containing the redemption entries
    output_dir (str): Directory path where the output file should be saved
    """
    import pandas as pd
    import os
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # Translation dictionaries
    STORE_TO_VENDOR = {
        "Miami Shores": "Carrot Express Miami Shores LLC",
        "Midtown": "Carrot Express Midtown LLC"
    }

    STORE_TO_ACCOUNT = {
        "Miami Shores": "6766054644",
        "Midtown": "690122067"
    }

    STORE_TO_ROUTING = {
        "Miami Shores": "63107513",
        "Midtown": "267084131"
    }

    # Initialize lists to store payment data
    payment_data = []

    # Process each store's data
    for store_name in ["Miami Shores", "Midtown"]:
        # Get entries for this store
        store_entries = redemption_entries[
            (redemption_entries['DetailLocation'].str.contains(store_name, case=False, na=False)) &
            (redemption_entries['Account'].str.startswith('Checking', na=False))
        ]

        if not store_entries.empty:
            # Calculate total amount for the store
            total_amount = round(store_entries['Debit'].sum(), 2)

            if total_amount > 0:  # Only create entry if there's an amount to pay
                payment_data.append({
                    'Pay $': total_amount,
                    'SEC Code': 'CCD',
                    'Location Account Number': '30000480952',
                    'Location Subsidiary': '0952 CR CCD/PPD CARROT LEADERSHIP LLC',
                    'Vendor Display Name': STORE_TO_VENDOR[store_name],
                    'Vendor Account Number': STORE_TO_ACCOUNT[store_name],
                    'Vendor Routing Number': STORE_TO_ROUTING[store_name]
                })

    if payment_data:
        # Create output filename
        today = datetime.now().strftime('%m%d%Y')
        output_filename = f"ACHB_PX_{today}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        # Create DataFrame
        df = pd.DataFrame(payment_data)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active

        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header

        # Write data
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            column = ws[column_letter]

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        print(f"Created ACHB Payment file: {output_filename}")
        return output_filename

    return None
