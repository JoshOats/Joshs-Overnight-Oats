import os
import sys
import re
from datetime import datetime
from PyQt5.QtWidgets import (QVBoxLayout, QPushButton, QLabel,
                            QFileDialog, QMessageBox, QTextEdit, QApplication, QListWidget)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QIcon
from retro_style import RetroWindow, create_retro_central_widget




def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class TipsReconcileWindow(RetroWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.selected_files = []
        self.error_handler = AIErrorHandler()  # Initialize error handler

        icon_path = resource_path(os.path.join('assets', 'icon.png'))
        self.setWindowIcon(QIcon(icon_path))

    def initUI(self):
        central_widget = create_retro_central_widget(self)
        layout = QVBoxLayout(central_widget)

        # Title
        title_label = QLabel("Tips Reconciliation", self)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setObjectName("title_label")
        layout.addWidget(title_label)

        # Instructions button
        self.instructions_button = QPushButton('Instructions')
        self.instructions_button.clicked.connect(self.show_instructions)
        layout.addWidget(self.instructions_button)

        # Input Files button and list
        input_layout = QVBoxLayout()
        self.input_button = QPushButton('Input Files')
        self.input_button.clicked.connect(self.select_files)
        input_layout.addWidget(self.input_button)
        self.file_list = QListWidget()
        input_layout.addWidget(self.file_list)
        self.file_list.setFixedHeight(150)
        layout.addLayout(input_layout)

        # Run button
        self.run_button = QPushButton('RUN')
        self.run_button.clicked.connect(self.run_reconciliation)
        layout.addWidget(self.run_button)

        # Console output
        self.console_output = QTextEdit()
        self.console_output.setReadOnly(True)
        layout.addWidget(self.console_output)

        self.setWindowTitle('Tips Reconciliation')
        self.resize(1000, 738)
        self.center()

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Input Files", "", "Excel/CSV Files (*.xlsx *.csv)"
        )
        if files:
            self.selected_files = files
            self.update_file_list()


    def update_file_list(self):
        self.file_list.clear()
        for file in self.selected_files:
            self.file_list.addItem(os.path.basename(file))

    def run_reconciliation(self):
        if not self.selected_files:
            QMessageBox.warning(self, "Error", "Please select input files.")
            return

        self.console_output.clear()
        self.console_output.append("Starting tips reconciliation process...")
        self.run_button.setEnabled(False)

        # Get Downloads path and create filename directly
        downloads_path = os.path.expanduser("~/Downloads")
        current_date = datetime.now().strftime("%m%d%Y")
        output_file = os.path.join(downloads_path, f"Tips_Reconciliation_{current_date}.xlsx")

        self.reconcile_thread = TipsReconcileThread(
            self.selected_files,
            output_file  # Pass the file path instead of directory
        )
        self.reconcile_thread.update_signal.connect(self.update_console)
        self.reconcile_thread.finished_signal.connect(self.reconciliation_finished)
        self.reconcile_thread.start()

    def update_console(self, message):
        self.console_output.append(message)

    def reconciliation_finished(self, success, message):
        self.run_button.setEnabled(True)
        if success:
            QMessageBox.information(self, "Success", message)
        else:
            # Try to extract filename from error message using different patterns
            filename = None

            # Check for file patterns in the traceback
            file_patterns = [
                r'File ".*?([^\\\/]+\.(?:csv|xlsx))"',  # Matches filenames in traceback
                r"Error (?:processing|reading) file ['\"](.+?)['\"]",  # Matches direct file references
                r"reading file:? ([^'\n]+\.(?:csv|xlsx))",  # Matches "reading file" messages
                r"Cannot access file:? ([^'\n]+\.(?:csv|xlsx))"  # Matches access errors
            ]

            for pattern in file_patterns:
                match = re.search(pattern, message)
                if match:
                    filename = match.group(1)
                    break

            # Use the AI error handler to interpret the error
            title, friendly_message = self.error_handler.format_error_for_qmessage(message, filename)

            error_box = QMessageBox(self)
            error_box.setIcon(QMessageBox.Critical)
            error_box.setWindowTitle(title)
            error_box.setText(friendly_message)
            error_box.setDetailedText(message)  # Original error message in details
            error_box.exec_()
        self.console_output.append(message)




    def show_instructions(self):

        instructions = """
1. Select all relevant input files:
   - Toast files (Order*.csv)
   - OLO files (Itemized_Orders*.csv, Transaction*.csv, Itemized_Cancelled*.csv)
   - GL file (GL*.csv)
   - Knock files (Billing*.xlsx)
   - Relay files (relay_carrotexpress*.xlsx)
   - The New company (Relacion*.xlsx)
   - Payroll file (Payroll*.csv)


2. Click RUN to process the reconciliation

The program will:
- Compare tips data across all systems
- Generate summary and discrepancy reports
- Highlight any mismatches or issues
- Create formatted Excel output files in your Downloads folder
        """
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Instructions")
        msg_box.setText(instructions)
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()

    def center(self):
        frame_geometry = self.frameGeometry()
        center_point = QApplication.desktop().availableGeometry().center()
        frame_geometry.moveCenter(center_point)
        self.move(frame_geometry.topLeft())

class TipsReconcileThread(QThread):

    update_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, input_files, output_file):  # Changed parameter name
        super().__init__()
        self.input_files = input_files
        self.output_file = output_file  # Store the full file path

        self.knock_mapping = {
            "Express Coral Glabes": "Coral Gables",
            "Boca Raton West": "West Boca",
            "Express Coconut Grove": "Coconut Grove",
            "Express Coconut Creek": "Coconut Creek",
            "Express Doral": "Doral",
            "Express Hollywood": "Hollywood",
            "Express Miami Garden": "Aventura (Miami Gardens)",
            "Express Miami Shores": "Miami Shores",
            "Express North Beach": "North Beach",
            "Express Pembroke Pines": "Pembroke Pines",
            "Express Plantation": "Plantation",
            "Express South Beach": "South Beach",
            "Boca Raton": "Boca Palmetto Park",
            "Express Brickell": "Brickell",
            "Express Midtown": "Midtown",
            "Express Dadeland": "Dadeland",
            "Express Downtown": "Downtown",
            "Express River landing": "River Landing",
            "Express Sunset": "South Miami (Sunset)"
        }

        self.olo_mapping = {
            "Carrot Express Aventura (Miami Gardens)": "Aventura (Miami Gardens)",
            "Carrot Express Aventura Mall": "Aventura Mall",
            "Carrot Express Boca Palmetto East": "Boca Palmetto Park",
            "Carrot Express Brickell": "Brickell",
            "Carrot Express Bryant Park": "Bryant Park",
            "Carrot Express Coconut Creek": "Coconut Creek",
            "Carrot Express Coconut Grove": "Coconut Grove",
            "Carrot Express Coral Gables": "Coral Gables",
            "Carrot Express Dadeland": "Dadeland",
            "Carrot Express Doral": "Doral",
            "Carrot Express Downtown Miami": "Downtown",
            "The Cushman School": "Downtown",
            "Carrot Express Flatiron": "Flatiron",
            "Carrot Express Hollywood": "Hollywood",
            "Carrot Express Las Olas": "Las Olas",
            "Carrot Express Lexington": "Lexington",
            "Carrot Express Miami Shores": "Miami Shores",
            "Carrot Express Midtown": "Midtown",
            "Carrot Express North Beach": "North Beach",
            "Carrot Express Pembroke Pines": "Pembroke Pines",
            "Carrot Express Plantation": "Plantation",
            "Carrot Express River Landing": "River Landing",
            "Carrot Express South Beach": "South Beach",
            "Carrot Express South Miami (Sunset)": "South Miami (Sunset)",
            "Carrot Express West Boca": "West Boca"
        }

        self.relay_mapping = {
            "bryantpark": "Bryant Park",
            "600ny": "Lexington",
            "flatiron": "Flatiron"
        }

        self.gl_mapping = {
            "Carrot Aventura Love LLC (Aventura)": "Aventura (Miami Gardens)",
            "Carrot Love Aventura Mall Operating LLC": "Aventura Mall",
            "Carrot Love Palmetto Park Operating LLC": "Boca Palmetto Park",
            "Carrot Love Brickell Operating LLC": "Brickell",
            "Carrot Love Bryant Park Operating LLC": "Bryant Park",
            "Carrot Love Coconut Creek Operating LLC": "Coconut Creek",
            "Carrot Love Coconut Grove Operating LLC": "Coconut Grove",
            "Carrot Coral GablesLove LLC (Coral Gabes)": "Coral Gables",
            "Carrot Love Dadeland Operating LLC": "Dadeland",
            "Carrot Love City Place Doral Operating LLC": "Doral",
            "Carrot Downtown Love Two LLC": "Downtown",
            "Carrot Flatiron Love Manhattan Operating LLC": "Flatiron",
            "Carrot Love Hollywood Operating LLC": "Hollywood",
            "Carrot Love Las Olas Operating LLC": "Las Olas",
            "Carrot Love 600 Lexington LLC": "Lexington",
            "Carrot Express Miami Shores LLC": "Miami Shores",
            "Carrot Express Midtown LLC": "Midtown",
            "Carrot North Beach Love LL (North Beach)": "North Beach",
            "Carrot Love Pembroke Pines Operating LLC": "Pembroke Pines",
            "Carrot Love Plantation Operating LLC": "Plantation",
            "Carrot Love River Lading Operating LLC": "River Landing",
            "Carrot Sobe Love South Florida Operating C LLC": "South Beach",
            "Carrot Love Sunset Operating LLC": "South Miami (Sunset)",
            "Carrot Love West Boca Operating LLC": "West Boca"
        }

    def safe_read_file(self, file_path, read_function='csv', **kwargs):
        """
        Safely read a file with error tracking and multiple encoding attempts.

        Args:
            file_path: Path to the file to read
            read_function: 'csv' or 'excel'
            **kwargs: Additional arguments to pass to read function

        Returns:
            pandas DataFrame
        """
        import pandas as pd
        filename = os.path.basename(file_path)
        self.update_signal.emit(f"Reading file: {filename}")

        try:
            if read_function == 'csv':
                # Remove encoding from kwargs if it exists
                kwargs.pop('encoding', None)

                # Try UTF-8 first
                try:
                    return pd.read_csv(file_path, encoding='utf-8', **kwargs)
                except UnicodeDecodeError:
                    # If UTF-8 fails, try latin1
                    return pd.read_csv(file_path, encoding='latin1', **kwargs)
            else:  # Excel
                return pd.read_excel(file_path, **kwargs)
        except Exception as e:
            raise ValueError(f"Error reading file {filename}: {str(e)}")

    def run(self):
        import pandas as pd
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        try:
            self.update_signal.emit("Starting reconciliation process...")

            # Get the directory paths
            first_file = self.input_files[0]
            base_directory = os.path.dirname(first_file)

            # Process files using your existing functions
            date_range = None
            if any(os.path.basename(f).lower().startswith('order') for f in self.input_files):
                (date_range, olo_tips, knock_tips, toast_relay_tips,
                 toast_metro_speedy_tips, toast_new_company_tips,
                 total_delivery, employee_tips) = self.read_toast_file(base_directory)

            knock_data = self.read_knock_files(base_directory, date_range, self.knock_mapping) if date_range else {}
            olo_data = self.read_olo_file(base_directory, self.olo_mapping, date_range) if date_range else {}
            r365_emp_tips, r365_del_tips = self.read_gl_file(base_directory, self.gl_mapping)
            relay_data = self.read_relay_file(base_directory, self.relay_mapping, date_range) if date_range else pd.DataFrame()
            metro_speedy_data = self.read_metro_speedy_file(base_directory, date_range) if date_range else pd.DataFrame()
            new_company_data = self.read_new_company_file(base_directory, date_range) if date_range else pd.DataFrame()

            # Initialize payroll variables
            payroll_emp_tips, payroll_del_tips = self.read_payroll_file(base_directory)

            # Create summary file with new company data
            summary = self.create_summary_file(
                date_range, total_delivery, employee_tips,
                olo_data, knock_data, relay_data, new_company_data,
                r365_emp_tips, r365_del_tips,
                payroll_emp_tips, payroll_del_tips
            )

            # Create discrepancy file with new company data
            discrepancy = self.create_discrepancy_file_with_relay(
                olo_tips, knock_tips, toast_relay_tips, toast_metro_speedy_tips,
                toast_new_company_tips, olo_data, knock_data, relay_data,
                metro_speedy_data, new_company_data
            )

            # Create workbook and add sheets
            wb = openpyxl.Workbook()

            # Setup Discrepancy sheet
            ws_discrepancy = wb.active
            ws_discrepancy.title = 'Tips Discrepancy'

            # Write discrepancy data to sheet
            for r_idx, row in enumerate(dataframe_to_rows(discrepancy, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_discrepancy.cell(row=r_idx, column=c_idx, value=value)

            # Format discrepancy sheet
            self.format_discrepancy_excel(discrepancy, self.output_file, ws_discrepancy)

            # Setup Summary sheet
            ws_summary = wb.create_sheet('Tips Summary')

            # Write summary data to sheet
            for r_idx, row in enumerate(dataframe_to_rows(summary, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_summary.cell(row=r_idx, column=c_idx, value=value)

            # Format summary sheet
            self.format_summary_excel(summary, self.output_file, ws_summary)

            # Save the workbook
            wb.save(self.output_file)

            success_message = f"Reconciliation completed! File saved to:\n{self.output_file}"
            self.finished_signal.emit(True, success_message)

        except Exception as e:
            import traceback
            self.update_signal.emit(f"Error details: {traceback.format_exc()}")
            self.finished_signal.emit(False, f"An error occurred: {str(e)}")

    def read_toast_file(self, directory):
        import pandas as pd
        def validate_toast_file(df, filename):
            """
            Validate required columns in Toast files.

            Args:
                df: pandas DataFrame to validate
                filename: name of the file being validated

            Raises:
                ValueError if required columns are missing
            """
            required_columns = [
                "Order #",
                "Location",
                "Opened",
                "Tab Names",
                "Server",
                "Dining Options",
                "Tip",
                "Gratuity"
            ]

            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                missing_cols_str = ", ".join(missing_columns)
                raise ValueError(
                    f"Missing required columns in Toast file '{filename}': {missing_cols_str}\n\n"
                    f"To fix this:\n"
                    f"1. Go to Toast website\n"
                    f"2. When downloading the Order Details report, make sure to include all required columns\n"
                    f"3. Required columns are: {', '.join(required_columns)}"
                )

        # Filter for files
        toast_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('order')]
        itemized_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('itemized_orders')]
        transaction_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('transaction')]
        cancelled_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('itemized_cancelled')]

        # Check for required files
        if not toast_files:
            raise ValueError("Toast file not found")

        # Check all required OLO files
        missing_files = []
        if not itemized_files:
            missing_files.append("Itemized Orders (starts with 'Itemized_Orders')")
        if not transaction_files:
            missing_files.append("Transaction (starts with 'Transaction')")
        if not cancelled_files:
            missing_files.append("Itemized Cancelled Orders (starts with 'Itemized_Cancelled')")

        if missing_files:
            missing_files_str = "\n   - ".join(missing_files)
            raise ValueError(
                "Required OLO files (Itemized or Transaction) not found. "
                f"Missing files:\n   - {missing_files_str}"
            )

        # Read OLO files with error tracking
        olo_df = self.safe_read_file(itemized_files[0], read_function='csv')
        trans_df = self.safe_read_file(transaction_files[0], read_function='csv')

        # Get cancelled orders if file exists
        cancelled_order_ids = set()
        if cancelled_files:
            cancelled_df = self.safe_read_file(cancelled_files[0], read_function='csv')
            cancelled_order_ids = set(cancelled_df['Order ID'].astype(str))

        # Get all Order IDs that have RefundSale or VoidSale
        refund_void_orders = trans_df[
            trans_df['Transaction Type'].isin(['RefundSale', 'VoidSale'])
        ]['Order ID'].unique()

        # Filter out refunded/voided orders from the itemized data
        olo_df = olo_df[~olo_df['Order ID'].isin(refund_void_orders)]
        # Also filter out cancelled orders
        olo_df = olo_df[~olo_df['Order ID'].astype(str).isin(cancelled_order_ids)]

        # Create a mapping of Order ID to Type for Google orders
        google_order_types = olo_df[['Order ID', 'Type']].drop_duplicates()
        google_order_types['Order ID'] = google_order_types['Order ID'].astype(str)
        google_delivery_orders = set(google_order_types[google_order_types['Type'].isin(['Delivery', 'Dispatch'])]['Order ID'])

        # Process toast files
        all_toast_data = []
        for file in toast_files:
            df = self.safe_read_file(file, read_function='csv', encoding='latin1')
            # Validate columns before processing
            validate_toast_file(df, os.path.basename(file))
            df['Order #'] = df['Order #'].astype(str)
            df = df[~df['Order #'].isin(cancelled_order_ids)]
            all_toast_data.append(df)

        # Combine all dataframes
        df = pd.concat(all_toast_data, ignore_index=True)

        # Process data
        df['Order #'] = df['Order #'].astype(str)
        # Convert the date with the correct format for AM/PM time
        def parse_date(date_str):
            try:
                # First try full year format without AM/PM
                return pd.to_datetime(date_str, format='%m/%d/%Y %H:%M').strftime('%m/%d/%Y')
            except:
                try:
                    # Then try two-digit year with AM/PM
                    return pd.to_datetime(date_str, format='%m/%d/%y %I:%M %p').strftime('%m/%d/%Y')
                except:
                    # Fallback to letting pandas infer the format
                    return pd.to_datetime(date_str).strftime('%m/%d/%Y')

        df['Date'] = df['Opened'].apply(parse_date)
        df['Tips_Total'] = df['Tip'] + df['Gratuity']

        # Get overall date range from combined data
        start_date = df['Date'].min()
        end_date = df['Date'].max()
        date_range = (start_date, end_date)

        # Define locations
        ny_locations = ['Bryant Park', 'Lexington', 'Flatiron']
        new_company_locations = ['South Beach']  # 'North Beach' is now treated as a normal location

        # For regular locations:
        standard_olo_mask = (
            ~df['Location'].isin(new_company_locations) &
            df['Dining Options'].isin(['Google Online (Dispatch)', 'Online Ordering (Dispatch) *'])
        )
        google_delivery_mask = (
            ~df['Location'].isin(new_company_locations) &
            (df['Dining Options'] == 'Google Online Ordering') &
            df['Order #'].isin(google_delivery_orders)
        )
        special_delivery_mask = (
            ~df['Location'].isin(new_company_locations) &
            df['Dining Options'].isin(['Online Ordering (Delivery) *', 'Telephone - Delivery'])
        )

        # For North Beach and South Beach:
        new_company_mask = (
            df['Location'].isin(new_company_locations) &
            (
                df['Dining Options'].isin(['Google Online (Dispatch)', 'Online Ordering (Dispatch) *']) |
                (
                    (df['Dining Options'] == 'Google Online Ordering') &
                    df['Order #'].isin(google_delivery_orders)
                ) |
                df['Dining Options'].isin(['Online Ordering (Delivery) *', 'Telephone - Delivery'])
            )
        )

        # Calculate regular OLO tips
        standard_olo = df[standard_olo_mask].groupby(['Location', 'Date'])['Gratuity'].sum().reset_index()
        google_delivery_tips = df[google_delivery_mask].groupby(['Location', 'Date']).apply(
            lambda x: x['Tip'] + x['Gratuity']
        ).reset_index(name='Toast OLO Delivery Tips')
        special_delivery = df[special_delivery_mask].groupby(['Location', 'Date'])['Tip'].sum().reset_index()

        # Calculate new company tips - MODIFIED to sum all tips for each day
        new_company_tips = df[new_company_mask].groupby(['Location', 'Date']).agg({
            'Tip': 'sum',
            'Gratuity': 'sum'
        }).reset_index()
        new_company_tips['Toast New Company Delivery Tips'] = new_company_tips['Tip'] + new_company_tips['Gratuity']
        new_company_tips = new_company_tips[['Location', 'Date', 'Toast New Company Delivery Tips']]

        # Combine all OLO tips (excluding new company locations)
        olo_tips = pd.concat([
            standard_olo.rename(columns={'Gratuity': 'Toast OLO Delivery Tips'}),
            google_delivery_tips,
            special_delivery.rename(columns={'Tip': 'Toast OLO Delivery Tips'})
        ]).groupby(['Location', 'Date'])['Toast OLO Delivery Tips'].sum().reset_index()

        # Knock tips calculation (excluding NY locations)
        knock_mask = df['Dining Options'].isin(['Olo Catering (Self-Delivery))', 'EZ Cater (Delivery)']) & \
                    ~df['Location'].isin(ny_locations)
        knock_tips = df[knock_mask].groupby(['Location', 'Date'])['Tip'].sum().reset_index()
        knock_tips = knock_tips.rename(columns={'Tip': 'Toast Knock Delivery Tips'})

        # Metro Speedy tips calculation (NY locations only)
        metro_speedy_mask = df['Dining Options'].isin(['Olo Catering (Self-Delivery))', 'EZ Cater (Delivery)']) & \
                        df['Location'].isin(ny_locations)
        metro_speedy_tips = df[metro_speedy_mask].groupby(['Location', 'Date'])['Tip'].sum().reset_index()
        metro_speedy_tips = metro_speedy_tips.rename(columns={'Tip': 'Toast Metro Speedy Delivery Tips'})

        # Relay tips calculation
        relay_mask = df['Dining Options'] == 'Grubhub (Delivery)'
        relay_tips = df[relay_mask].groupby(['Location', 'Date'])['Tip'].sum().reset_index()
        relay_tips = relay_tips.rename(columns={'Tip': 'Toast Relay Delivery Tips'})

        # Calculate total delivery tips
        total_delivery = pd.concat([
            olo_tips,
            knock_tips,
            metro_speedy_tips,
            relay_tips,
            new_company_tips
        ]).groupby('Location')[
            ['Toast OLO Delivery Tips', 'Toast Knock Delivery Tips',
            'Toast Metro Speedy Delivery Tips', 'Toast Relay Delivery Tips',
            'Toast New Company Delivery Tips']
        ].sum().reset_index()

        total_delivery = total_delivery.fillna(0)
        total_delivery['Total Toast Delivery Tips'] = (
            total_delivery['Toast OLO Delivery Tips'] +
            total_delivery['Toast Knock Delivery Tips'] +
            total_delivery['Toast Metro Speedy Delivery Tips'] +
            total_delivery['Toast Relay Delivery Tips'] +
            total_delivery['Toast New Company Delivery Tips']
        )

        # Set OLO tips to 0 for new company locations
        olo_tips.loc[olo_tips['Location'].isin(new_company_locations), 'Toast OLO Delivery Tips'] = 0

        # Calculate employee tips
        all_tips = df.groupby(['Location'])['Tips_Total'].sum().reset_index()
        employee_tips = all_tips.copy()
        employee_tips['Total Toast Employee Tips'] = employee_tips['Tips_Total'] - total_delivery['Total Toast Delivery Tips']
        employee_tips = employee_tips[['Location', 'Total Toast Employee Tips']]

        return date_range, olo_tips, knock_tips, relay_tips, metro_speedy_tips, new_company_tips, total_delivery, employee_tips



    def read_metro_speedy_file(self, directory, toast_date_range):
        import pandas as pd
        start_date, end_date = pd.to_datetime(toast_date_range[0]), pd.to_datetime(toast_date_range[1])
        metro_speedy_data = pd.DataFrame(columns=['Date', 'Tip', 'Location'])

        # Filter for Metro Speedy files
        metro_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('metro_speedy')]

        if not metro_files:
            self.update_signal.emit("No Metro Speedy files found")
            return metro_speedy_data

        all_metro_data = []
        for file in metro_files:
            try:
                # Read with error tracking
                data = self.safe_read_file(file, read_function='excel')
                data['Date'] = pd.to_datetime(data['Date']).dt.strftime('%m/%d/%Y')

                # Filter data within toast date range
                data['DateForFilter'] = pd.to_datetime(data['Date'])
                data = data[(data['DateForFilter'] >= start_date) & (data['DateForFilter'] <= end_date)]
                all_metro_data.append(data)
            except Exception as e:
                self.update_signal.emit(f"Error processing Metro Speedy file {os.path.basename(file)}: {str(e)}")
                raise  # Re-raise the exception to be caught by the main error handler

        if all_metro_data:
            metro_speedy_data = pd.concat(all_metro_data, ignore_index=True)

        return metro_speedy_data


    def read_knock_files(self, directory, toast_date_range, knock_mapping):
        import pandas as pd
        start_date, end_date = pd.to_datetime(toast_date_range[0]), pd.to_datetime(toast_date_range[1])
        tips_data = {}

        # Filter for Billing files from selected files
        billing_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('billing')]

        for file in billing_files:
            try:
                # Read with error tracking
                first_row = self.safe_read_file(file, read_function='excel', nrows=1)
                location = first_row.columns[0].replace("Carrot ", "")
                location = knock_mapping.get(location, location)

                # Read and clean data
                data = self.safe_read_file(file, read_function='excel', skiprows=2)
                data = self.clean_file_content(data)  # Now passing DataFrame instead of file path

                # Filter data within toast date range
                data['Date'] = pd.to_datetime(data['Date'])
                data = data[(data['Date'] >= start_date) & (data['Date'] <= end_date)]
                grouped_tips = data.groupby(data['Date'].dt.strftime('%m/%d/%Y'))['Tip'].sum().to_dict()

                if location in tips_data:
                    for date, tip in grouped_tips.items():
                        tips_data[location][date] = tips_data[location].get(date, 0) + tip
                else:
                    tips_data[location] = grouped_tips
            except Exception as e:
                self.update_signal.emit(f"Error processing file {os.path.basename(file)}: {e}")

        return tips_data

    def clean_file_content(self, data):
        """
        Clean and process a DataFrame

        Args:
            data: pandas DataFrame to clean

        Returns:
            cleaned pandas DataFrame
        """
        import pandas as pd

        # Remove completely empty rows and columns
        data = data.dropna(how='all', axis=0).dropna(how='all', axis=1)

        # Clean column names
        data.columns = data.columns.str.strip()
        data.columns = data.columns.str.replace(r'\s+', ' ', regex=True)

        if 'Date' in data.columns:
            # Remove TOTALS row and convert dates
            data = data[data['Date'].notna() & (data['Date'] != 'TOTALS')]
            data['Date'] = pd.to_datetime(data['Date'], errors='coerce').dt.strftime('%m/%d/%Y')
            data = data[data['Date'].notna()]

        # Convert numeric columns
        for col in ['Tip', 'Fee', 'Original Amount', 'Total to pay']:
            if col in data.columns:
                data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0)

        return data

    def read_olo_file(self, directory, location_mapping, toast_date_range):
        import pandas as pd
        start_date, end_date = pd.to_datetime(toast_date_range[0]), pd.to_datetime(toast_date_range[1])

        # Find required files
        itemized_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('itemized_orders')]
        transaction_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('transaction')]
        cancelled_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('itemized_cancelled')]

        if not itemized_files or not transaction_files:
            raise ValueError("Required OLO files (Itemized or Transaction) not found")

        # Read files with error tracking
        df = self.safe_read_file(itemized_files[0], read_function='csv')
        trans_df = self.safe_read_file(transaction_files[0], read_function='csv')

        # Get cancelled orders if file exists
        cancelled_order_ids = set()
        if cancelled_files:
            cancelled_df = self.safe_read_file(cancelled_files[0], read_function='csv')
            cancelled_order_ids = set(cancelled_df['Order ID'].astype(str))

        # Filter out refunded/voided/cancelled orders
        refund_void_orders = trans_df[
            trans_df['Transaction Type'].isin(['RefundSale', 'VoidSale'])
        ]['Order ID'].unique()
        df = df[~df['Order ID'].isin(refund_void_orders)]
        df = df[~df['Order ID'].astype(str).isin(cancelled_order_ids)]

        # Filter for Dispatch and Delivery orders
        df = df[df['Type'].isin(['Dispatch', 'Delivery'])]

        # Define NY locations and new company locations
        ny_locations = ['Bryant Park', 'Lexington', 'Flatiron']
        new_company_locations = ['South Beach']

        def get_date(row):
            if 'Immediate' in str(row['Time Wanted']):
                return pd.to_datetime(row['Time Placed']).strftime('%m/%d/%Y')
            return pd.to_datetime(row['Time Wanted']).strftime('%m/%d/%Y')

        df['Date'] = df.apply(get_date, axis=1)
        df['DateForFilter'] = pd.to_datetime(df['Date'])
        df = df[(df['DateForFilter'] >= start_date) & (df['DateForFilter'] <= end_date)]
        df['Location'] = df['Store Name'].map(location_mapping)

        # Create masks for filtering
        ny_delivery_mask = (df['Location'].isin(ny_locations)) & (df['Type'] == 'Delivery')
        catering_mask = df['Store Name'].str.contains('Catering', na=False)
        non_ny_mask = ~df['Location'].isin(ny_locations)
        new_company_mask = df['Location'].isin(new_company_locations)

        # Create a copy of the dataframe for NY employee tips
        ny_employee_df = df[ny_delivery_mask].copy()

        # Filter out catering orders for non-NY locations
        delivery_df = df[
            ~(catering_mask & non_ny_mask) &  # Exclude catering orders for non-NY locations
            ~ny_delivery_mask  # Remove NY delivery orders as before
        ]

        # Calculate regular delivery tips (for non-new company locations)
        regular_delivery_tips = delivery_df[~new_company_mask].groupby(['Location', 'Date'])['Tip'].sum().reset_index()

        # Calculate new company delivery tips (using same logic as regular OLO)
        new_company_delivery_tips = delivery_df[new_company_mask].groupby(['Location', 'Date'])['Tip'].sum().reset_index()

        # Combine all delivery tips
        delivery_tips = pd.concat([regular_delivery_tips, new_company_delivery_tips])

        # Calculate NY employee tips
        ny_employee_tips = ny_employee_df.groupby(['Location', 'Date'])['Tip'].sum().reset_index()
        ny_employee_tips = ny_employee_tips.rename(columns={'Tip': 'Employee_Tips'})

        return {'delivery_tips': delivery_tips, 'ny_employee_tips': ny_employee_tips}

    def read_gl_file(self, directory, location_mapping):
        import pandas as pd
        gl_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('gl')]
        if not gl_files:
            raise ValueError("GL file not found")

        df = self.safe_read_file(gl_files[0], read_function='csv', skiprows=3, thousands=',')
        df['LocationName1'] = df['LocationName1'].map(location_mapping)
        emp_tips = df[df['ParentAccountName'] == '21250 - Employee Tips Payable'].groupby('LocationName1')['Credit1'].first()
        del_tips = df[df['ParentAccountName'] == '21251 - Payable Delivery Tips'].groupby('LocationName1')['Credit1'].first()
        return emp_tips.rename('R365 Employee Tips Payable'), del_tips.rename('R365 Delivery Tips Payable')

    def read_relay_file(self, directory, relay_mapping, toast_date_range):
        import pandas as pd
        start_date, end_date = pd.to_datetime(toast_date_range[0]), pd.to_datetime(toast_date_range[1])

        relay_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('relay_carrotexpress')]
        all_relay_tips = []

        for file in relay_files:
            try:
                file_name = os.path.basename(file)
                location_key = file_name.split("relay_carrotexpress")[1].split("_invoice")[0].lower()
                location = relay_mapping.get(location_key)
                if not location:
                    continue

                # Read with error tracking
                data = self.safe_read_file(file, read_function='excel', sheet_name=1)
                data['Date'] = pd.to_datetime(data['Time']).dt.strftime('%m/%d/%Y')
                # Filter data within toast date range
                data['DateForFilter'] = pd.to_datetime(data['Date'])
                data = data[(data['DateForFilter'] >= start_date) & (data['DateForFilter'] <= end_date)]
                data = data[data['Status'] != "VOIDED"]

                daily_tips = data.groupby('Date')['Tip'].sum().reset_index()
                daily_tips['Location'] = location
                all_relay_tips.append(daily_tips)

            except Exception as e:
                self.update_signal.emit(f"Error processing relay file {file}: {e}")

        if all_relay_tips:
            return pd.concat(all_relay_tips, ignore_index=True)
        else:
            return pd.DataFrame(columns=['Date', 'Tip', 'Location'])

    def read_payroll_file(self, directory):
        import pandas as pd

        # Filter for Payroll files
        payroll_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('payroll')]

        if not payroll_files:
            self.update_signal.emit("No Payroll file found")
            return {}, {}

        # Read with error tracking
        df = self.safe_read_file(payroll_files[0], read_function='csv')

        # Initialize dictionaries to store tips by location
        employee_tips = {}
        delivery_tips = {}

        # Process each row
        current_location = None
        for _, row in df.iterrows():
            # Get location from column A
            if pd.notna(row.iloc[0]):
                current_location = row.iloc[0]

            # Check column B for indicators
            indicator = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""

            if indicator == "Total Tips  DIA repartir":
                # Get employee tips from column AZ
                tips_amount = float(row.iloc[51]) if pd.notna(row.iloc[51]) else 0
                if current_location:
                    employee_tips[current_location] = tips_amount

            elif indicator == "Delivery":
                # Get delivery tips from column AZ
                tips_amount = float(row.iloc[51]) if pd.notna(row.iloc[51]) else 0
                if current_location:
                    delivery_tips[current_location] = tips_amount

        return employee_tips, delivery_tips

    def read_new_company_file(self, directory, toast_date_range):
        """Read the new company file (Relacion para Carrot Express.xlsx)"""
        import pandas as pd
        new_company_data = pd.DataFrame(columns=['Date', 'Tip', 'Location'])

        # Filter for new company files
        new_company_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('relacion')]

        if not new_company_files:
            self.update_signal.emit("No New Company files found")
            return new_company_data

        location_mapping = {
            "Resumen entregas Carrot North Beach": "North Beach",
            "Resumen entregas Carrot South Beach": "South Beach"
        }

        all_new_company_data = []
        for file in new_company_files:
            try:
                # Read first sheet to get location with error tracking
                first_sheet = self.safe_read_file(file, read_function='excel', sheet_name=0, nrows=1)
                location = first_sheet.columns[0]

                # Map location name
                for key in location_mapping:
                    if key in location:
                        location = location_mapping[key]
                        break

                # Read second sheet for tip data with error tracking
                data = self.safe_read_file(file, read_function='excel', sheet_name=1, skiprows=1)
                data = data[['Fecha', 'Tip']].copy()
                data['Location'] = location
                data = data.rename(columns={'Fecha': 'Date'})

                # Convert date format
                data['Date'] = pd.to_datetime(data['Date']).dt.strftime('%m/%d/%Y')

                # Filter data within toast date range if provided
                if toast_date_range:
                    start_date, end_date = pd.to_datetime(toast_date_range[0]), pd.to_datetime(toast_date_range[1])
                    data['DateForFilter'] = pd.to_datetime(data['Date'])
                    data = data[(data['DateForFilter'] >= start_date) & (data['DateForFilter'] <= end_date)]
                    data = data.drop('DateForFilter', axis=1)

                all_new_company_data.append(data)

            except Exception as e:
                self.update_signal.emit(f"Error processing New Company file {os.path.basename(file)}: {str(e)}")
                raise  # Re-raise the exception to be caught by the main error handler

        if all_new_company_data:
            new_company_data = pd.concat(all_new_company_data, ignore_index=True)
            # Group by location and date to get totals
            new_company_data = new_company_data.groupby(['Location', 'Date'])['Tip'].sum().reset_index()

        return new_company_data


    def create_summary_file(self, date_range, total_delivery, employee_tips, olo_payable_data, knock_payable,
                       relay_payable, new_company_data, r365_emp_tips, r365_del_tips,
                       payroll_emp_tips, payroll_del_tips):
        import pandas as pd

        # Extract delivery tips from olo_payable_data dictionary
        olo_payable = olo_payable_data['delivery_tips']
        ny_employee_tips = olo_payable_data['ny_employee_tips']

        excluded_locations = ["Weston", "West Kendall (London Square)", "Pinecrest"]
        total_delivery = total_delivery[~total_delivery['Location'].isin(excluded_locations)]
        employee_tips = employee_tips[~employee_tips['Location'].isin(excluded_locations)]
        olo_payable = olo_payable[~olo_payable['Location'].isin(excluded_locations)]
        knock_payable = {k: v for k, v in knock_payable.items() if k not in excluded_locations}
        r365_emp_tips = r365_emp_tips[~r365_emp_tips.index.isin(excluded_locations)]
        r365_del_tips = r365_del_tips[~r365_del_tips.index.isin(excluded_locations)]

        formatted_date_range = f"{date_range[0]} - {date_range[1]}"
        summary = pd.merge(total_delivery, employee_tips, on='Location')
        summary['Date range'] = formatted_date_range

        # Calculate third party tips by location and individual totals
        third_party_by_location = {}
        olo_totals = {}
        knock_totals = {}
        relay_totals = {}
        new_company_totals = {}

        # Get available locations in new_company_data
        new_company_locations = set(new_company_data['Location'].unique()) if not new_company_data.empty else set()

        for location in summary['Location']:
            # Calculate OLO total
            if location not in ['South Beach']:
                olo_total = olo_payable[olo_payable['Location'] == location]['Tip'].sum()
            else:
                olo_total = 0
            olo_totals[location] = olo_total

            # Calculate Knock total
            knock_total = sum(knock_payable.get(location, {}).values())
            knock_totals[location] = knock_total

            # Calculate Relay total
            relay_total = relay_payable[relay_payable['Location'] == location]['Tip'].sum() if not relay_payable.empty else 0
            relay_totals[location] = relay_total

            # Calculate New Company total based on availability of Relacion file
            if location in ['South Beach']:
                if location in new_company_locations:
                    # Use data from Relacion file
                    new_company_total = new_company_data[new_company_data['Location'] == location]['Tip'].sum()
                else:
                    # Fall back to OLO data
                    new_company_total = olo_payable[olo_payable['Location'] == location]['Tip'].sum()
            else:
                new_company_total = 0
            new_company_totals[location] = new_company_total

            # Calculate total third party tips
            third_party_by_location[location] = (
                olo_total + knock_total + relay_total + new_company_total
            )

        # Add all totals to summary DataFrame
        summary['Total 3rd Party Delivery Tips'] = summary['Location'].map(third_party_by_location)
        summary['Total OLO'] = summary['Location'].map(olo_totals)
        summary['Total Knock Knock'] = summary['Location'].map(knock_totals)
        summary['Total Relay'] = summary['Location'].map(relay_totals)
        summary['Total The New Company'] = summary['Location'].map(new_company_totals)

        summary = summary.merge(pd.DataFrame(r365_emp_tips), left_on='Location', right_index=True, how='left')
        summary = summary.merge(pd.DataFrame(r365_del_tips), left_on='Location', right_index=True, how='left')
        summary = summary.fillna(0)

        summary['Difference Toast - 3P Delivery Tips'] = (
            summary['Total Toast Delivery Tips'] - summary['Total 3rd Party Delivery Tips']
        )
        summary['Difference Toast - R365 Delivery Tips'] = (
            summary['Total Toast Delivery Tips'] - summary['R365 Delivery Tips Payable']
        )
        summary['Difference Toast - R365 Employee Tips'] = (
            summary['Total Toast Employee Tips'] - summary['R365 Employee Tips Payable']
        )

        # Round all numeric columns to 2 decimal places
        numeric_columns = summary.select_dtypes(include=['float64', 'int64']).columns
        summary[numeric_columns] = summary[numeric_columns].round(2)

        summary.loc[abs(summary['Difference Toast - 3P Delivery Tips']) < 0.01, 'Difference Toast - 3P Delivery Tips'] = 0
        summary.loc[abs(summary['Difference Toast - R365 Delivery Tips']) < 0.01, 'Difference Toast - R365 Delivery Tips'] = 0
        summary.loc[abs(summary['Difference Toast - R365 Employee Tips']) < 0.01, 'Difference Toast - R365 Employee Tips'] = 0

        # Add payroll comparisons
        summary['Toast - Payroll: Employee Tips'] = 0.0
        summary['Toast - Payroll: Delivery Tips'] = 0.0

        for index, row in summary.iterrows():
            location = row['Location']
            if location in payroll_emp_tips:
                summary.at[index, 'Toast - Payroll: Employee Tips'] = (
                    row['Total Toast Employee Tips'] - payroll_emp_tips[location]
                )
            if location in payroll_del_tips:
                summary.at[index, 'Toast - Payroll: Delivery Tips'] = (
                    row['Total Toast Delivery Tips'] - payroll_del_tips[location]
                )

        columns = [
            'Location', 'Date range', 'Total Toast Delivery Tips', 'Total Toast Employee Tips',
            'R365 Employee Tips Payable', 'R365 Delivery Tips Payable', 'Total 3rd Party Delivery Tips',
            'Total OLO', 'Total Knock Knock', 'Total Relay', 'Total The New Company',
            'Difference Toast - 3P Delivery Tips', 'Difference Toast - R365 Delivery Tips',
            'Difference Toast - R365 Employee Tips', 'Toast - Payroll: Employee Tips',
            'Toast - Payroll: Delivery Tips'
        ]

        return summary[columns]

    def create_discrepancy_file_with_relay(self, olo_tips, knock_tips, toast_relay_tips,
                                    toast_metro_speedy_tips, toast_new_company_tips,
                                    olo_data, knock_payable, relay_payable,
                                    metro_speedy_data, new_company_data):
        import pandas as pd
        import re
        # Get the components from olo_data dictionary
        olo_payable = olo_data['delivery_tips']
        ny_employee_tips = olo_data['ny_employee_tips']

        new_company_locations = ['South Beach']
        excluded_locations = ["Weston", "West Kendall (London Square)", "Pinecrest"]
        ny_locations = ['Bryant Park', 'Lexington', 'Flatiron']

        olo_tips = olo_tips[~olo_tips['Location'].isin(excluded_locations)]
        knock_tips = knock_tips[~knock_tips['Location'].isin(excluded_locations)]
        olo_payable = olo_payable[~olo_payable['Location'].isin(excluded_locations)]
        knock_payable = {k: v for k, v in knock_payable.items() if k not in excluded_locations}

        # Get toast data from all toast files
        toast_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('order')]
        relay_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('relay_carrotexpress')]
        itemized_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('itemized_orders')]
        transaction_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('transaction')]
        cancelled_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('itemized_cancelled')]

        if not toast_files:
            raise ValueError("Toast file not found")

        # Initialize list to store dataframes from each file
        all_toast_data = []
        all_relay_data = []

        # Process toast files
        for file in toast_files:
            df = pd.read_csv(file, encoding='latin1')
            df['Date'] = pd.to_datetime(df['Opened']).dt.strftime('%m/%d/%Y')
            df['Tips_Total'] = df['Tip'] + df['Gratuity']
            all_toast_data.append(df)

        # Process relay files
        for file in relay_files:
            try:
                data = pd.read_excel(file, sheet_name=1)
                location_key = os.path.basename(file).split("relay_carrotexpress")[1].split("_invoice")[0].lower()
                location = self.relay_mapping.get(location_key)
                if location:
                    data['Location'] = location
                    data['Date'] = pd.to_datetime(data['Time']).dt.strftime('%m/%d/%Y')
                    all_relay_data.append(data)
            except Exception as e:
                self.update_signal.emit(f"Error processing relay file {file}: {e}")

        # Combine all toast data
        toast_df = pd.concat(all_toast_data, ignore_index=True)
        relay_df = pd.concat(all_relay_data, ignore_index=True) if all_relay_data else pd.DataFrame()

        # Process OLO refunds and cancellations
        refund_cancelled_orders_dict = {}
        if itemized_files and transaction_files:
            olo_df = pd.read_csv(itemized_files[0])
            trans_df = pd.read_csv(transaction_files[0])

            # Get cancelled orders from cancelled file
            cancelled_orders_info = {}
            if cancelled_files:
                cancelled_df = pd.read_csv(cancelled_files[0])
                # Create a mapping of Order ID to status "Cancelled"
                for _, row in cancelled_df.iterrows():
                    cancelled_orders_info[str(row['Order ID'])] = "Cancelled"

            # Get refunded orders from transaction file
            refunded_orders_info = {}
            for _, row in trans_df[trans_df['Transaction Type'].isin(['RefundSale', 'VoidSale'])].iterrows():
                refunded_orders_info[str(row['Order ID'])] = "Refunded/Adjusted"

            # Combine all problematic orders
            problematic_orders = set(cancelled_orders_info.keys()).union(set(refunded_orders_info.keys()))

            # Create a filtered DataFrame with problematic orders, excluding Pickup orders
            filtered_problematic_df = olo_df[
                (olo_df['Order ID'].astype(str).isin(problematic_orders)) &
                (olo_df['Type'] != 'Pickup')
            ]

            # Group orders by location and date
            filtered_problematic_df['Date'] = pd.to_datetime(filtered_problematic_df['Time Placed']).dt.strftime('%m/%d/%Y')

            for _, row in filtered_problematic_df.iterrows():
                location = self.olo_mapping.get(row['Store Name'])
                if location:
                    key = (location, row['Date'])
                    if key not in refund_cancelled_orders_dict:
                        refund_cancelled_orders_dict[key] = []

                    order_id = str(row['Order ID'])
                    # Determine if it's cancelled or refunded
                    if order_id in cancelled_orders_info:
                        refund_cancelled_orders_dict[key].append(f"OLO#{order_id}-Cancelled")
                    elif order_id in refunded_orders_info:
                        refund_cancelled_orders_dict[key].append(f"OLO#{order_id}-Refunded/Adjusted")
                    else:
                        # Fallback (shouldn't happen, but just in case)
                        refund_cancelled_orders_dict[key].append(f"OLO#{order_id}")

            # Get available locations in new_company_data
            new_company_locations_with_data = set(new_company_data['Location'].unique()) if not new_company_data.empty else set()

            # Find orders with incorrect dining options
            incorrect_dining_mask = (
                (
                    # Regular online/google orders
                    toast_df['Dining Options'].str.contains('Online|Google', case=False, na=False) &
                    # Exclude when it's OLO Catering with Default Online Ordering server
                    ~(
                        (toast_df['Dining Options'] == 'Olo Catering (Self-Delivery))') &
                        (toast_df['Server'] == 'Default Online Ordering')
                    ) &
                    ~toast_df['Server'].isin(['Default Online Ordering', 'Online Ordering Online Ordering', '']) &
                    toast_df['Server'].notna()
                )
            )

            # Process incorrect dining options for each location and date
            incorrect_orders_dict = {}
            relay_voided_orders_dict = {}
            relay_mismatch_dict = {}  # New dictionary for relay mismatches

            for location in toast_df['Location'].unique():
                for date in toast_df[toast_df['Location'] == location]['Date'].unique():
                    incorrect_orders = toast_df[
                        (toast_df['Location'] == location) &
                        (toast_df['Date'] == date) &
                        incorrect_dining_mask
                    ]['Order #'].astype(str).tolist()

                    # For NY locations, check relay names against toast tab names
                    if location in ny_locations and not relay_df.empty:
                        # Track voided relay orders
                        voided_orders = relay_df[
                            (relay_df['Location'] == location) &
                            (relay_df['Date'] == date) &
                            (relay_df['Status'] == "VOIDED")
                        ]

                        if not voided_orders.empty:
                            voided_list = []
                            # Get matching toast orders for the same location/date
                            toast_orders = toast_df[
                                (toast_df['Location'] == location) &
                                (toast_df['Date'] == date) &
                                (toast_df['Dining Options'] == 'Grubhub (Delivery)')
                            ]

                            for _, voided_order in voided_orders.iterrows():
                                consumer_name = str(voided_order['Consumer'])
                                if pd.notna(consumer_name) and consumer_name.strip():
                                    name_parts = consumer_name.strip().split()
                                    if len(name_parts) >= 2:
                                        first_name = name_parts[0].lower()
                                        last_initial = name_parts[1][0].lower()
                                        name_pattern = f"{re.escape(first_name)} {re.escape(last_initial)}.*grubhub"

                                        # Find matching toast order
                                        matching_toast_order = None
                                        for _, toast_row in toast_orders.iterrows():
                                            tab_name = str(toast_row['Tab Names']).lower()
                                            if pd.notna(tab_name) and re.search(name_pattern, tab_name):
                                                matching_toast_order = toast_row['Order #']
                                                break

                                        if matching_toast_order:
                                            voided_list.append(f"Relay#{voided_order['ID']} (Toast#{matching_toast_order})")
                                        else:
                                            voided_list.append(f"Relay#{voided_order['ID']}")

                            if voided_list:
                                relay_voided_orders_dict[(location, date)] = voided_list

                        # Process non-voided relay orders for tip comparison
                        relay_orders = relay_df[
                            (relay_df['Location'] == location) &
                            (relay_df['Date'] == date) &
                            (relay_df['Status'] != "VOIDED")
                        ]

                        if not relay_orders.empty:
                            # Get matching toast orders for the same location/date
                            toast_orders = toast_df[
                                (toast_df['Location'] == location) &
                                (toast_df['Date'] == date) &
                                (toast_df['Dining Options'] == 'Grubhub (Delivery)')
                            ]

                            for _, relay_order in relay_orders.iterrows():
                                consumer_name = str(relay_order['Consumer'])
                                if pd.notna(consumer_name) and consumer_name.strip():
                                    name_parts = consumer_name.strip().split()
                                    if len(name_parts) >= 2:
                                        first_name = name_parts[0].lower()
                                        last_initial = name_parts[1][0].lower()
                                        name_pattern = f"{re.escape(first_name)} {re.escape(last_initial)}.*grubhub"

                                        # Find matching toast order
                                        matching_toast_order = None
                                        toast_tip = None
                                        for _, toast_row in toast_orders.iterrows():
                                            tab_name = str(toast_row['Tab Names']).lower()
                                            if pd.notna(tab_name) and re.search(name_pattern, tab_name):
                                                matching_toast_order = toast_row['Order #']
                                                toast_tip = toast_row['Tip']
                                                break

                                        if matching_toast_order:
                                            # Compare tips between relay and toast
                                            relay_tip = relay_order['Tip']
                                            if abs(float(toast_tip) - float(relay_tip)) > 0.01:
                                                key = (location, date)
                                                if key not in relay_mismatch_dict:
                                                    relay_mismatch_dict[key] = []
                                                relay_mismatch_dict[key].append(
                                                    f"Relay#{relay_order['ID']} (Toast#{matching_toast_order}): "
                                                    f"Relay=${relay_tip:.2f}, Toast=${toast_tip:.2f}"
                                                )

                        # Add any incorrect dining option orders
                        if incorrect_orders:
                            if (location, date) in incorrect_orders_dict:
                                incorrect_orders_dict[(location, date)].extend([f"#{order}" for order in incorrect_orders])
                            else:
                                incorrect_orders_dict[(location, date)] = [f"#{order}" for order in incorrect_orders]

        def find_matching_tip_orders(row, toast_df):
            knock_diff = abs(row['Toast Knock Delivery Tips'] - row['Knock Payable Delivery Tips'])
            if knock_diff > 0.01:
                location_date_orders = toast_df[
                    (toast_df['Location'] == row['Location']) &
                    (toast_df['Date'] == row['Date']) &
                    (abs(toast_df['Tip'] - knock_diff) < 0.01)
                ]['Order #'].astype(str).tolist()

                if location_date_orders:
                    return ', '.join([f"#{order}" for order in location_date_orders])
            return ''

        # Start with OLO and knock tips merge
        discrepancy = pd.merge(olo_tips, knock_tips, on=['Location', 'Date'], how='outer').fillna(0)
        discrepancy = pd.merge(discrepancy, toast_relay_tips, on=['Location', 'Date'], how='outer').fillna(0)
        discrepancy = pd.merge(discrepancy, toast_metro_speedy_tips, on=['Location', 'Date'], how='outer').fillna(0)
        discrepancy = pd.merge(discrepancy, toast_new_company_tips, on=['Location', 'Date'], how='outer').fillna(0)

        # Handle New Company tips based on Relacion file availability
        new_company_tips = pd.DataFrame()

        for location in new_company_locations:
            if location in new_company_locations_with_data:
                location_data = new_company_data[new_company_data['Location'] == location].copy()
                location_data = location_data.rename(columns={'Tip': 'New Company Delivery Tips'})
            else:
                location_data = olo_payable[olo_payable['Location'] == location].copy()
                location_data = location_data.rename(columns={'Tip': 'New Company Delivery Tips'})

            if new_company_tips.empty:
                new_company_tips = location_data
            else:
                new_company_tips = pd.concat([new_company_tips, location_data])

        # Add New Company payable tips
        discrepancy = pd.merge(discrepancy, new_company_tips, on=['Location', 'Date'], how='outer').fillna(0)

        # Then add OLO payable tips
        olo_payable = olo_payable.rename(columns={'Tip': 'OLO Payable Delivery Tips'})
        discrepancy = pd.merge(discrepancy, olo_payable, on=['Location', 'Date'], how='outer').fillna(0)

        # Set OLO Payable Delivery Tips to 0 for new company locations
        discrepancy.loc[discrepancy['Location'].isin(new_company_locations), 'OLO Payable Delivery Tips'] = 0

        # Set Toast OLO Delivery Tips to 0 for new company locations
        discrepancy.loc[discrepancy['Location'].isin(new_company_locations), 'Toast OLO Delivery Tips'] = 0

        # Add knock payable tips
        knock_df = pd.DataFrame([
            {'Location': loc, 'Date': date, 'Knock Payable Delivery Tips': tip}
            for loc, date_tips in knock_payable.items()
            for date, tip in date_tips.items()
        ])
        if not knock_df.empty:
            discrepancy = pd.merge(discrepancy, knock_df, on=['Location', 'Date'], how='outer').fillna(0)
        else:
            discrepancy['Knock Payable Delivery Tips'] = 0

        # Add relay payable tips
        relay_payable = relay_payable.rename(columns={'Tip': 'Relay Payable Delivery Tips'})
        discrepancy = pd.merge(discrepancy, relay_payable, on=['Location', 'Date'], how='outer').fillna(0)

        # Add Metro Speedy payable tips
        if not metro_speedy_data.empty:
            metro_speedy_data = metro_speedy_data.rename(columns={'Tip': 'Metro Speedy Payable Delivery Tips'})
            discrepancy = pd.merge(discrepancy, metro_speedy_data, on=['Location', 'Date'], how='outer').fillna(0)
        else:
            discrepancy['Metro Speedy Payable Delivery Tips'] = 0

        # In the column setup section:
        discrepancy['Incorrect Dining Option'] = ''
        discrepancy['Relay Voided Orders'] = ''
        discrepancy['OLO Refunded/Cancelled'] = ''
        discrepancy['Relay Mismatch Tip'] = ''  # New column

        # When populating the columns:
        for idx, row in discrepancy.iterrows():
            key = (row['Location'], row['Date'])

            # Incorrect dining options
            if key in incorrect_orders_dict:
                discrepancy.at[idx, 'Incorrect Dining Option'] = ', '.join(incorrect_orders_dict[key])

            # Relay voided orders
            if key in relay_voided_orders_dict:
                discrepancy.at[idx, 'Relay Voided Orders'] = ', '.join(relay_voided_orders_dict[key])

            # OLO refunded/cancelled orders
            if key in refund_cancelled_orders_dict:
                discrepancy.at[idx, 'OLO Refunded/Cancelled'] = ', '.join(refund_cancelled_orders_dict[key])

            # Relay mismatches
            if key in relay_mismatch_dict:
                discrepancy.at[idx, 'Relay Mismatch Tip'] = ', '.join(relay_mismatch_dict[key])


            # Add orders from knock tip difference check if any
            matching_orders = find_matching_tip_orders(row, toast_df)
            if matching_orders:
                if discrepancy.at[idx, 'Incorrect Dining Option']:
                    discrepancy.at[idx, 'Incorrect Dining Option'] += f", {matching_orders}"
                else:
                    discrepancy.at[idx, 'Incorrect Dining Option'] = matching_orders

        # Add Total 3rd Party Delivery Tips
        discrepancy['Total 3rd Party Delivery Tips'] = discrepancy.apply(
            lambda row: (
                (0 if row['Location'] in new_company_locations else row['OLO Payable Delivery Tips']) +
                    row['Knock Payable Delivery Tips'] +
                    row['Relay Payable Delivery Tips'] +
                    row['Metro Speedy Payable Delivery Tips'] +
                    (row['New Company Delivery Tips'] if row['Location'] in new_company_locations else 0)
                ),
                axis=1
            ).round(2)

        # Calculate Toast delivery tips
        discrepancy['Total Toast Delivery Tips'] = discrepancy.apply(
            lambda row: (
            (0 if row['Location'] in new_company_locations else row['Toast OLO Delivery Tips']) +
                    row['Toast Knock Delivery Tips'] +
                    row['Toast Metro Speedy Delivery Tips'] +
                    row['Toast Relay Delivery Tips'] +
                    (row['Toast New Company Delivery Tips'] if row['Location'] in new_company_locations else 0)
                ),
                axis=1
            )

        # Calculate daily total tips and employee tips from Toast using the combined data
        daily_total_tips = toast_df.groupby(['Location', 'Date'])['Tips_Total'].sum().reset_index()
        daily_employee_tips = daily_total_tips.copy()
        daily_employee_tips = pd.merge(daily_employee_tips, discrepancy[['Location', 'Date', 'Total Toast Delivery Tips']],
                                    on=['Location', 'Date'], how='left')

        daily_employee_tips['Total Toast Employee Tips'] = (
            daily_employee_tips['Tips_Total'] -
            daily_employee_tips['Total Toast Delivery Tips']
        )

        daily_employee_tips = daily_employee_tips[['Location', 'Date', 'Total Toast Employee Tips']]

        # Add Total Toast Employee Tips to discrepancy DataFrame
        discrepancy = pd.merge(discrepancy, daily_employee_tips, on=['Location', 'Date'], how='left').fillna(0)

        # Get R365 tips directly from GL file
        gl_files = [f for f in self.input_files if os.path.basename(f).lower().startswith('gl')]
        if not gl_files:
            raise ValueError("GL file not found")
        gl_df = pd.read_csv(gl_files[0], skiprows=3, thousands=',')

        gl_df['LocationName1'] = gl_df['LocationName1'].map(self.gl_mapping)
        gl_df['TrxDate'] = pd.to_datetime(gl_df['TrxDate']).dt.strftime('%m/%d/%Y')

        # Get R365 delivery tips
        r365_del_tips = gl_df[gl_df['ParentAccountName'] == '21251 - Payable Delivery Tips'].groupby(['LocationName1', 'TrxDate'])['Credit'].sum().reset_index()
        r365_del_tips = r365_del_tips[~r365_del_tips['LocationName1'].isin(excluded_locations)]
        r365_del_tips.columns = ['Location', 'Date', 'R365 Delivery Tips Payable']

        # Get R365 employee tips
        r365_emp_tips = gl_df[gl_df['ParentAccountName'] == '21250 - Employee Tips Payable'].groupby(['LocationName1', 'TrxDate'])['Credit'].sum().reset_index()
        r365_emp_tips = r365_emp_tips[~r365_emp_tips['LocationName1'].isin(excluded_locations)]
        r365_emp_tips.columns = ['Location', 'Date', 'R365 Employee Tips Payable']

        # Add R365 tips to the dataframe
        discrepancy = pd.merge(discrepancy, r365_del_tips, on=['Location', 'Date'], how='left').fillna(0)
        discrepancy = pd.merge(discrepancy, r365_emp_tips, on=['Location', 'Date'], how='left').fillna(0)

        # Calculate differences
        discrepancy['Difference Toast - 3P Delivery Tips'] = (
            discrepancy['Total Toast Delivery Tips'] -
            discrepancy['Total 3rd Party Delivery Tips']
        )
        discrepancy['Difference Toast - R365 Delivery Tips'] = (
            discrepancy['Total Toast Delivery Tips'] -
            discrepancy['R365 Delivery Tips Payable']
        )
        discrepancy['Difference Toast - R365 Employee Tips'] = (
            discrepancy['Total Toast Employee Tips'] -
            discrepancy['R365 Employee Tips Payable']
        )

        # Filter for significant differences
        significant_diff = (
            (abs(discrepancy['Difference Toast - 3P Delivery Tips']) > 0.01) |
            (abs(discrepancy['Difference Toast - R365 Delivery Tips']) > 0.01) |
            (abs(discrepancy['Difference Toast - R365 Employee Tips']) > 0.01)
        )
        discrepancy = discrepancy[significant_diff]

        # Final deduplication
        discrepancy = discrepancy.groupby(['Location', 'Date']).first().reset_index()

        columns = [
            'Location', 'Date',
            'OLO Payable Delivery Tips', 'Knock Payable Delivery Tips',
            'Metro Speedy Payable Delivery Tips', 'Relay Payable Delivery Tips',
            'New Company Delivery Tips',
            'Toast OLO Delivery Tips', 'Toast Knock Delivery Tips',
            'Toast Metro Speedy Delivery Tips', 'Toast Relay Delivery Tips',
            'Toast New Company Delivery Tips',
            'R365 Delivery Tips Payable', 'R365 Employee Tips Payable',
            'Total 3rd Party Delivery Tips', 'Total Toast Delivery Tips',
            'Total Toast Employee Tips',
            'Difference Toast - R365 Delivery Tips', 'Difference Toast - R365 Employee Tips',
            'Difference Toast - 3P Delivery Tips',
            'Incorrect Dining Option', 'Relay Voided Orders', 'OLO Refunded/Cancelled', 'Relay Mismatch Tip'
        ]

        # Create a final column mapping to rename columns
        column_mapping = {
            'Incorrect Dining Option': 'Incorrect Dining Option/Tip Amount'
        }
        discrepancy = discrepancy[columns].rename(columns=column_mapping)


        # Round all numeric columns to 2 decimal places
        numeric_columns = discrepancy.select_dtypes(include=['float64', 'int64']).columns
        discrepancy[numeric_columns] = discrepancy[numeric_columns].round(2)

        # Zero out any tiny differences (less than a penny)
        discrepancy.loc[abs(discrepancy['Difference Toast - 3P Delivery Tips']) < 0.01, 'Difference Toast - 3P Delivery Tips'] = 0
        discrepancy.loc[abs(discrepancy['Difference Toast - R365 Delivery Tips']) < 0.01, 'Difference Toast - R365 Delivery Tips'] = 0
        discrepancy.loc[abs(discrepancy['Difference Toast - R365 Employee Tips']) < 0.01, 'Difference Toast - R365 Employee Tips'] = 0

        return discrepancy

    def format_summary_excel(self, summary_df, output_path, ws):
        import openpyxl
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter

        # Rename the difference columns (remove the word "Difference")
        renamed_columns = {
            'Difference Toast - 3P Delivery Tips': 'Toast - 3P Delivery Tips',
            'Difference Toast - R365 Delivery Tips': 'Toast - R365 Delivery Tips',
            'Difference Toast - R365 Employee Tips': 'Toast - R365 Employee Tips'
        }
        summary_df = summary_df.rename(columns=renamed_columns)


        # Insert new row at top for grouped headers
        ws.insert_rows(1)

        # Apply Aptos Narrow font to all cells
        for row in ws.rows:
            for cell in row:
                cell.font = Font(name='Aptos Narrow')

        # Set column width
        EXCEL_COLUMN_WIDTH = 19
        WIDE_COLUMN_WIDTH = 25  # approximately 170 pixels
        # Set row height for row 2
        ws.row_dimensions[2].height = 35

        # Set default width for all columns
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            # Special width for columns A and B
            if col <= 2:
                ws.column_dimensions[column_letter].width = WIDE_COLUMN_WIDTH
            else:
                ws.column_dimensions[column_letter].width = EXCEL_COLUMN_WIDTH

        # Define colors
        light_blue = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        light_orange = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        super_light_red = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        # Define sections and their headers
        sections = [
            ('A', 'B', None, None),  # Location, Date range
            ('C', 'D', 'Toast Totals', light_orange),
            ('E', 'F', 'R365 Totals', light_green),
            ('G', 'K', '3P Delivery Total', light_blue),
            ('L', 'N', 'DIFFERENCES', super_light_red ),
            ('O', 'P', 'PAYROLL DIFFERENCES', yellow)
        ]

        # Define center alignment
        center_align = Alignment(horizontal='center', vertical='center')
        thick = Side(style='thick')

        # Apply merged headers, colors, and borders for row 1
        for start_col, end_col, header_text, color in sections:
            if header_text:  # Skip the first section
                # Merge cells and add header text
                if start_col != end_col:
                    ws.merge_cells(f'{start_col}1:{end_col}1')
                cell = ws[f'{start_col}1']
                cell.value = header_text
                cell.alignment = center_align

                # Apply color if specified
                if color:
                    for col in range(ord(start_col) - 64, ord(end_col) - 64 + 1):
                        col_letter = get_column_letter(col)
                        ws[f'{col_letter}1'].fill = color

                # Apply borders
                if start_col == end_col:
                    ws[f'{start_col}1'].border = Border(top=thick, right=thick, bottom=thick, left=thick)
                else:
                    ws[f'{start_col}1'].border = Border(left=thick, top=thick, bottom=thick)
                    ws[f'{end_col}1'].border = Border(right=thick, top=thick, bottom=thick)
                    for col in range(ord(start_col) - 64 + 1, ord(end_col) - 64):
                        col_letter = get_column_letter(col)
                        ws[f'{col_letter}1'].border = Border(top=thick, bottom=thick)

        # Apply text wrapping and alignment to row 2
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}2']
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Color the specific cells in row 2
        ws['G2'].fill = light_blue   # Total 3rd Party Delivery Tips
        ws['L2'].fill = light_blue   # Toast - 3P Delivery Tips
        ws['M2'].fill = light_green # Toast - R365 Delivery Tips
        ws['N2'].fill = light_green  # Toast - R365 Employee Tips

        # Apply borders to the data section
        max_row = ws.max_row
        for start_col, end_col, _, _ in sections[1:]:  # Skip first section
            for row in range(2, max_row + 1):
                for col in range(ord(start_col) - 64, ord(end_col) - 64 + 1):
                    col_letter = get_column_letter(col)
                    cell = ws[f'{col_letter}{row}']
                    if row == 2:
                        # Skip borders for A2 and B2
                        if col_letter not in ['A', 'B']:
                            cell.border = Border(top=thick, bottom=thick,
                                                 left=thick if col_letter == start_col else None,
                                                 right=thick if col_letter == end_col else None)
                    else:
                        cell.border = Border(left=thick if col_letter == start_col else None,
                                             right=thick if col_letter == end_col else None,
                                             bottom=thick if row == max_row else None)
        # Remove borders from cells A2 and B2
        ws['A2'].border = Border()
        ws['B2'].border = Border()




    def format_discrepancy_excel(self, discrepancy_df, output_path, ws):
        import openpyxl
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter

        # Insert new row at top for grouped headers
        ws.insert_rows(1)

        # Apply Aptos Narrow font to all cells
        for row in ws.rows:
            for cell in row:
                cell.font = Font(name='Aptos Narrow')

        # Set exact column width of 115 pixels for all columns
        EXCEL_COLUMN_WIDTH = 19
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = EXCEL_COLUMN_WIDTH

        # Set row height for row 2
        ws.row_dimensions[2].height = 35

        # Apply center alignment to A2 and B2 (both horizontal and vertical)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

        # Remove all borders and fills first
        no_border = Border()
        no_fill = PatternFill(fill_type=None)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = no_border
                cell.fill = no_fill

        # Define colors
        light_blue = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        light_orange = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        super_light_red = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        # Define comparison highlight colors
        comp_light_blue = PatternFill(start_color="DCEBF7", end_color="DCEBF7", fill_type="solid")
        comp_light_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        comp_light_orange = PatternFill(start_color="FFE4CC", end_color="FFE4CC", fill_type="solid")
        comp_light_green = PatternFill(start_color="E8F3E8", end_color="E8F3E8", fill_type="solid")

        # Define sections and their headers
        sections = [
            ('C', 'G', '3rd Parties', light_blue),
            ('H', 'L', 'Toast', light_orange),
            ('M', 'N', 'R365', light_green),
            ('O', 'Q', 'TOTALS', super_light_red),
            ('R', 'T', 'DIFFERENCES', super_light_red),
            ('U', 'X', 'ORDER NUMBER', None)  # Updated to span four columns U-X
        ]

        # Define center alignment
        center_align = Alignment(horizontal='center', vertical='center')
        thick = Side(style='thick')
        thin = Side(style='thin')

        # Apply thick border around entire dataset
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                cell = ws[f'{col_letter}{row}']

                # Default thin borders
                border_left = thin
                border_right = thin
                border_top = thin
                border_bottom = thin

                # Thick borders for edges
                if col == 1:  # First column
                    border_left = thick
                if col == ws.max_column:  # Last column
                    border_right = thick
                if row == 1:  # First row
                    border_top = thick
                if row == ws.max_row:  # Last row
                    border_bottom = thick
                if row == 2:  # Second row
                    border_top = thick
                    border_bottom = thick

                cell.border = Border(
                    left=border_left,
                    right=border_right,
                    top=border_top,
                    bottom=border_bottom
                )

        # Apply merged headers, colors, and borders for row 1
        for start_col, end_col, header_text, color in sections:
            # Merge cells and add header text
            if start_col != end_col:
                ws.merge_cells(f'{start_col}1:{end_col}1')
            cell = ws[f'{start_col}1']
            cell.value = header_text
            cell.alignment = center_align

            # Apply color if specified
            if color:
                for col in range(ord(start_col) - 64, ord(end_col) - 64 + 1):
                    col_letter = get_column_letter(col)
                    ws[f'{col_letter}1'].fill = color

        # Apply thick borders between sections and within ORDER NUMBER section
        for start_col, end_col, _, _ in sections:
            start_num = ord(start_col) - 64
            if start_col != 'A':  # Don't add left border to first column
                for row in range(1, ws.max_row + 1):
                    # Add thick left border to section start
                    cell = ws[f'{start_col}{row}']
                    current_border = cell.border
                    cell.border = Border(
                        left=thick,
                        right=current_border.right,
                        top=current_border.top,
                        bottom=current_border.bottom
                    )

                    # Add thin vertical borders between ORDER NUMBER columns
                    if start_col == 'U':  # ORDER NUMBER section
                        # Add thin border between U and V
                        cell_v = ws[f'V{row}']
                        current_v_border = cell_v.border
                        cell_v.border = Border(
                            left=thin,
                            right=current_v_border.right,
                            top=current_v_border.top,
                            bottom=current_v_border.bottom
                        )

                        # Add thin border between V and W
                        cell_w = ws[f'W{row}']
                        current_w_border = cell_w.border
                        cell_w.border = Border(
                            left=thin,
                            right=thin,  # Change this to thin since it's not the last column anymore
                            top=current_w_border.top,
                            bottom=current_w_border.bottom
                        )

                        # Add border for column X (the actual last column)
                        cell_x = ws[f'X{row}']
                        current_x_border = cell_x.border
                        cell_x.border = Border(
                            left=thin,
                            right=thick,  # This will create the thick right border
                            top=current_x_border.top,
                            bottom=current_x_border.bottom
                        )

        # Apply text wrapping and alignment to row 2
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}2']
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Apply colors to specific cells in row 2
        ws['O2'].fill = light_blue  # Total 3rd Party Delivery Tips
        ws['P2'].fill = light_blue  # Total Toast Delivery Tips
        ws['Q2'].fill = light_orange  # Total Toast Employee Tips
        ws['R2'].fill = light_green  # Toast - R365 Delivery Tips
        ws['S2'].fill = light_green  # Toast - R365 Employee Tips
        ws['T2'].fill = light_blue  # Toast - 3P Delivery Tips

        # Define the pairs to compare and their corresponding highlight colors
        pair_configs = [
            ('Toast OLO Delivery Tips', 'OLO Payable Delivery Tips', comp_light_blue),
            ('Toast Knock Delivery Tips', 'Knock Payable Delivery Tips', comp_light_yellow),
            ('Toast Metro Speedy Delivery Tips', 'Metro Speedy Payable Delivery Tips', comp_light_orange),
            ('Toast Relay Delivery Tips', 'Relay Payable Delivery Tips', comp_light_green),
            ('Toast New Company Delivery Tips', 'New Company Delivery Tips', light_blue)
        ]

        # Apply highlighting for each row
        for row in range(3, ws.max_row + 1):
            for toast_col, payable_col, highlight_color in pair_configs:
                if toast_col in discrepancy_df.columns and payable_col in discrepancy_df.columns:
                    col_letters = {}
                    for col in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        header_cell = ws[f"{col_letter}2"]
                        if header_cell.value:
                            col_letters[header_cell.value] = col_letter

                    if toast_col in col_letters and payable_col in col_letters:
                        toast_cell = ws[f"{col_letters[toast_col]}{row}"]
                        payable_cell = ws[f"{col_letters[payable_col]}{row}"]

                        try:
                            toast_value = round(float(toast_cell.value or 0), 2)
                            payable_value = round(float(payable_cell.value or 0), 2)

                            if abs(toast_value - payable_value) > 0.01:
                                toast_cell.fill = highlight_color
                                payable_cell.fill = highlight_color
                        except (ValueError, TypeError):
                            continue

        return ws



from typing import Dict, Tuple


class AIErrorHandler:
    def __init__(self):
        """Initialize the error handler with common error patterns and their user-friendly messages.
        The patterns are regular expressions that match common errors, paired with
        user-friendly messages explaining what went wrong and how to fix it."""

        self.file_type_requirements = {
            r"order.*": ".csv",
            r"itemized_orders.*": ".csv",
            r"transaction.*": ".csv",
            r"itemized_cancelled.*": ".csv",
            r"gl.*": ".csv",
            r"billing.*": ".xlsx",
            r"relay_carrotexpress.*": ".xlsx",
            r"relacion.*": ".xlsx",
            r"payroll.*": ".csv"
        }

        self.error_patterns: Dict[str, Tuple[str, str]] = {
            # File format errors
            r"UnicodeDecodeError: 'utf-8' codec can't decode bytes.*File \".*?([^\\\/]+\.(?:csv|xlsx))\",": (
                "File Encoding Error",
                "The file '{0}' was saved in the wrong format. "
                "This usually happens when you save an Excel file (.xlsx) as a regular CSV instead of UTF-8 CSV.\n\n"
                "To fix this:\n"
                "1. Open '{0}' in Excel\n"
                "2. Click 'Save As'\n"
                "3. Choose 'CSV UTF-8' as the file type\n"
                "4. Save and try again"
            ),
            r"Error tokenizing data|Expected \d+ fields in line \d+, saw \d+": (
                "CSV Format Error",
                "There seems to be an issue with the structure of your CSV file. This usually happens when:\n"
                "1. The file was saved in the wrong format\n"
                "2. The file contains merged cells\n"
                "3. There are extra commas in the data\n\n"
                "Please open the file in Excel, ensure no cells are merged, and save it as a CSV file."
            ),
            r"Unsupported format, or corrupt file: Expected BOF record; found": (
                "Excel Format Error",
                "The file you're trying to open as an Excel file (.xlsx) appears to be in a different format. "
                "This usually happens when a file is saved as .csv but renamed to .xlsx.\n\n"
                "To fix this:\n"
                "1. Open the file in Excel\n"
                "2. Click 'Save As'\n"
                "3. Select 'Excel Workbook (.xlsx)' as the file type\n"
                "4. Save and try again"
            ),
            r"File is not a zip file": (
                "Excel File Error",
                "An Excel file appears to be corrupted or in the wrong format. "
                "This often happens when a CSV file is simply renamed to .xlsx instead of being properly saved as Excel.\n\n"
                "To fix this:\n"
                "1. Open the original file in Excel\n"
                "2. Use 'Save As' and select 'Excel Workbook (.xlsx)'\n"
                "3. Try again with the new file"
            ),

            # General file errors
            r"No such file or directory: '(.*)'": (
                "File Not Found",
                "Unable to find the file: {0}\n"
                "Please check that the file exists and you have selected it."
            ),
            r"Permission denied: '(.*)'": (
                "File Access Error",
                "Cannot access the file: {0}\n"
                "This usually happens when:\n"
                "1. The file is open in another program (like Excel)\n"
                "2. You don't have permission to access the file\n\n"
                "Please close any programs that might be using the file and try again."
            ),

            # Data format errors
            r"Invalid date format detected in '(.*?)'": (
                "Date Format Error",
                "Incorrect date format found in file: {0}\n"
                "Please ensure all dates are in MM/DD/YYYY format."
            ),
            r"could not convert string '(.*)' to float": (
                "Number Format Error",
                "Found invalid number format: '{0}'\n"
                "This usually happens when:\n"
                "1. Numbers are formatted as text\n"
                "2. There are special characters in number fields\n"
                "3. Currency symbols or commas are present\n\n"
                "Please check your files for any non-numeric characters in number columns."
            ),

            # Missing data errors
            r"KeyError: '(.*)'": (
                "Missing Column",
                "Required column '{0}' is missing.\n"
                "Please check that all required columns are present and spelled correctly."
            ),
            r"Required column '(.*)' not found in file '(.*)'": (
                "Missing Required Column",
                "The column '{0}' is missing from file '{1}'.\n"
                "Please verify the file has all required columns."
            ),

            # Default patterns from before
            r"Toast file not found": (
                "Missing Toast File Error",
                "The required Toast file (starting with 'Order') was not found. Please make sure you've selected a Toast file that starts with 'Order' in its name."
            ),
            r"Required OLO files \(Itemized or Transaction\) not found": (
                "Missing OLO Files Error",
                "The required OLO files are missing. Please make sure you've selected both:\n1. An Itemized Orders file (starts with 'Itemized_Orders')\n2. A Transaction file (starts with 'Transaction')"
            ),
            r"GL file not found": (
                "Missing GL File Error",
                "The GL file was not found. Please make sure you've selected a GL file that starts with 'GL' in its name."
            )
        }

    def check_file_format(self, filename: str) -> Tuple[bool, str]:
        """
        Checks if a file is in the correct format based on its name.

        Args:
            filename: Name of the file to check

        Returns:
            Tuple of (is_correct_format, error_message)
        """
        base_name = os.path.basename(filename).lower()
        actual_ext = os.path.splitext(base_name)[1].lower()

        for pattern, required_ext in self.file_type_requirements.items():
            if re.match(pattern, base_name):
                if actual_ext != required_ext:
                    return False, (
                        f"File format error: '{os.path.basename(filename)}' should be a {required_ext} file.\n\n"
                        f"To fix this:\n"
                        f"1. Open the file in Excel\n"
                        f"2. Click 'Save As'\n"
                        f"3. Select '{required_ext[1:].upper()} format' as the file type\n"
                        f"4. Save and try again"
                    )
        return True, ""

    def interpret_error(self, error_message: str, filename: str = None) -> Tuple[str, str]:
        """
        Interprets an error message and returns a user-friendly explanation.
        Prioritizes specific error messages over generic guidance.
        """
        # Check for specific error patterns first

        # 1. Missing columns error
        missing_columns_match = re.search(r"Missing required columns in Toast file '([^']+)': (.*?)(?:\n|$)", error_message)
        if missing_columns_match:
            filename = missing_columns_match.group(1)
            missing_cols = missing_columns_match.group(2)
            return "Missing Required Column", (
                f"The file '{filename}' is missing the following required column: {missing_cols}\n\n"
                f"To fix this:\n"
                f"1. Go to Toast website\n"
                f"2. When downloading the Order Details report, make sure to include all required columns\n"
                f"3. Required columns are: Order #, Location, Opened, Tab Names, Server, Dining Options, Tip, Gratuity"
            )

        # 2. File format error
        if filename:
            is_correct_format, format_error = self.check_file_format(filename)
            if not is_correct_format:
                return "File Format Error", format_error

        # 3. File encoding error
        encoding_match = re.search(r"UnicodeDecodeError: 'utf-8' codec can't decode bytes.*File \".*?([^\\\/]+\.(?:csv|xlsx))\"", error_message)
        if encoding_match:
            file = encoding_match.group(1)
            return "File Encoding Error", (
                f"The file '{file}' needs to be saved in UTF-8 format.\n\n"
                f"To fix this:\n"
                f"1. Open '{file}' in Excel\n"
                f"2. Click 'Save As'\n"
                f"3. Choose 'CSV UTF-8' as the file type\n"
                f"4. Save and try again"
            )

        # 4. Missing required files
        if "Toast file not found" in error_message:
            return "Missing Toast File", (
                "No Toast files were found in your selection.\n\n"
                "To fix this:\n"
                "1. Make sure to select your Toast files (starting with 'Order')\n"
                "2. Try running the program again with the correct files selected"
            )

        if "Required OLO files (Itemized or Transaction) not found" in error_message:
            return "Missing OLO Files", (
                "Required OLO files are missing from your selection.\n\n"
                "To fix this:\n"
                "1. Make sure to select all three required files:\n"
                "   - Itemized Orders file (starts with 'Itemized_Orders')\n"
                "   - Itemized Cancelled Orders file (starts with 'Itemized_Cancelled')\n"
                "   - Transaction file (starts with 'Transaction')\n"
                "2. Try running the program again with all required files selected"
            )

        if "GL file not found" in error_message:
            return "Missing GL File", (
                "No GL file was found in your selection.\n\n"
                "To fix this:\n"
                "1. Make sure to select your GL file (starts with 'GL')\n"
                "2. Try running the program again with the GL file selected"
            )

        # 5. Date format errors
        date_format_match = re.search(r"time data \"([^\"]+)\" doesn't match format", error_message)
        if date_format_match:
            bad_date = date_format_match.group(1)
            return "Date Format Error", (
                f"Found an unexpected date format: '{bad_date}'\n\n"
                f"To fix this:\n"
                f"1. Check that your dates are in the correct format (MM/DD/YYYY)\n"
                f"2. Make sure you're using the correct export settings when downloading your files"
            )

        # 6. CSV structure errors
        if "Error tokenizing data" in error_message or "Expected" in error_message and "fields in line" in error_message:
            affected_file = filename if filename else "CSV file"
            return "CSV Format Error", (
                f"There's an issue with the structure of {affected_file}.\n\n"
                f"To fix this:\n"
                f"1. Open the file in Excel\n"
                f"2. Check for merged cells or extra commas\n"
                f"3. Save as a regular CSV file and try again"
            )

        # 7. Access errors
        if "Permission denied" in error_message:
            affected_file = filename if filename else "file"
            return "File Access Error", (
                f"Cannot access {affected_file} - it might be open in another program.\n\n"
                f"To fix this:\n"
                f"1. Close any programs that might be using the file\n"
                f"2. Try running the program again"
            )

        # If no specific error pattern matches, then and only then show the generic message
        default_message = (
            "An unexpected error occurred. Here's what you can try:\n\n"
            "1. Check that all files are in the correct format:\n"
            "   - Toast files (Order*.csv)\n"
            "   - OLO files (Itemized*.csv, Transaction*.csv)\n"
            "   - Knock files (Billing*.xlsx)\n"
            "   - Relay files (relay_carrotexpress*.xlsx)\n"
            "   - GL files (GL*.csv)\n"
            "2. Ensure no files are open in other programs\n"
            "3. Try saving the files again in the correct format\n\n"
            f"Technical details:\n{error_message}"
        )

        if filename:
            default_message = f"Error in file: {os.path.basename(filename)}\n\n{default_message}"

        return "Unexpected Error", default_message

    def format_error_for_qmessage(self, error_message: str, filename: str = None) -> Tuple[str, str]:
        """
        Formats the error message specifically for QMessageBox display.

        Args:
            error_message: The original error message
            filename: Optional filename where the error occurred

        Returns:
            Tuple containing (title, formatted_message)
        """
        title, message = self.interpret_error(error_message, filename)
        return title, message


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = TipsReconcileWindow()
    ex.show()
    sys.exit(app.exec_())
